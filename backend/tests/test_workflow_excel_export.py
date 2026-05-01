"""
Tests for the workflow Excel export — verifying the output matches
mom's FA_Summary_paroll.xlsx format exactly.

Two test layers:
  1. Unit tests against _build_mom_excel() directly — no DB required, fast.
  2. HTTP-level smoke tests via TestClient — verifies routing + 404 behavior.

Checked properties:
  - Sheet name: "Payroll_Summary" (underscore)
  - Row 1: "MM/DD/YYYY - MM/DD/YYYY - Week N" format, no merge
  - Row 2: exact 10 headers in correct order
  - Header fill: #2563EB (blue), bold, white text, all columns centered
  - Data rows: currency format ("$"#,##0.00) on money columns E/F/G/I/J
  - TOTALS row: fill #A24B10 (orange-brown), bold, white text
  - Column widths: match mom's file within ±0.5 chars

Run:
    PYTHONPATH=<project-root> pytest backend/tests/test_workflow_excel_export.py -v
"""

from __future__ import annotations

import io
import os
import sys
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import BigInteger, Integer, Text, create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

# ── Project root on sys.path ─────────────────────────────────────────────────
_PROJECT_ROOT = str(Path(__file__).resolve().parents[2])
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

os.environ.setdefault(
    "ZPAY_SECRET_KEY",
    "test-secret-key-for-workflow-excel-export-tests-long",
)
os.environ.setdefault("DATABASE_URL", "sqlite://")

from backend.db.models import Base, ZRateOverride  # noqa: E402

# ── Metadata patches (same as other test modules) ────────────────────────────
Base.metadata.tables["z_rate_override"].c["effective_during"].type = Text()

for _tbl in Base.metadata.tables.values():
    for _col in _tbl.columns:
        if _col.primary_key and isinstance(_col.type, BigInteger):
            _col.type = Integer()

for _tbl in Base.metadata.tables.values():
    for _col in _tbl.columns:
        if _col.server_default is not None:
            _sd = _col.server_default
            try:
                _arg = _sd.arg.text if hasattr(_sd, "arg") and hasattr(_sd.arg, "text") else ""
            except Exception:
                _arg = ""
            if "NOW()" in _arg:
                _col.nullable = True
                _col.server_default = None

import re as _re  # noqa: E402

_engine = create_engine(
    "sqlite://",
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)


def _sqlite_regexp_replace(value: str, pattern: str, repl: str, flags: str) -> str:
    if value is None:
        return ""
    if "g" in (flags or ""):
        return _re.sub(pattern or "", repl or "", str(value))
    return _re.sub(pattern or "", repl or "", str(value), count=1)


@event.listens_for(_engine, "connect")
def _register_udfs(dbapi_conn, rec):
    dbapi_conn.create_function("regexp_replace", 4, _sqlite_regexp_replace)


Base.metadata.create_all(_engine)
_SessionFactory = sessionmaker(bind=_engine, autoflush=False, autocommit=False)

from fastapi.testclient import TestClient  # noqa: E402

from backend.app import app  # noqa: E402
from backend.db import get_db  # noqa: E402
from backend.db.models import PayrollBatch, Person, Ride  # noqa: E402
from backend.middleware.auth import COOKIE_NAME, create_session  # noqa: E402
from backend.routes.workflow import _build_mom_excel  # noqa: E402


def _override_get_db():
    db = _SessionFactory()
    try:
        yield db
    finally:
        db.close()


# NOTE: do NOT set app.dependency_overrides here at module level — doing so
# would stomp the override set by test_manual_adjustments.py when pytest
# collects both modules in the same process (they share one in-memory SQLite
# via StaticPool).  Instead, each test class that needs the DB override
# installs/removes it in setup_method/teardown_method.
def _install_db_override():
    app.dependency_overrides[get_db] = _override_get_db


def _remove_db_override():
    app.dependency_overrides.pop(get_db, None)

_SESSION_COOKIE = create_session(
    username="testadmin",
    display_name="Test Admin",
    color="#333",
    initials="TA",
    role="admin",
)
_AUTH = {COOKIE_NAME: _SESSION_COOKIE}
client = TestClient(app, raise_server_exceptions=True)

_NOW = datetime.now(timezone.utc)

# ── Reference values from mom's FA_Summary_paroll.xlsx ───────────────────────

MOM_HEADERS = [
    "Driver Name", "Pay Code", "Rides", "Miles",
    "Partner Pays", "Driver Pay", "Deduction",
    "Withheld (Y/N)", "Carried Over", "Paid This Period",
]
# Exact column widths (tolerance ±0.5)
MOM_COL_WIDTHS = [36.0, 12.0, 12.16, 13.16, 13.16, 10.16, 9.33, 8.66, 10.83, 13.33]
# Money columns (1-based): E=5 F=6 G=7 I=9 J=10
MONEY_COLS = {5, 6, 7, 9, 10}
# Colors with openpyxl alpha prefix "FF"
HEADER_FILL = "FF2563EB"
TOTALS_FILL = "FFA24B10"

# ── Minimal fixture data ──────────────────────────────────────────────────────

_SAMPLE_ROWS = [
    {
        "person": "Abbas Driver",
        "code": "1031",
        "rides": 7,
        "miles": 131.0,
        "partner_pays": 414.25,
        "driver_pay": 332.0,
        "deduction": 0.0,
        "withheld": False,
        "from_last_period": 0.0,
        "pay_this_period": 332.0,
    },
    {
        "person": "Zara Driver",
        "code": "1099",
        "rides": 2,
        "miles": 22.0,
        "partner_pays": 76.0,
        "driver_pay": 76.0,
        "deduction": 0.0,
        "withheld": True,
        "from_last_period": 0.0,
        "pay_this_period": 0.0,
    },
]
_SAMPLE_TOTALS = {
    "rides": 9,
    "miles": 153.0,
    "partner_pays": 490.25,
    "driver_pay": 408.0,
    "deduction": 0.0,
    "carried_over": 76.0,
    "pay_this_period": 332.0,
}
_SAMPLE_PERIOD = "03/21/2026 - 03/27/2026 - Week 12"


def _make_wb(rows=None, totals=None, period=None) -> openpyxl.Workbook:
    """Call _build_mom_excel and return the populated workbook."""
    wb = openpyxl.Workbook()
    _build_mom_excel(
        wb,
        rows   if rows   is not None else _SAMPLE_ROWS,
        totals if totals is not None else _SAMPLE_TOTALS,
        period if period is not None else _SAMPLE_PERIOD,
    )
    return wb


def _ws(wb: openpyxl.Workbook):
    return wb["Payroll_Summary"]


# ── HTTP-level helpers ────────────────────────────────────────────────────────

def _db():
    return _SessionFactory()


def _wipe():
    sess = _db()
    try:
        sess.query(Ride).delete(synchronize_session=False)
        sess.query(PayrollBatch).delete(synchronize_session=False)
        sess.query(Person).delete(synchronize_session=False)
        sess.commit()
    finally:
        sess.close()


def _seed_batch(sess, batch_id=1) -> PayrollBatch:
    b = PayrollBatch(
        payroll_batch_id=batch_id,
        source="acumen",
        company_name="FirstAlt",
        batch_ref=f"W-test-{batch_id}",
        status="approved",
        period_start=date(2026, 3, 21),
        period_end=date(2026, 3, 27),
        week_start=date(2026, 3, 21),
        week_end=date(2026, 3, 27),
        currency="USD",
        uploaded_at=_NOW,
    )
    sess.add(b)
    sess.flush()
    return b


# ── Unit tests: _build_mom_excel ─────────────────────────────────────────────

class TestBuildMomExcel:
    """Unit tests that call _build_mom_excel() directly — no DB, fast."""

    def test_sheet_name_uses_underscore(self):
        wb = _make_wb()
        assert "Payroll_Summary" in wb.sheetnames, (
            f"Expected 'Payroll_Summary', got: {wb.sheetnames}"
        )

    def test_row1_matches_period_label(self):
        wb = _make_wb()
        ws = _ws(wb)
        assert ws.cell(row=1, column=1).value == _SAMPLE_PERIOD

    def test_row1_height_is_22(self):
        wb = _make_wb()
        ws = _ws(wb)
        assert ws.row_dimensions[1].height == 22.0

    def test_row1_not_merged(self):
        wb = _make_wb()
        ws = _ws(wb)
        assert len(list(ws.merged_cells.ranges)) == 0, "Row 1 must not be merged"

    def test_header_row_is_row2(self):
        wb = _make_wb()
        ws = _ws(wb)
        row2 = [ws.cell(row=2, column=c).value for c in range(1, 11)]
        assert row2 == MOM_HEADERS, (
            f"Row 2 mismatch.\nExpected: {MOM_HEADERS}\nGot:      {row2}"
        )

    def test_header_column_count_is_10(self):
        wb = _make_wb()
        ws = _ws(wb)
        row2 = [ws.cell(row=2, column=c).value for c in range(1, 11)]
        assert len([v for v in row2 if v]) == 10

    def test_header_fill_is_blue(self):
        wb = _make_wb()
        ws = _ws(wb)
        for col in range(1, 11):
            fg = ws.cell(row=2, column=col).fill.fgColor.rgb
            assert fg == HEADER_FILL, f"Col {col} header fill: {fg!r}"

    def test_header_font_bold_white(self):
        wb = _make_wb()
        ws = _ws(wb)
        for col in range(1, 11):
            cell = ws.cell(row=2, column=col)
            assert cell.font.bold, f"Col {col} header not bold"
            assert cell.font.color.rgb in ("FFFFFFFF", "FFFFFF"), (
                f"Col {col} header font color: {cell.font.color.rgb!r}"
            )

    def test_header_alignment_centered(self):
        wb = _make_wb()
        ws = _ws(wb)
        for col in range(1, 11):
            cell = ws.cell(row=2, column=col)
            assert cell.alignment.horizontal == "center", (
                f"Col {col} header alignment: {cell.alignment.horizontal!r}"
            )

    def test_data_starts_at_row3(self):
        wb = _make_wb()
        ws = _ws(wb)
        val = ws.cell(row=3, column=1).value
        assert val not in (None, "", "TOTALS"), f"Row 3 A unexpected: {val!r}"

    def test_money_columns_currency_format(self):
        """Columns E/F/G/I/J must have a dollar-sign number format."""
        wb = _make_wb()
        ws = _ws(wb)
        for col in MONEY_COLS:
            fmt = ws.cell(row=3, column=col).number_format
            assert "$" in (fmt or ""), (
                f"Col {col} data row money format missing $: {fmt!r}"
            )

    def test_totals_row_label(self):
        wb = _make_wb()
        ws = _ws(wb)
        totals_row = next(
            (r[0].row for r in ws.iter_rows() if r[0].value == "TOTALS"), None
        )
        assert totals_row is not None, "TOTALS row not found"

    def test_totals_fill_is_orange_brown(self):
        wb = _make_wb()
        ws = _ws(wb)
        totals_row = next(r[0].row for r in ws.iter_rows() if r[0].value == "TOTALS")
        for col in range(1, 11):
            fg = ws.cell(row=totals_row, column=col).fill.fgColor.rgb
            assert fg == TOTALS_FILL, f"TOTALS col {col} fill: {fg!r}"

    def test_totals_font_bold_white(self):
        wb = _make_wb()
        ws = _ws(wb)
        totals_row = next(r[0].row for r in ws.iter_rows() if r[0].value == "TOTALS")
        cell = ws.cell(row=totals_row, column=1)
        assert cell.font.bold
        assert cell.font.color.rgb in ("FFFFFFFF", "FFFFFF")

    def test_column_widths_match_mom(self):
        wb = _make_wb()
        ws = _ws(wb)
        for idx, expected in enumerate(MOM_COL_WIDTHS, start=1):
            letter = openpyxl.utils.get_column_letter(idx)
            actual = ws.column_dimensions[letter].width
            assert abs(actual - expected) <= 0.5, (
                f"Col {letter} width: expected {expected}, got {actual}"
            )

    def test_withheld_driver_shows_yes(self):
        wb = _make_wb()
        ws = _ws(wb)
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] and "Zara" in str(row[0]):
                assert row[7] == "Yes", f"Withheld col H: {row[7]!r}"
                return
        pytest.fail("Withheld driver 'Zara Driver' not found")

    def test_paid_driver_shows_no(self):
        wb = _make_wb()
        ws = _ws(wb)
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] and "Abbas" in str(row[0]):
                assert row[7] == "No", f"Paid col H: {row[7]!r}"
                return
        pytest.fail("Paid driver 'Abbas Driver' not found")

    def test_totals_money_columns_have_currency_format(self):
        """Currency format must also appear on the TOTALS row."""
        wb = _make_wb()
        ws = _ws(wb)
        totals_row = next(r[0].row for r in ws.iter_rows() if r[0].value == "TOTALS")
        for col in MONEY_COLS:
            fmt = ws.cell(row=totals_row, column=col).number_format
            assert "$" in (fmt or ""), (
                f"TOTALS col {col} money format missing $: {fmt!r}"
            )

    def test_period_label_date_format(self):
        """Period label must follow MM/DD/YYYY - MM/DD/YYYY - Week N."""
        import re
        wb = _make_wb(period="03/21/2026 - 03/27/2026 - Week 12")
        ws = _ws(wb)
        val = ws.cell(row=1, column=1).value
        assert re.match(r"\d{2}/\d{2}/\d{4} - \d{2}/\d{2}/\d{4} - Week \d+", val), (
            f"Period label format mismatch: {val!r}"
        )

    def test_roundtrip_serialise_and_reopen(self):
        """Workbook must survive a save/reload cycle without corruption."""
        wb = _make_wb()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        wb2 = openpyxl.load_workbook(buf)
        assert "Payroll_Summary" in wb2.sheetnames
        ws2 = wb2["Payroll_Summary"]
        assert ws2.cell(row=2, column=1).value == "Driver Name"


# ── HTTP smoke tests ──────────────────────────────────────────────────────────
#
# _build_summary uses `func.coalesce(cast(Ride.ride_start_ts, Date))` — a
# single-argument coalesce that SQLite doesn't support.  These tests patch
# _build_summary out so the HTTP layer (routing, auth, content-type, 404) is
# verified without triggering the PostgreSQL-only query.

from unittest.mock import patch as _patch  # noqa: E402


_STUB_SUMMARY = {
    "rows": _SAMPLE_ROWS,
    "totals": _SAMPLE_TOTALS,
}


class TestExportEndpoint:
    """Light HTTP-level checks — routing, auth, 404."""

    def setup_method(self):
        _install_db_override()
        _wipe()

    def teardown_method(self):
        _remove_db_override()

    def test_404_for_unknown_batch(self):
        resp = client.get("/api/data/workflow/9999/export-excel", cookies=_AUTH)
        assert resp.status_code == 404

    def test_200_and_xlsx_content_type_for_known_batch(self):
        """Seeding a batch with no rides must still return 200 xlsx."""
        sess = _db()
        try:
            _seed_batch(sess, batch_id=42)
            sess.commit()
        finally:
            sess.close()
        with _patch("backend.routes.workflow._build_summary", return_value=_STUB_SUMMARY):
            resp = client.get("/api/data/workflow/42/export-excel", cookies=_AUTH)
        assert resp.status_code == 200
        ct = resp.headers.get("content-type", "")
        assert "spreadsheetml" in ct, f"Unexpected content-type: {ct}"

    def test_response_is_valid_xlsx(self):
        """The response body must be a parseable openpyxl workbook.

        Acumen batches now produce 3 tabs: SP PAY SUMMARY, SP ITEMIZED REPORT,
        Payroll Summary (space, not underscore — changed in feat/payroll-excel-mom-format-v2).
        """
        sess = _db()
        try:
            _seed_batch(sess, batch_id=43)
            sess.commit()
        finally:
            sess.close()
        with _patch("backend.routes.workflow._build_summary", return_value=_STUB_SUMMARY):
            resp = client.get("/api/data/workflow/43/export-excel", cookies=_AUTH)
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        # Acumen export now has 3 tabs; Payroll Summary uses a space (not underscore)
        assert "Payroll Summary" in wb.sheetnames or "Payroll_Summary" in wb.sheetnames

    def test_content_disposition_has_filename(self):
        """Content-Disposition header must include a filename."""
        sess = _db()
        try:
            _seed_batch(sess, batch_id=44)
            sess.commit()
        finally:
            sess.close()
        with _patch("backend.routes.workflow._build_summary", return_value=_STUB_SUMMARY):
            resp = client.get("/api/data/workflow/44/export-excel", cookies=_AUTH)
        assert resp.status_code == 200
        cd = resp.headers.get("content-disposition", "")
        assert "filename=" in cd, f"No filename in Content-Disposition: {cd}"

    def test_paychex_exported_at_stamped_on_download(self):
        """paychex_exported_at must be set on the batch after a successful download."""
        sess = _db()
        try:
            b = _seed_batch(sess, batch_id=50)
            assert b.paychex_exported_at is None, "Precondition: not yet stamped"
            sess.commit()
        finally:
            sess.close()

        with _patch("backend.routes.workflow._build_summary", return_value=_STUB_SUMMARY):
            resp = client.get("/api/data/workflow/50/export-excel", cookies=_AUTH)
        assert resp.status_code == 200

        # Reload the batch from DB and verify the timestamp was written
        sess = _db()
        try:
            from backend.db.models import PayrollBatch as _PB
            batch = sess.query(_PB).filter(_PB.payroll_batch_id == 50).first()
            assert batch is not None
            assert batch.paychex_exported_at is not None, (
                "paychex_exported_at was not stamped after export-excel download"
            )
        finally:
            sess.close()

    def test_paychex_exported_at_not_overwritten_on_second_download(self):
        """A second download must not overwrite the original stamp time."""
        sess = _db()
        try:
            b = _seed_batch(sess, batch_id=51)
            sess.commit()
        finally:
            sess.close()

        with _patch("backend.routes.workflow._build_summary", return_value=_STUB_SUMMARY):
            client.get("/api/data/workflow/51/export-excel", cookies=_AUTH)

        # Record the timestamp after first download
        sess = _db()
        try:
            from backend.db.models import PayrollBatch as _PB
            first_stamp = sess.query(_PB).filter(_PB.payroll_batch_id == 51).first().paychex_exported_at
        finally:
            sess.close()

        with _patch("backend.routes.workflow._build_summary", return_value=_STUB_SUMMARY):
            client.get("/api/data/workflow/51/export-excel", cookies=_AUTH)

        sess = _db()
        try:
            from backend.db.models import PayrollBatch as _PB
            second_stamp = sess.query(_PB).filter(_PB.payroll_batch_id == 51).first().paychex_exported_at
        finally:
            sess.close()

        assert first_stamp == second_stamp, (
            f"paychex_exported_at was overwritten on second download: {first_stamp} → {second_stamp}"
        )

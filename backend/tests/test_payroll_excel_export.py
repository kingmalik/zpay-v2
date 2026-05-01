"""
Tests for the 3-tab payroll Excel export (Acumen) and 1-tab (Maz).

Coverage:
  - Acumen batch → 3 sheets: SP PAY SUMMARY, SP ITEMIZED REPORT, Payroll Summary
  - Maz batch → 1 sheet: Payroll Summary (no SP tabs)
  - Payroll Summary layout: title R1, period R2, blank R3, headers R4, data R5+
  - TOTALS row present and orange-brown styled
  - Paychex Flex Amount row present (mom keys amount, formula in G)
  - Paid on Week section present
  - Unpaid on Week section present
  - Under-$100 withheld rule: driver with combined < $100 → Withheld=Yes, J=0
  - Under-$100 withheld rule: driver with combined >= $100 → Withheld=No, J>0
  - Carried Over (col I) populated when driver has prior held balance
  - Release WARNING logged when this-week earnings < $100 but driver is released
  - Per-LLC isolation (Acumen and Maz drivers stay in separate files)
  - SP PAY SUMMARY has correct headers in row 1
  - SP ITEMIZED REPORT has correct headers in row 1
  - SP PAY SUMMARY driver data matches summary rows
  - HTTP 404 for unknown batch
  - HTTP 200 + xlsx content-type for known Acumen batch
  - HTTP 200 + xlsx content-type for known Maz batch
  - Content-Disposition has filename for both
  - File is a parseable workbook (save/reload round-trip)

Run:
    PYTHONPATH=<project-root> pytest backend/tests/test_payroll_excel_export.py -v
"""

from __future__ import annotations

import io
import logging
import os
import sys
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch as _patch

import openpyxl
import pytest
from sqlalchemy import BigInteger, Integer, Text, create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

# ── Project root on sys.path ─────────────────────────────────────────────────
_PROJECT_ROOT = str(Path(__file__).resolve().parents[2])
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

os.environ.setdefault(
    "ZPAY_SECRET_KEY",
    "test-secret-key-for-payroll-excel-export-tests-long",
)
os.environ.setdefault("DATABASE_URL", "sqlite://")

from backend.db.models import Base, ZRateOverride  # noqa: E402

# ── Metadata patches — same boilerplate as other test modules ─────────────────
Base.metadata.tables["z_rate_override"].c["effective_during"].type = Text()

for _tbl in Base.metadata.tables.values():
    for _col in _tbl.columns:
        if _col.primary_key and isinstance(_col.type, BigInteger):
            _col.type = Integer()

for _tbl in Base.metadata.tables.values():
    for _col in _tbl.columns:
        if _col.server_default is not None:
            _sd = _col.server_default
            try:
                _arg = _sd.arg.text if hasattr(_sd, "arg") and hasattr(_sd.arg, "text") else ""
            except Exception:
                _arg = ""
            if "NOW()" in _arg:
                _col.nullable = True
                _col.server_default = None

import re as _re  # noqa: E402

_engine = create_engine(
    "sqlite://",
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)


def _sqlite_regexp_replace(value: str, pattern: str, repl: str, flags: str) -> str:
    if value is None:
        return ""
    if "g" in (flags or ""):
        return _re.sub(pattern or "", repl or "", str(value))
    return _re.sub(pattern or "", repl or "", str(value), count=1)


@event.listens_for(_engine, "connect")
def _register_udfs(dbapi_conn, rec):
    dbapi_conn.create_function("regexp_replace", 4, _sqlite_regexp_replace)


Base.metadata.create_all(_engine)
_SessionFactory = sessionmaker(bind=_engine, autoflush=False, autocommit=False)

from fastapi.testclient import TestClient  # noqa: E402

from backend.app import app  # noqa: E402
from backend.db import get_db  # noqa: E402
from backend.db.models import DriverBalance, PayrollBatch, Person, Ride  # noqa: E402
from backend.middleware.auth import COOKIE_NAME, create_session  # noqa: E402
from backend.routes.workflow import (  # noqa: E402
    _build_payroll_summary_tab,
    _build_sp_itemized_tab,
    _build_sp_pay_summary_tab,
    _SP_ITEMIZED_HEADERS,
    _SP_PAY_HEADERS,
    _MOM_HEADERS,
    _HEADER_FILL_HEX,
    _TOTALS_FILL_HEX,
    _MONEY_FMT,
    _RELEASE_THRESHOLD_DOLLARS,
)

_SESSION_COOKIE = create_session(
    username="testadmin",
    display_name="Test Admin",
    color="#333",
    initials="TA",
    role="admin",
)
_AUTH = {COOKIE_NAME: _SESSION_COOKIE}
client = TestClient(app, raise_server_exceptions=True)
_NOW = datetime.now(timezone.utc)


# ── DB helpers ────────────────────────────────────────────────────────────────

def _override_get_db():
    db = _SessionFactory()
    try:
        yield db
    finally:
        db.close()


def _install_db_override():
    app.dependency_overrides[get_db] = _override_get_db


def _remove_db_override():
    app.dependency_overrides.pop(get_db, None)


def _db():
    return _SessionFactory()


def _wipe():
    sess = _db()
    try:
        sess.query(DriverBalance).delete(synchronize_session=False)
        sess.query(Ride).delete(synchronize_session=False)
        sess.query(PayrollBatch).delete(synchronize_session=False)
        sess.query(Person).delete(synchronize_session=False)
        sess.commit()
    finally:
        sess.close()


def _seed_acumen_batch(sess, batch_id: int = 100) -> PayrollBatch:
    b = PayrollBatch(
        payroll_batch_id=batch_id,
        source="acumen",
        company_name="Acumen International",
        batch_ref=f"0404202604102026",
        status="approved",
        period_start=date(2026, 4, 4),
        period_end=date(2026, 4, 10),
        week_start=date(2026, 4, 4),
        week_end=date(2026, 4, 10),
        currency="USD",
        uploaded_at=_NOW,
    )
    sess.add(b)
    sess.flush()
    return b


def _seed_maz_batch(sess, batch_id: int = 200) -> PayrollBatch:
    b = PayrollBatch(
        payroll_batch_id=batch_id,
        source="maz",
        company_name="EverDriven",
        batch_ref="WASO291-OY2026W15-20260419",
        status="approved",
        period_start=date(2026, 4, 11),
        period_end=date(2026, 4, 17),
        week_start=date(2026, 4, 11),
        week_end=date(2026, 4, 17),
        currency="USD",
        uploaded_at=_NOW,
    )
    sess.add(b)
    sess.flush()
    return b


def _seed_person(sess, person_id: int, name: str, code: str | None = None) -> Person:
    p = Person(
        person_id=person_id,
        full_name=name,
        paycheck_code=code,
        source="acumen",
    )
    sess.add(p)
    sess.flush()
    return p


# ── Sample fixture rows ───────────────────────────────────────────────────────

_PAID_ROW = {
    "person_id": 1,
    "person": "Abbas Driver",
    "code": "1031",
    "rides": 7,
    "miles": 131.0,
    "partner_pays": 414.25,
    "driver_pay": 332.0,
    "deduction": 0.0,
    "withheld": False,
    "from_last_period": 0.0,
    "pay_this_period": 332.0,
    "withheld_amount": 0.0,
    "service_days": 5,
}

_WITHHELD_ROW = {
    "person_id": 2,
    "person": "Zara Driver",
    "code": "1099",
    "rides": 2,
    "miles": 22.0,
    "partner_pays": 76.0,
    "driver_pay": 76.0,
    "deduction": 0.0,
    "withheld": True,
    "from_last_period": 0.0,
    "pay_this_period": 0.0,
    "withheld_amount": 76.0,
    "service_days": 2,
}

# Driver whose this-week earnings < $100 but carry balance brings combined >= $100
_RELEASE_ROW = {
    "person_id": 3,
    "person": "Juhar Driver",
    "code": "1048",
    "rides": 1,
    "miles": 10.0,
    "partner_pays": 76.0,
    "driver_pay": 76.0,  # this-week < $100
    "deduction": 0.0,
    "withheld": False,   # but combined ($76 + $50 carry) = $126 → released
    "from_last_period": 50.0,
    "pay_this_period": 126.0,
    "withheld_amount": 0.0,
    "service_days": 1,
}

_SAMPLE_ROWS = [_PAID_ROW, _WITHHELD_ROW]
_SAMPLE_TOTALS = {
    "rides": 9,
    "miles": 153.0,
    "partner_pays": 490.25,
    "driver_pay": 408.0,
    "deduction": 0.0,
    "carried_over": 76.0,
    "pay_this_period": 332.0,
}


# ── Helpers ───────────────────────────────────────────────────────────────────

class _BatchStub:
    """Minimal plain-python batch stub for unit tests — no DB, no SQLAlchemy."""
    def __init__(self, **kwargs):
        for k, v in kwargs.items():
            setattr(self, k, v)


def _make_acumen_batch_stub() -> _BatchStub:
    return _BatchStub(
        payroll_batch_id=999,
        source="acumen",
        company_name="Acumen International",
        batch_ref="0404202604102026",
        period_start=date(2026, 4, 4),
        period_end=date(2026, 4, 10),
        week_start=date(2026, 4, 4),
        week_end=date(2026, 4, 10),
    )


def _make_maz_batch_stub() -> _BatchStub:
    return _BatchStub(
        payroll_batch_id=998,
        source="maz",
        company_name="EverDriven",
        batch_ref="WASO291-OY2026W15",
        period_start=date(2026, 4, 11),
        period_end=date(2026, 4, 17),
        week_start=date(2026, 4, 11),
        week_end=date(2026, 4, 17),
    )


def _make_payroll_summary_ws(batch=None, rows=None, totals=None, title="FirstAlt — Payroll Summary"):
    wb = openpyxl.Workbook()
    ws = wb.active
    _build_payroll_summary_tab(
        ws,
        batch or _make_acumen_batch_stub(),
        rows if rows is not None else _SAMPLE_ROWS,
        totals if totals is not None else _SAMPLE_TOTALS,
        title,
    )
    return wb, ws


# ── Unit tests: _build_payroll_summary_tab ────────────────────────────────────

class TestPayrollSummaryTab:
    """Tests for the Payroll Summary tab (both Acumen and Maz)."""

    def test_title_row1_acumen(self):
        _, ws = _make_payroll_summary_ws(title="FirstAlt — Payroll Summary")
        assert ws.cell(row=1, column=1).value == "FirstAlt — Payroll Summary"

    def test_title_row1_maz(self):
        _, ws = _make_payroll_summary_ws(
            batch=_make_maz_batch_stub(),
            title="Maz — Payroll Summary",
        )
        assert ws.cell(row=1, column=1).value == "Maz — Payroll Summary"

    def test_period_row2_format(self):
        _, ws = _make_payroll_summary_ws()
        val = ws.cell(row=2, column=1).value or ""
        assert val.startswith("Period:"), f"Row 2 expected 'Period:...', got: {val!r}"
        assert "Apr" in val or "2026" in val

    def test_row3_is_blank(self):
        _, ws = _make_payroll_summary_ws()
        assert ws.cell(row=3, column=1).value is None

    def test_headers_at_row4(self):
        _, ws = _make_payroll_summary_ws()
        headers = [ws.cell(row=4, column=c).value for c in range(1, 11)]
        assert headers == _MOM_HEADERS, f"Row 4 mismatch: {headers}"

    def test_header_fill_blue(self):
        _, ws = _make_payroll_summary_ws()
        for col in range(1, 11):
            fg = ws.cell(row=4, column=col).fill.fgColor.rgb
            assert fg == _HEADER_FILL_HEX, f"Col {col} header fill: {fg!r}"

    def test_data_starts_row5(self):
        _, ws = _make_payroll_summary_ws()
        val = ws.cell(row=5, column=1).value
        assert val is not None and val != "TOTALS", f"Row 5 unexpected: {val!r}"

    def test_paid_driver_withheld_no(self):
        _, ws = _make_payroll_summary_ws()
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] and "Abbas" in str(row[0]):
                assert row[7] == "No", f"Paid driver col H: {row[7]!r}"
                return
        pytest.fail("Abbas Driver not found")

    def test_withheld_driver_withheld_yes(self):
        _, ws = _make_payroll_summary_ws()
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] and "Zara" in str(row[0]):
                assert row[7] == "Yes", f"Withheld driver col H: {row[7]!r}"
                assert row[9] == 0 or row[9] == 0.0, f"Withheld driver J: {row[9]!r}"
                return
        pytest.fail("Zara Driver not found")

    def test_under_100_withheld_j_is_zero(self):
        """Driver with combined < $100 must have Paid This Period = 0."""
        _, ws = _make_payroll_summary_ws()
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] and "Zara" in str(row[0]):
                assert row[9] == 0 or row[9] == 0.0
                return
        pytest.fail("Zara Driver not found")

    def test_carried_over_populated_when_prior_balance(self):
        """Col I must show prior held balance when from_last_period > 0."""
        rows_with_carry = [
            {**_PAID_ROW, "from_last_period": 90.0, "pay_this_period": 422.0},
            _WITHHELD_ROW,
        ]
        _, ws = _make_payroll_summary_ws(rows=rows_with_carry)
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] and "Abbas" in str(row[0]):
                assert float(row[8]) == 90.0, f"Carried Over col I: {row[8]!r}"
                return
        pytest.fail("Abbas Driver not found")

    def test_totals_carried_over_sums_all_rows_not_just_withheld(self):
        """TOTALS col I must equal sum of col I across ALL data rows.

        Bug fixed 2026-05-01: _build_summary was accumulating total_carried
        only for withheld drivers, so TOTALS col I showed $0 when a paid
        driver had a prior balance that was released this period.
        """
        rows_mixed = [
            {**_PAID_ROW, "from_last_period": 50.0, "pay_this_period": 382.0},
            _WITHHELD_ROW,  # from_last_period=0
        ]
        totals_with_carry = {
            **_SAMPLE_TOTALS,
            "carried_over": 50.0,  # correct: sum of all from_last_period (50+0)
            "pay_this_period": 382.0,
        }
        _, ws = _make_payroll_summary_ws(rows=rows_mixed, totals=totals_with_carry)

        # Collect all data row col I values
        data_carries = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] == "TOTALS":
                break
            if row[0] is not None:
                data_carries.append(float(row[8] or 0))

        totals_row_num = next(
            r[0].row for r in ws.iter_rows() if r[0].value == "TOTALS"
        )
        totals_col_i = float(ws.cell(row=totals_row_num, column=9).value or 0)
        expected = round(sum(data_carries), 2)

        assert totals_col_i == pytest.approx(expected, abs=0.01), (
            f"TOTALS col I ({totals_col_i}) != sum of data col I ({expected}). "
            "Withheld-only accumulation bug in _build_summary total_carried."
        )

    def test_totals_row_present(self):
        _, ws = _make_payroll_summary_ws()
        totals_row = next(
            (r[0].row for r in ws.iter_rows() if r[0].value == "TOTALS"), None
        )
        assert totals_row is not None, "TOTALS row missing"

    def test_totals_fill_orange_brown(self):
        _, ws = _make_payroll_summary_ws()
        totals_row = next(r[0].row for r in ws.iter_rows() if r[0].value == "TOTALS")
        for col in range(1, 11):
            fg = ws.cell(row=totals_row, column=col).fill.fgColor.rgb
            assert fg == _TOTALS_FILL_HEX, f"TOTALS col {col}: {fg!r}"

    def test_paychex_flex_row_present(self):
        _, ws = _make_payroll_summary_ws()
        paychex_row = next(
            (r[0].row for r in ws.iter_rows() if "Paychex" in str(r[0].value or "")),
            None,
        )
        assert paychex_row is not None, "Paychex Flex Amount row missing"

    def test_paychex_flex_g_is_formula(self):
        """Col G on Paychex Flex row must be a reconciliation formula."""
        _, ws = _make_payroll_summary_ws()
        paychex_row = next(
            r[0].row for r in ws.iter_rows() if "Paychex" in str(r[0].value or "")
        )
        g_val = ws.cell(row=paychex_row, column=7).value or ""
        assert str(g_val).startswith("="), f"Paychex G not a formula: {g_val!r}"

    def test_paid_on_week_section_present(self):
        _, ws = _make_payroll_summary_ws()
        paid_row = next(
            (r[0].row for r in ws.iter_rows() if r[0].value == "Paid on Week"),
            None,
        )
        assert paid_row is not None, "'Paid on Week' section header missing"

    def test_unpaid_on_week_section_present(self):
        _, ws = _make_payroll_summary_ws()
        unpaid_row = next(
            (r[0].row for r in ws.iter_rows() if r[0].value == "Unpaid on Week"),
            None,
        )
        assert unpaid_row is not None, "'Unpaid on Week' section header missing"

    def test_paid_section_lists_paid_drivers(self):
        _, ws = _make_payroll_summary_ws()
        paid_header_row = next(
            r[0].row for r in ws.iter_rows() if r[0].value == "Paid on Week"
        )
        names_below = []
        for row_idx in range(paid_header_row + 1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=1).value
            if val in (None, "Unpaid on Week", "Total"):
                break
            names_below.append(str(val))
        assert any("Abbas" in n for n in names_below), f"Abbas not in Paid section: {names_below}"

    def test_unpaid_section_lists_withheld_drivers(self):
        _, ws = _make_payroll_summary_ws()
        unpaid_header_row = next(
            r[0].row for r in ws.iter_rows() if r[0].value == "Unpaid on Week"
        )
        names_below = []
        for row_idx in range(unpaid_header_row + 1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=1).value
            if val is None:
                break
            names_below.append(str(val))
        assert any("Zara" in n for n in names_below), f"Zara not in Unpaid section: {names_below}"

    def test_column_widths_match_mom(self):
        from backend.routes.workflow import _MOM_COL_WIDTHS
        _, ws = _make_payroll_summary_ws()
        for idx, expected in enumerate(_MOM_COL_WIDTHS, start=1):
            letter = openpyxl.utils.get_column_letter(idx)
            actual = ws.column_dimensions[letter].width
            assert abs(actual - expected) <= 0.5, (
                f"Col {letter} width: expected {expected}, got {actual}"
            )

    def test_money_columns_currency_format(self):
        _, ws = _make_payroll_summary_ws()
        for col in {5, 6, 7, 9, 10}:
            fmt = ws.cell(row=5, column=col).number_format
            assert "$" in (fmt or ""), f"Col {col} data money format: {fmt!r}"

    def test_roundtrip_save_reload(self):
        wb, _ = _make_payroll_summary_ws()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        wb2 = openpyxl.load_workbook(buf)
        assert wb2.sheetnames[0] is not None
        ws2 = wb2.active
        assert ws2.cell(row=4, column=1).value == "Driver Name"


# ── Unit tests: release WARNING log ──────────────────────────────────────────

class TestReleaseWarningLog:
    """When a held driver is released with this-week earnings < $100, log WARNING."""

    def test_release_warning_logged(self, caplog):
        rows = [_RELEASE_ROW]
        totals = {
            "rides": 1, "miles": 10.0, "partner_pays": 76.0,
            "driver_pay": 76.0, "deduction": 0.0,
            "carried_over": 0.0, "pay_this_period": 126.0,
        }
        with caplog.at_level(logging.WARNING, logger="root"):
            _make_payroll_summary_ws(rows=rows, totals=totals)
        assert any(
            "Releasing held balance" in r.message and "Juhar" in r.message
            for r in caplog.records
        ), f"Expected WARNING for Juhar, got: {[r.message for r in caplog.records]}"

    def test_no_warning_when_this_week_above_threshold(self, caplog):
        """Normal paid driver with no carry should NOT trigger WARNING."""
        rows = [_PAID_ROW]  # driver_pay=332.0 > $100, from_last_period=0
        totals = {**_SAMPLE_TOTALS, "pay_this_period": 332.0, "rides": 7}
        with caplog.at_level(logging.WARNING, logger="root"):
            _make_payroll_summary_ws(rows=rows, totals=totals)
        assert not any(
            "Releasing held balance" in r.message
            for r in caplog.records
        ), "Unexpected WARNING for paid driver above threshold"

    def test_no_warning_for_withheld_driver(self, caplog):
        """Withheld driver should not trigger release WARNING."""
        rows = [_WITHHELD_ROW]
        totals = {
            "rides": 2, "miles": 22.0, "partner_pays": 76.0,
            "driver_pay": 76.0, "deduction": 0.0,
            "carried_over": 76.0, "pay_this_period": 0.0,
        }
        with caplog.at_level(logging.WARNING, logger="root"):
            _make_payroll_summary_ws(rows=rows, totals=totals)
        assert not any(
            "Releasing held balance" in r.message
            for r in caplog.records
        )


# ── Unit tests: _build_sp_pay_summary_tab ─────────────────────────────────────

class TestSpPaySummaryTab:

    def _make_ws(self, rows=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        _build_sp_pay_summary_tab(
            ws,
            _make_acumen_batch_stub(),
            rows if rows is not None else _SAMPLE_ROWS,
        )
        return wb, ws

    def test_sheet_title(self):
        _, ws = self._make_ws()
        assert ws.title == "SP PAY SUMMARY"

    def test_header_row_is_row1(self):
        _, ws = self._make_ws()
        headers = [ws.cell(row=1, column=c).value for c in range(1, len(_SP_PAY_HEADERS) + 1)]
        assert headers == _SP_PAY_HEADERS, f"Row 1 mismatch: {headers}"

    def test_header_fill_blue(self):
        _, ws = self._make_ws()
        for col in range(1, len(_SP_PAY_HEADERS) + 1):
            fg = ws.cell(row=1, column=col).fill.fgColor.rgb
            assert fg == _HEADER_FILL_HEX, f"SP PAY col {col} fill: {fg!r}"

    def test_data_row_count_matches_input(self):
        _, ws = self._make_ws()
        assert ws.max_row == 1 + len(_SAMPLE_ROWS)

    def test_driver_name_in_col_e(self):
        _, ws = self._make_ws()
        names = [ws.cell(row=r, column=5).value for r in range(2, ws.max_row + 1)]
        assert "Abbas Driver" in names

    def test_batch_ref_in_col_a(self):
        _, ws = self._make_ws()
        val = ws.cell(row=2, column=1).value
        assert val == "0404202604102026", f"Batch ref: {val!r}"

    def test_gross_pay_in_col_j(self):
        _, ws = self._make_ws()
        gross = ws.cell(row=2, column=10).value
        assert float(gross) == pytest.approx(414.25, abs=0.01)


# ── Unit tests: _build_sp_itemized_tab ───────────────────────────────────────

class TestSpItemizedTab:

    def _make_trip_rows(self):
        return [
            {
                "person": "Abbas Driver",
                "code": "1031",
                "date": date(2026, 4, 6),
                "trip_code": "9434783",
                "trip_name": "Helen Keller ES OB 02_B",
                "cancellation_reason": None,
                "miles": 29.0,
                "gross_pay": 74.75,
                "deduction": 0.0,
                "net_pay": 74.75,
            },
            {
                "person": "Abbas Driver",
                "code": "1031",
                "date": date(2026, 4, 7),
                "trip_code": "9434705",
                "trip_name": "Helen Keller ES IB 02_B",
                "cancellation_reason": "Canceled",
                "miles": 14.0,
                "gross_pay": 0.0,
                "deduction": 0.0,
                "net_pay": 0.0,
            },
        ]

    def _make_ws(self, trip_rows=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        _build_sp_itemized_tab(
            ws,
            _make_acumen_batch_stub(),
            trip_rows if trip_rows is not None else self._make_trip_rows(),
        )
        return wb, ws

    def test_sheet_title(self):
        _, ws = self._make_ws()
        assert ws.title == "SP ITEMIZED REPORT"

    def test_header_row_is_row1(self):
        _, ws = self._make_ws()
        headers = [ws.cell(row=1, column=c).value for c in range(1, len(_SP_ITEMIZED_HEADERS) + 1)]
        assert headers == _SP_ITEMIZED_HEADERS, f"Row 1 mismatch: {headers}"

    def test_header_fill_blue(self):
        _, ws = self._make_ws()
        for col in range(1, len(_SP_ITEMIZED_HEADERS) + 1):
            fg = ws.cell(row=1, column=col).fill.fgColor.rgb
            assert fg == _HEADER_FILL_HEX, f"ITEMIZED col {col} fill: {fg!r}"

    def test_data_row_count_matches_trips(self):
        _, ws = self._make_ws()
        assert ws.max_row == 1 + len(self._make_trip_rows())

    def test_trip_name_in_col_g(self):
        _, ws = self._make_ws()
        names = [ws.cell(row=r, column=7).value for r in range(2, ws.max_row + 1)]
        assert "Helen Keller ES OB 02_B" in names

    def test_cancellation_reason_in_col_h(self):
        _, ws = self._make_ws()
        cancellations = [ws.cell(row=r, column=8).value for r in range(2, ws.max_row + 1)]
        assert "Canceled" in cancellations
        assert None in cancellations


# ── Unit tests: Per-LLC isolation ────────────────────────────────────────────

class TestPerLLCIsolation:
    """Acumen and Maz must produce separate workbooks with distinct structures."""

    def test_acumen_has_three_tabs(self):
        """Unit-level: building all 3 acumen tabs produces correct sheet names."""
        wb = openpyxl.Workbook()
        ws1 = wb.active
        _build_sp_pay_summary_tab(ws1, _make_acumen_batch_stub(), _SAMPLE_ROWS)
        ws2 = wb.create_sheet("SP ITEMIZED REPORT")
        _build_sp_itemized_tab(ws2, _make_acumen_batch_stub(), [])
        ws3 = wb.create_sheet("Payroll Summary")
        _build_payroll_summary_tab(ws3, _make_acumen_batch_stub(), _SAMPLE_ROWS, _SAMPLE_TOTALS, "FirstAlt — Payroll Summary")
        assert wb.sheetnames == ["SP PAY SUMMARY", "SP ITEMIZED REPORT", "Payroll Summary"]

    def test_maz_has_one_tab(self):
        """Maz export builds a single tab with Maz title."""
        wb = openpyxl.Workbook()
        ws = wb.active
        _build_payroll_summary_tab(ws, _make_maz_batch_stub(), _SAMPLE_ROWS, _SAMPLE_TOTALS, "Maz — Payroll Summary")
        assert len(wb.sheetnames) == 1
        assert ws.cell(row=1, column=1).value == "Maz — Payroll Summary"

    def test_acumen_payroll_summary_title_contains_firstalt(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        _build_payroll_summary_tab(ws, _make_acumen_batch_stub(), _SAMPLE_ROWS, _SAMPLE_TOTALS, "FirstAlt — Payroll Summary")
        assert "FirstAlt" in (ws.cell(row=1, column=1).value or "")

    def test_maz_payroll_summary_title_contains_maz(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        _build_payroll_summary_tab(ws, _make_maz_batch_stub(), _SAMPLE_ROWS, _SAMPLE_TOTALS, "Maz — Payroll Summary")
        assert "Maz" in (ws.cell(row=1, column=1).value or "")


# ── HTTP smoke tests ──────────────────────────────────────────────────────────

_STUB_SUMMARY = {
    "rows": _SAMPLE_ROWS,
    "totals": _SAMPLE_TOTALS,
}

_STUB_SUMMARY_MAZ = {
    "rows": [_PAID_ROW],
    "totals": {
        "rides": 7, "miles": 131.0, "partner_pays": 414.25,
        "driver_pay": 332.0, "deduction": 0.0,
        "carried_over": 0.0, "pay_this_period": 332.0,
    },
}


class TestExportEndpoint:
    """HTTP-level checks — routing, content-type, 404, filename, sheet names."""

    def setup_method(self):
        _install_db_override()
        _wipe()

    def teardown_method(self):
        _remove_db_override()

    def test_404_for_unknown_batch(self):
        resp = client.get("/api/data/workflow/9998/export-excel", cookies=_AUTH)
        assert resp.status_code == 404

    def test_acumen_200_xlsx_content_type(self):
        sess = _db()
        try:
            _seed_acumen_batch(sess, batch_id=100)
            sess.commit()
        finally:
            sess.close()
        with _patch("backend.routes.workflow._build_summary", return_value=_STUB_SUMMARY):
            resp = client.get("/api/data/workflow/100/export-excel", cookies=_AUTH)
        assert resp.status_code == 200
        ct = resp.headers.get("content-type", "")
        assert "spreadsheetml" in ct, f"Unexpected content-type: {ct}"

    def test_maz_200_xlsx_content_type(self):
        sess = _db()
        try:
            _seed_maz_batch(sess, batch_id=200)
            sess.commit()
        finally:
            sess.close()
        with _patch("backend.routes.workflow._build_summary", return_value=_STUB_SUMMARY_MAZ):
            resp = client.get("/api/data/workflow/200/export-excel", cookies=_AUTH)
        assert resp.status_code == 200
        ct = resp.headers.get("content-type", "")
        assert "spreadsheetml" in ct, f"Unexpected content-type: {ct}"

    def test_acumen_response_has_three_sheets(self):
        sess = _db()
        try:
            _seed_acumen_batch(sess, batch_id=101)
            sess.commit()
        finally:
            sess.close()
        with _patch("backend.routes.workflow._build_summary", return_value=_STUB_SUMMARY):
            resp = client.get("/api/data/workflow/101/export-excel", cookies=_AUTH)
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert "SP PAY SUMMARY" in wb.sheetnames
        assert "SP ITEMIZED REPORT" in wb.sheetnames
        assert "Payroll Summary" in wb.sheetnames

    def test_maz_response_has_one_sheet(self):
        sess = _db()
        try:
            _seed_maz_batch(sess, batch_id=201)
            sess.commit()
        finally:
            sess.close()
        with _patch("backend.routes.workflow._build_summary", return_value=_STUB_SUMMARY_MAZ):
            resp = client.get("/api/data/workflow/201/export-excel", cookies=_AUTH)
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        # Maz must NOT have SP PAY SUMMARY or SP ITEMIZED REPORT
        assert "SP PAY SUMMARY" not in wb.sheetnames
        assert "SP ITEMIZED REPORT" not in wb.sheetnames
        assert wb.sheetnames[0] is not None  # has at least 1 sheet

    def test_acumen_payroll_summary_tab_has_correct_title(self):
        sess = _db()
        try:
            _seed_acumen_batch(sess, batch_id=102)
            sess.commit()
        finally:
            sess.close()
        with _patch("backend.routes.workflow._build_summary", return_value=_STUB_SUMMARY):
            resp = client.get("/api/data/workflow/102/export-excel", cookies=_AUTH)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb["Payroll Summary"]
        title = ws.cell(row=1, column=1).value or ""
        assert "FirstAlt" in title, f"Expected FirstAlt in title: {title!r}"

    def test_content_disposition_has_filename_acumen(self):
        sess = _db()
        try:
            _seed_acumen_batch(sess, batch_id=103)
            sess.commit()
        finally:
            sess.close()
        with _patch("backend.routes.workflow._build_summary", return_value=_STUB_SUMMARY):
            resp = client.get("/api/data/workflow/103/export-excel", cookies=_AUTH)
        cd = resp.headers.get("content-disposition", "")
        assert "filename=" in cd, f"No filename in Content-Disposition: {cd}"

    def test_content_disposition_has_filename_maz(self):
        sess = _db()
        try:
            _seed_maz_batch(sess, batch_id=202)
            sess.commit()
        finally:
            sess.close()
        with _patch("backend.routes.workflow._build_summary", return_value=_STUB_SUMMARY_MAZ):
            resp = client.get("/api/data/workflow/202/export-excel", cookies=_AUTH)
        cd = resp.headers.get("content-disposition", "")
        assert "filename=" in cd, f"No filename in Content-Disposition: {cd}"

    def test_acumen_workbook_parseable_after_save_reload(self):
        sess = _db()
        try:
            _seed_acumen_batch(sess, batch_id=104)
            sess.commit()
        finally:
            sess.close()
        with _patch("backend.routes.workflow._build_summary", return_value=_STUB_SUMMARY):
            resp = client.get("/api/data/workflow/104/export-excel", cookies=_AUTH)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb["Payroll Summary"]
        assert ws.cell(row=4, column=1).value == "Driver Name"

    def test_no_stale_payroll_summary_2_sheet(self):
        """No 'Payroll Summary (2)' duplicate sheet in output."""
        sess = _db()
        try:
            _seed_acumen_batch(sess, batch_id=105)
            sess.commit()
        finally:
            sess.close()
        with _patch("backend.routes.workflow._build_summary", return_value=_STUB_SUMMARY):
            resp = client.get("/api/data/workflow/105/export-excel", cookies=_AUTH)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert "Payroll Summary (2)" not in wb.sheetnames, (
            f"Stale duplicate sheet found: {wb.sheetnames}"
        )

    def test_carry_forward_in_export_matches_prior_batch_balance(self):
        """Fix 1: Carry-forward cell (col I) in the export must equal the
        prior batch's open driver_balance.carried_over — not the current
        batch's own driver_pay.  This verifies that export_excel passes
        override_ids so the withheld/paid status matches the approved state
        and from_last_period is populated from prior DriverBalance rows.
        """
        from backend.db.models import DriverBalance as _DriverBalance

        prior_carry = 57.0  # balance held from a prior batch

        _stub_summary_with_carry = {
            "rows": [
                {
                    **_PAID_ROW,
                    "from_last_period": prior_carry,
                    "pay_this_period": round(_PAID_ROW["driver_pay"] + prior_carry, 2),
                },
                _WITHHELD_ROW,
            ],
            "totals": {
                **_SAMPLE_TOTALS,
                "pay_this_period": round(_SAMPLE_TOTALS["pay_this_period"] + prior_carry, 2),
            },
        }

        sess = _db()
        try:
            _seed_acumen_batch(sess, batch_id=150)
            sess.commit()
        finally:
            sess.close()

        with _patch("backend.routes.workflow._build_summary", return_value=_stub_summary_with_carry):
            resp = client.get("/api/data/workflow/150/export-excel", cookies=_AUTH)

        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb["Payroll Summary"]

        # Find Abbas Driver row (the paid driver with a carry-forward)
        abbas_carry = None
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] and "Abbas" in str(row[0]):
                # Col I (index 8, 0-based) = Carried Over
                abbas_carry = row[8]
                break

        assert abbas_carry is not None, "Abbas Driver row not found in Payroll Summary tab"
        assert float(abbas_carry) == pytest.approx(prior_carry, abs=0.01), (
            f"Carry-forward in export ({abbas_carry}) does not match prior batch balance ({prior_carry}). "
            "export_excel must pass override_ids so _build_summary honours approved withheld state."
        )

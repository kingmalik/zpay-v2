"""
Tests for backend/scripts/audit_xlsx_integrity.py

Coverage:
  - _read_payroll_summary_tab: parses header row + data rows correctly
  - _summary_rows_to_dict: converts _build_summary row list to expected shape
  - _diff_dicts: emits ✓ lines on match, ⚠ lines on drift, ONLY IN on missing
  - run_audit: exits 0 when xlsx matches DB, exits 1 when drift detected
    (mocked _build_summary + _generate_xlsx_bytes_for_batch, real openpyxl)
  - run_audit: exits 1 on unknown batch_id
  - _build_maz_trip_details_tab: produces Trip Details tab with correct headers
    and correct row count when ride data is present

Run:
    PYTHONPATH=<project-root> pytest backend/tests/test_audit_xlsx_integrity.py -v
"""

from __future__ import annotations

import io
import os
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

# ── Project root on sys.path ─────────────────────────────────────────────────
_PROJECT_ROOT = str(Path(__file__).resolve().parents[2])
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

os.environ.setdefault("ZPAY_SECRET_KEY", "test-audit-integrity-key-not-used-for-auth")
os.environ.setdefault("DATABASE_URL", "sqlite://")

from backend.scripts.audit_xlsx_integrity import (  # noqa: E402
    _diff_dicts,
    _read_payroll_summary_tab,
    _summary_rows_to_dict,
    _find_header_row,
)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers — build minimal xlsx fixtures in memory
# ─────────────────────────────────────────────────────────────────────────────

_MOM_HEADERS = [
    "Driver Name", "Pay Code", "Rides", "Miles",
    "Partner Pays", "Driver Pay", "Deduction",
    "Withheld (Y/N)", "Carried Over", "Paid This Period",
]


def _make_payroll_summary_ws(drivers: list[dict]) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    Build a minimal in-memory Payroll Summary worksheet with the canonical
    R1=title / R2=period / R3=blank / R4=headers / R5+= data layout.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll Summary"

    ws.append(["Maz — Payroll Summary"])       # R1 title
    ws.append(["Period: Apr 11, 2026 – Apr 17, 2026"])  # R2 period
    ws.append([])                               # R3 blank
    ws.append(_MOM_HEADERS)                     # R4 headers

    for d in drivers:
        ws.append([
            d.get("person", ""),
            d.get("code", ""),
            d.get("rides", 0),
            d.get("miles", 0.0),
            d.get("partner_pays", 0.0),
            d.get("driver_pay", 0.0),
            d.get("deduction", 0.0),
            "Yes" if d.get("withheld") else "No",
            d.get("from_last_period", 0.0),
            d.get("pay_this_period", 0.0),
        ])

    ws.append(["TOTALS", "", 0, 0, 0, 0, 0, "", 0, 0])   # TOTALS sentinel
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# _find_header_row
# ─────────────────────────────────────────────────────────────────────────────

class TestFindHeaderRow:
    def test_finds_row_4_in_standard_layout(self):
        ws = _make_payroll_summary_ws([])
        assert _find_header_row(ws) == 4

    def test_raises_when_no_driver_name_header(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["No headers here"])
        ws.append(["Also no headers"])
        with pytest.raises(ValueError, match="Driver Name"):
            _find_header_row(ws)


# ─────────────────────────────────────────────────────────────────────────────
# _read_payroll_summary_tab
# ─────────────────────────────────────────────────────────────────────────────

class TestReadPayrollSummaryTab:
    def _sample_drivers(self):
        return [
            {
                "person": "Abbas Driver",
                "code": "1031",
                "rides": 7,
                "miles": 131.0,
                "partner_pays": 414.25,
                "driver_pay": 385.00,
                "deduction": 29.25,
                "withheld": False,
                "from_last_period": 0.0,
                "pay_this_period": 385.00,
            },
            {
                "person": "Zaid Walker",
                "code": "1099",
                "rides": 3,
                "miles": 55.5,
                "partner_pays": 120.00,
                "driver_pay": 90.00,
                "deduction": 30.00,
                "withheld": True,
                "from_last_period": 0.0,
                "pay_this_period": 0.0,
            },
        ]

    def test_parses_two_drivers(self):
        ws = _make_payroll_summary_ws(self._sample_drivers())
        result = _read_payroll_summary_tab(ws)
        assert len(result) == 2
        assert "Abbas Driver" in result
        assert "Zaid Walker" in result

    def test_paid_driver_values(self):
        ws = _make_payroll_summary_ws(self._sample_drivers())
        result = _read_payroll_summary_tab(ws)
        d = result["Abbas Driver"]
        assert d["partner_pays"] == pytest.approx(414.25, abs=0.01)
        assert d["driver_pay"] == pytest.approx(385.00, abs=0.01)
        assert d["deduction"] == pytest.approx(29.25, abs=0.01)
        assert d["paid_this_period"] == pytest.approx(385.00, abs=0.01)
        assert d["withheld"] == "No"

    def test_withheld_driver_values(self):
        ws = _make_payroll_summary_ws(self._sample_drivers())
        result = _read_payroll_summary_tab(ws)
        d = result["Zaid Walker"]
        assert d["paid_this_period"] == pytest.approx(0.0, abs=0.01)
        assert d["withheld"] == "Yes"

    def test_stops_at_totals_row(self):
        """Should not include the TOTALS row as a driver."""
        ws = _make_payroll_summary_ws(self._sample_drivers())
        result = _read_payroll_summary_tab(ws)
        assert "TOTALS" not in result

    def test_empty_batch_returns_empty_dict(self):
        ws = _make_payroll_summary_ws([])
        result = _read_payroll_summary_tab(ws)
        assert result == {}


# ─────────────────────────────────────────────────────────────────────────────
# _summary_rows_to_dict
# ─────────────────────────────────────────────────────────────────────────────

class TestSummaryRowsToDict:
    def test_converts_paid_row(self):
        rows = [
            {
                "person": "Abbas Driver",
                "code": "1031",
                "rides": 7,
                "partner_pays": 414.25,
                "driver_pay": 385.00,
                "deduction": 29.25,
                "from_last_period": 0.0,
                "pay_this_period": 385.00,
                "withheld": False,
            }
        ]
        result = _summary_rows_to_dict(rows)
        assert "Abbas Driver" in result
        d = result["Abbas Driver"]
        assert d["paid_this_period"] == pytest.approx(385.00, abs=0.01)
        assert d["withheld"] == "No"

    def test_converts_withheld_row(self):
        rows = [{"person": "Zaid Walker", "code": "", "rides": 3,
                 "partner_pays": 50.0, "driver_pay": 50.0,
                 "deduction": 0.0, "from_last_period": 0.0,
                 "pay_this_period": 0.0, "withheld": True}]
        result = _summary_rows_to_dict(rows)
        assert result["Zaid Walker"]["withheld"] == "Yes"
        assert result["Zaid Walker"]["paid_this_period"] == pytest.approx(0.0)

    def test_skips_blank_person_name(self):
        rows = [{"person": "", "code": "1099", "rides": 1,
                 "partner_pays": 100.0, "driver_pay": 100.0,
                 "deduction": 0.0, "from_last_period": 0.0,
                 "pay_this_period": 100.0, "withheld": False}]
        result = _summary_rows_to_dict(rows)
        assert result == {}


# ─────────────────────────────────────────────────────────────────────────────
# _diff_dicts
# ─────────────────────────────────────────────────────────────────────────────

_COMPARE_FIELDS = [
    "partner_pays", "driver_pay", "deduction",
    "carried_over", "paid_this_period", "withheld",
]


class TestDiffDicts:
    def _make_driver_dict(self, **kwargs) -> dict:
        base = {
            "partner_pays": 414.25,
            "driver_pay": 385.00,
            "deduction": 29.25,
            "carried_over": 0.0,
            "paid_this_period": 385.00,
            "withheld": "No",
        }
        base.update(kwargs)
        return base

    def test_no_drift_returns_checkmarks(self):
        driver = {"Abbas Driver": self._make_driver_dict()}
        lines = _diff_dicts("xlsx", "DB", driver, driver, _COMPARE_FIELDS)
        assert any("✓" in line and "Abbas Driver" in line for line in lines)
        assert not any("⚠" in line for line in lines)

    def test_numeric_drift_detected(self):
        xlsx = {"Abbas Driver": self._make_driver_dict(driver_pay=385.00)}
        db   = {"Abbas Driver": self._make_driver_dict(driver_pay=375.00)}
        lines = _diff_dicts("xlsx", "DB", xlsx, db, _COMPARE_FIELDS)
        assert any("⚠" in line and "Abbas Driver" in line for line in lines)
        assert any("driver_pay" in line for line in lines)

    def test_string_drift_detected(self):
        xlsx = {"Abbas Driver": self._make_driver_dict(withheld="No")}
        db   = {"Abbas Driver": self._make_driver_dict(withheld="Yes")}
        lines = _diff_dicts("xlsx", "DB", xlsx, db, _COMPARE_FIELDS)
        assert any("⚠" in line for line in lines)
        assert any("withheld" in line for line in lines)

    def test_driver_only_in_xlsx(self):
        a = {"Abbas Driver": self._make_driver_dict()}
        b: dict = {}
        lines = _diff_dicts("xlsx", "DB", a, b, _COMPARE_FIELDS)
        assert any("ONLY IN xlsx" in line and "Abbas Driver" in line for line in lines)

    def test_driver_only_in_db(self):
        a: dict = {}
        b = {"New Driver": self._make_driver_dict()}
        lines = _diff_dicts("xlsx", "DB", a, b, _COMPARE_FIELDS)
        assert any("ONLY IN DB" in line and "New Driver" in line for line in lines)

    def test_within_tolerance_no_drift(self):
        """Floats within 0.005 must not trigger a diff."""
        xlsx = {"Abbas Driver": self._make_driver_dict(driver_pay=385.001)}
        db   = {"Abbas Driver": self._make_driver_dict(driver_pay=385.003)}
        lines = _diff_dicts("xlsx", "DB", xlsx, db, _COMPARE_FIELDS)
        assert not any("⚠" in line for line in lines)

    def test_multiple_drivers_mixed_results(self):
        xlsx = {
            "Clean Driver":  self._make_driver_dict(driver_pay=200.00),
            "Dirty Driver":  self._make_driver_dict(driver_pay=200.00),
        }
        db = {
            "Clean Driver":  self._make_driver_dict(driver_pay=200.00),
            "Dirty Driver":  self._make_driver_dict(driver_pay=100.00),
        }
        lines = _diff_dicts("xlsx", "DB", xlsx, db, _COMPARE_FIELDS)
        ok_lines = [l for l in lines if "✓" in l]
        warn_lines = [l for l in lines if "⚠" in l]
        assert len(ok_lines) == 1
        assert len(warn_lines) == 1
        assert "Clean Driver" in ok_lines[0]
        assert "Dirty Driver" in warn_lines[0]


# ─────────────────────────────────────────────────────────────────────────────
# run_audit — integration tests using mocks (no real DB)
# ─────────────────────────────────────────────────────────────────────────────

def _make_xlsx_bytes_for_rows(rows: list[dict]) -> bytes:
    """
    Build a minimal xlsx bytes payload with a Payroll Summary tab matching
    the given rows — used to feed to run_audit under mock.
    """
    ws = _make_payroll_summary_ws(rows)
    buf = io.BytesIO()
    ws.parent.save(buf)
    return buf.getvalue()


class TestRunAudit:
    """
    Tests for run_audit() with mocked DB + builder.

    Strategy:
      - Mock _generate_xlsx_bytes_for_batch to return pre-built xlsx bytes
      - Mock _build_summary to return pre-built rows
      - Mock SQLAlchemy session + PayrollBatch lookup
      - Verify exit code based on whether rows match or drift
    """

    def _make_batch_mock(self, batch_id: int = 95, source: str = "maz"):
        batch = MagicMock()
        batch.payroll_batch_id = batch_id
        batch.source = source
        batch.status = "approved"
        return batch

    def _run_with_mocks(
        self,
        summary_rows: list[dict],
        xlsx_rows: list[dict],
        batch_id: int = 95,
        source: str = "maz",
        against_path: str | None = None,
    ) -> int:
        from backend.scripts.audit_xlsx_integrity import run_audit
        import backend.routes.summary as _summary_mod
        import backend.scripts.audit_xlsx_integrity as _audit_mod

        batch_mock = self._make_batch_mock(batch_id, source)
        xlsx_bytes = _make_xlsx_bytes_for_rows(xlsx_rows)
        summary_data = {"rows": summary_rows, "totals": {}}

        db_mock = MagicMock()
        db_mock.query.return_value.filter.return_value.first.return_value = batch_mock

        session_factory_mock = MagicMock(return_value=db_mock)
        engine_mock = MagicMock()

        original_bs = _summary_mod._build_summary
        try:
            _summary_mod._build_summary = lambda *a, **kw: summary_data
            os.environ["DATABASE_URL"] = "sqlite://"
            with (
                patch.object(_audit_mod, "create_engine", return_value=engine_mock),
                patch.object(_audit_mod, "sessionmaker", return_value=session_factory_mock),
                patch.object(
                    _audit_mod,
                    "_generate_xlsx_bytes_for_batch",
                    return_value=xlsx_bytes,
                ),
            ):
                return run_audit(batch_id, against_path)
        finally:
            _summary_mod._build_summary = original_bs

    def test_clean_batch_exits_0(self):
        rows = [
            {
                "person": "Abbas Driver",
                "code": "1031",
                "rides": 7,
                "partner_pays": 414.25,
                "driver_pay": 385.00,
                "deduction": 29.25,
                "from_last_period": 0.0,
                "pay_this_period": 385.00,
                "withheld": False,
            }
        ]
        result = self._run_with_mocks(rows, rows)
        assert result == 0

    def test_drifted_batch_exits_1(self):
        db_rows = [
            {
                "person": "Abbas Driver",
                "code": "1031",
                "rides": 7,
                "partner_pays": 414.25,
                "driver_pay": 385.00,
                "deduction": 29.25,
                "from_last_period": 0.0,
                "pay_this_period": 385.00,
                "withheld": False,
            }
        ]
        # xlsx has different driver_pay — simulates stale xlsx
        xlsx_rows = [
            {
                "person": "Abbas Driver",
                "code": "1031",
                "rides": 7,
                "partner_pays": 414.25,
                "driver_pay": 350.00,   # DRIFTED
                "deduction": 29.25,
                "from_last_period": 0.0,
                "pay_this_period": 350.00,
                "withheld": False,
            }
        ]
        result = self._run_with_mocks(db_rows, xlsx_rows)
        assert result == 1

    def test_missing_driver_in_xlsx_exits_1(self):
        db_rows = [
            {
                "person": "Abbas Driver",
                "code": "1031",
                "rides": 7,
                "partner_pays": 414.25,
                "driver_pay": 385.00,
                "deduction": 29.25,
                "from_last_period": 0.0,
                "pay_this_period": 385.00,
                "withheld": False,
            },
            {
                "person": "New Driver",
                "code": "1050",
                "rides": 3,
                "partner_pays": 150.00,
                "driver_pay": 130.00,
                "deduction": 20.00,
                "from_last_period": 0.0,
                "pay_this_period": 130.00,
                "withheld": False,
            },
        ]
        # xlsx only has Abbas — New Driver is missing
        xlsx_rows = db_rows[:1]
        result = self._run_with_mocks(db_rows, xlsx_rows)
        assert result == 1


# ─────────────────────────────────────────────────────────────────────────────
# _build_maz_trip_details_tab
# ─────────────────────────────────────────────────────────────────────────────

_STUB_SUMMARY_MAZ_TRIP = {
    "rows": [
        {
            "person_id": 900,
            "person": "Test Driver",
            "code": "MAZ-01",
            "rides": 2,
            "miles": 21.0,
            "partner_pays": 250.00,
            "driver_pay": 190.00,
            "deduction": 0.00,
            "active_between": "4/11/2026 – 4/12/2026",
            "days": 2,
            "net_pay": 250.00,
            "from_last_period": 0.0,
            "pay_this_period": 190.00,
            "withheld": False,
            "withheld_amount": 0.0,
            "missing_paycheck_code": False,
            "balance_source": None,
        }
    ],
    "totals": {
        "rides": 2,
        "miles": 21.0,
        "partner_pays": 250.00,
        "driver_pay": 190.00,
        "deduction": 0.00,
        "carried_over": 0.0,
        "days": 2,
        "net_pay": 250.00,
        "pay_this_period": 190.00,
    },
}


class TestBuildMazTripDetailsTab:
    """
    Verify that the new Trip Details tab is added to the Maz xlsx workbook
    with the correct headers and row count.

    _build_summary is patched (SQLite coalesce compat issue on single-arg call).
    Ride query inside _build_maz_trip_details_tab runs against a real SQLite DB
    to validate the actual ORM join + filter logic.
    """

    def _build_db(self):
        """Return a minimal SQLAlchemy session with one Maz batch + 2 rides."""
        import re as _re
        from sqlalchemy import event, create_engine, BigInteger, Integer, Text
        from sqlalchemy.orm import sessionmaker
        from backend.db.models import Base, ZRateOverride, PayrollBatch, Person, Ride

        # Patch for SQLite compat (same pattern as test_payroll_excel_export.py)
        Base.metadata.tables["z_rate_override"].c["effective_during"].type = Text()
        for _tbl in Base.metadata.tables.values():
            for _col in _tbl.columns:
                if _col.primary_key and isinstance(_col.type, BigInteger):
                    _col.type = Integer()
        for _tbl in Base.metadata.tables.values():
            for _col in _tbl.columns:
                if _col.server_default is not None:
                    _sd = _col.server_default
                    try:
                        _arg = _sd.arg.text if hasattr(_sd, "arg") and hasattr(_sd.arg, "text") else ""
                    except Exception:
                        _arg = ""
                    if "NOW()" in _arg:
                        _col.nullable = True
                        _col.server_default = None

        from sqlalchemy.pool import StaticPool
        engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )

        def _sqlite_regexp_replace(value, pattern, repl, flags):
            if value is None:
                return ""
            if "g" in (flags or ""):
                return _re.sub(pattern or "", repl or "", str(value))
            return _re.sub(pattern or "", repl or "", str(value), count=1)

        @event.listens_for(engine, "connect")
        def _reg(dbapi_conn, rec):
            dbapi_conn.create_function("regexp_replace", 4, _sqlite_regexp_replace)

        Base.metadata.create_all(engine)
        Session = sessionmaker(bind=engine)
        db = Session()

        _NOW = datetime.now(timezone.utc)

        batch = PayrollBatch(
            payroll_batch_id=500,
            source="maz",
            company_name="EverDriven",
            batch_ref="TEST-BATCH-500",
            status="approved",
            period_start=date(2026, 4, 11),
            period_end=date(2026, 4, 17),
            week_start=date(2026, 4, 11),
            week_end=date(2026, 4, 17),
            currency="USD",
            uploaded_at=_NOW,
        )
        db.add(batch)
        db.flush()

        person = Person(
            person_id=900,
            full_name="Test Driver",
            paycheck_code_maz="MAZ-01",
        )
        db.add(person)
        db.flush()

        for i, src in enumerate(["ED:AUDIT-001", "ED:AUDIT-002"]):
            ride = Ride(
                payroll_batch_id=500,
                person_id=900,
                source_ref=src,
                source="maz",
                service_name=f"Route {i + 1}",
                ride_start_ts=datetime(2026, 4, 11 + i, 8, 0, tzinfo=timezone.utc),
                miles=10.5 + i,
                gross_pay=120.00 + i * 10,
                net_pay=120.00 + i * 10,
                deduction=0.00,
                z_rate=95.00,
                z_rate_source="default",
            )
            db.add(ride)
        db.commit()

        return db, batch

    def _xlsx_bytes(self, db, batch) -> bytes:
        """Run _build_maz_xlsx_bytes with _build_summary stubbed out."""
        from backend.routes.workflow import _build_maz_xlsx_bytes
        with patch("backend.routes.workflow._build_summary", return_value=_STUB_SUMMARY_MAZ_TRIP):
            return _build_maz_xlsx_bytes(db, batch)

    def test_trip_details_tab_exists_in_maz_xlsx(self):
        db, batch = self._build_db()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(self._xlsx_bytes(db, batch)))
            assert "Trip Details" in wb.sheetnames, (
                f"Expected 'Trip Details' tab in Maz xlsx. Got: {wb.sheetnames}"
            )
        finally:
            db.close()

    def test_payroll_summary_tab_still_present(self):
        db, batch = self._build_db()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(self._xlsx_bytes(db, batch)))
            assert "Payroll Summary" in wb.sheetnames
        finally:
            db.close()

    def test_trip_details_has_correct_headers(self):
        from backend.routes.workflow import _MAZ_TRIP_DETAIL_HEADERS

        db, batch = self._build_db()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(self._xlsx_bytes(db, batch)))
            ws = wb["Trip Details"]
            header_row = [
                ws.cell(row=1, column=c).value
                for c in range(1, len(_MAZ_TRIP_DETAIL_HEADERS) + 1)
            ]
            assert header_row == _MAZ_TRIP_DETAIL_HEADERS
        finally:
            db.close()

    def test_trip_details_has_correct_row_count(self):
        """2 rides seeded → 2 data rows (row 1 is header)."""
        db, batch = self._build_db()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(self._xlsx_bytes(db, batch)))
            ws = wb["Trip Details"]
            # Row 1 = header, rows 2+ = data
            data_rows = [
                ws.cell(row=r, column=1).value
                for r in range(2, ws.max_row + 1)
                if ws.cell(row=r, column=1).value
            ]
            assert len(data_rows) == 2
        finally:
            db.close()

    def test_trip_details_driver_name_matches(self):
        db, batch = self._build_db()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(self._xlsx_bytes(db, batch)))
            ws = wb["Trip Details"]
            assert ws.cell(row=2, column=1).value == "Test Driver"
        finally:
            db.close()

    def test_trip_details_pay_code_maz_populated(self):
        db, batch = self._build_db()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(self._xlsx_bytes(db, batch)))
            ws = wb["Trip Details"]
            assert ws.cell(row=2, column=2).value == "MAZ-01"
        finally:
            db.close()

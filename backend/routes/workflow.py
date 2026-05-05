"""
Workflow API endpoints for the guided payroll workflow.
All routes under /api/data/workflow/* return JSON.
"""
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from backend.db import get_db
from backend.db.models import (
    PayrollBatch, Ride, Person, EmailSendLog, ZRateService, BatchWorkflowLog, DriverBalance,
)
from backend.services.workflow import (
    STAGE_ORDER, advance_batch, reopen_batch, check_gate, next_stage,
)
from backend.routes.summary import _build_summary
from backend.utils.week_label import week_label as _wl
from backend.utils.roles import require_role

router = APIRouter(prefix="/api/data/workflow", tags=["workflow"])


def _safe_slug(s: str) -> str:
    import re
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s or "batch"


def _fmt_period(batch) -> str:
    """Return a filename-safe period string like 'mar17_mar21_2026'."""
    ws = getattr(batch, "week_start", None) or getattr(batch, "period_start", None)
    we = getattr(batch, "week_end", None) or getattr(batch, "period_end", None)
    if ws and we:
        return f"{ws.strftime('%b%d').lower()}_{we.strftime('%b%d_%Y').lower()}"
    if ws:
        return ws.strftime("%Y_%m_%d")
    return f"batch_{batch.payroll_batch_id}"


# ── Rate management helpers ──────────────────────────────────────────────────

@router.post("/rates/create")
async def workflow_create_rate(request=None, db: Session = Depends(get_db)):
    """Create a new z_rate_service and apply it to matching rides."""
    import re
    from decimal import Decimal

    body = await request.json()
    service_name = body.get("service_name", "").strip()
    source = body.get("source", "")
    company_name = body.get("company_name", "")
    rate = body.get("default_rate")

    if not service_name or not rate:
        return JSONResponse({"error": "service_name and default_rate required"}, status_code=400)

    # Check if already exists
    existing = db.query(ZRateService).filter(ZRateService.service_name == service_name).first()
    if existing:
        existing.default_rate = Decimal(str(rate))
        svc = existing
    else:
        svc = ZRateService(
            service_key=re.sub(r"[^a-z0-9]+", "_", service_name.lower()).strip("_"),
            service_name=service_name,
            source=source,
            company_name=company_name,
            default_rate=Decimal(str(rate)),
        )
        db.add(svc)
        db.flush()

    # Update all rides matching this service_name with z_rate=0
    updated = (
        db.query(Ride)
        .filter(Ride.service_name == service_name, Ride.z_rate == 0)
        .update({
            "z_rate": float(rate),
            "z_rate_service_id": svc.z_rate_service_id,
        }, synchronize_session=False)
    )
    db.commit()

    return JSONResponse({"ok": True, "service_id": svc.z_rate_service_id, "rides_updated": updated})


@router.post("/rates/apply-batch/{batch_id}")
def workflow_apply_batch_rates(batch_id: int, db: Session = Depends(get_db)):
    """Re-apply z_rate_service rates to all z_rate=0 rides in a batch."""
    unpriced = (
        db.query(Ride)
        .filter(Ride.payroll_batch_id == batch_id, Ride.z_rate == 0)
        .all()
    )

    updated = 0
    for ride in unpriced:
        svc = (
            db.query(ZRateService)
            .filter(ZRateService.service_name == ride.service_name)
            .first()
        )
        if svc and float(svc.default_rate) > 0:
            ride.z_rate = float(svc.default_rate)
            ride.z_rate_service_id = svc.z_rate_service_id
            updated += 1

    db.commit()
    return JSONResponse({"ok": True, "rides_updated": updated})


def _display_company(raw: str) -> str:
    co = (raw or "").lower()
    if "ever" in co:
        return "EverDriven"
    return "FirstAlt"


def _batch_summary(db: Session, batch: PayrollBatch) -> dict:
    """Build a summary dict for a single batch."""
    bid = batch.payroll_batch_id

    ride_stats = db.query(
        func.count(Ride.ride_id).label("rides"),
        func.coalesce(func.sum(Ride.net_pay), 0).label("revenue"),
        func.coalesce(func.sum(Ride.z_rate), 0).label("cost"),
        func.count(Ride.ride_id).filter(Ride.z_rate == 0).label("unpriced"),
    ).filter(Ride.payroll_batch_id == bid).one()

    # Driver count
    driver_count = (
        db.query(func.count(func.distinct(Ride.person_id)))
        .filter(Ride.payroll_batch_id == bid)
        .scalar() or 0
    )

    # Email send stats
    sent_count = (
        db.query(func.count(EmailSendLog.id))
        .filter(EmailSendLog.payroll_batch_id == bid, EmailSendLog.status == "sent")
        .scalar() or 0
    )
    failed_count = (
        db.query(func.count(EmailSendLog.id))
        .filter(EmailSendLog.payroll_batch_id == bid, EmailSendLog.status == "failed")
        .scalar() or 0
    )

    return {
        "batch_id": bid,
        "source": batch.source,
        "company": _display_company(batch.company_name or ""),
        "company_raw": batch.company_name,
        "batch_ref": batch.batch_ref,
        "status": batch.status,
        "week_label": f"Week {(db.query(func.count(PayrollBatch.payroll_batch_id)).filter(PayrollBatch.source == batch.source, PayrollBatch.period_start <= batch.period_start).scalar() or 1)}",
        "period_start": batch.period_start.isoformat() if batch.period_start else None,
        "period_end": batch.period_end.isoformat() if batch.period_end else None,
        "week_start": batch.week_start.isoformat() if batch.week_start else None,
        "week_end": batch.week_end.isoformat() if batch.week_end else None,
        "uploaded_at": batch.uploaded_at.isoformat() if batch.uploaded_at else None,
        "finalized_at": batch.finalized_at.isoformat() if batch.finalized_at else None,
        "paychex_exported_at": batch.paychex_exported_at.isoformat() if batch.paychex_exported_at else None,
        "rides": int(ride_stats.rides or 0),
        "revenue": round(float(ride_stats.revenue or 0), 2),
        "cost": round(float(ride_stats.cost or 0), 2),
        "margin": round(float((ride_stats.revenue or 0) - (ride_stats.cost or 0)), 2),
        "unpriced_rides": int(ride_stats.unpriced or 0),
        "driver_count": driver_count,
        "stubs_sent": sent_count,
        "stubs_failed": failed_count,
    }


# ── Active batches ───────────────────────────────────────────────────────────

@router.get("/active")
def workflow_active(db: Session = Depends(get_db)):
    """All batches not yet complete, ordered by most recent first."""
    batches = (
        db.query(PayrollBatch)
        .filter(PayrollBatch.status != "complete")
        .order_by(PayrollBatch.uploaded_at.desc())
        .all()
    )
    return JSONResponse({
        "batches": [_batch_summary(db, b) for b in batches],
    })


# ── Single batch status ─────────────────────────────────────────────────────

@router.get("/{batch_id}/status")
def workflow_status(batch_id: int, db: Session = Depends(get_db)):
    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    summary = _batch_summary(db, batch)

    # Check what's blocking the next advance
    nxt = next_stage(batch.status)
    blockers = []
    warnings = []
    if nxt:
        _, blockers, warnings = check_gate(db, batch, nxt)

    # Get workflow log
    logs = (
        db.query(BatchWorkflowLog)
        .filter(BatchWorkflowLog.payroll_batch_id == batch_id)
        .order_by(BatchWorkflowLog.created_at.asc())
        .all()
    )

    return JSONResponse({
        **summary,
        "next_stage": nxt,
        "blockers": blockers,
        "warnings": warnings,
        "stage_index": STAGE_ORDER.index(batch.status) if batch.status in STAGE_ORDER else 0,
        "stages": STAGE_ORDER,
        "logs": [
            {
                "from": log.from_status,
                "to": log.to_status,
                "by": log.triggered_by,
                "notes": log.notes,
                "at": log.created_at.isoformat() if log.created_at else None,
            }
            for log in logs
        ],
    })


# ── Advance batch ───────────────────────────────────────────────────────────

@router.post("/{batch_id}/advance")
async def workflow_advance(batch_id: int, request=None, db: Session = Depends(get_db)):
    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    force = False
    notes = None
    if request:
        try:
            body = await request.json()
            force = body.get("force", False)
            notes = body.get("notes")
        except Exception:
            pass

    success, new_status, blockers = advance_batch(db, batch, triggered_by="user", force=force, notes=notes)

    if not success:
        return JSONResponse({
            "ok": False,
            "status": batch.status,
            "blockers": blockers,
        }, status_code=400)

    return JSONResponse({
        "ok": True,
        "status": new_status,
        "blockers": blockers,
    })


# ── Reopen batch ─────────────────────────────────────────────────────────────

@router.post("/{batch_id}/reopen")
def workflow_reopen(batch_id: int, db: Session = Depends(get_db), _admin=Depends(require_role("admin"))):
    """Reset batch to payroll_review stage. Admin-only — wipes calculated balances."""
    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    # Block reopen if any real (non-test) stubs have been sent
    real_sends = (
        db.query(EmailSendLog)
        .filter(
            EmailSendLog.payroll_batch_id == batch_id,
            EmailSendLog.status == "sent",
            EmailSendLog.is_test == False,  # noqa: E712
        )
        .count()
    )
    if real_sends > 0:
        return JSONResponse(
            {"ok": False, "error": f"Cannot reopen — {real_sends} real stubs already sent to drivers."},
            status_code=400,
        )

    success, status = reopen_batch(db, batch)
    if not success:
        return JSONResponse({"ok": False, "error": status}, status_code=400)

    return JSONResponse({"ok": True, "status": status})


# ── Lock & Approve (explicit commit point for Review step) ──────────────────

@router.post("/{batch_id}/lock-and-approve")
async def workflow_lock_and_approve(batch_id: int, request: Request, db: Session = Depends(get_db)):
    """Explicit 'Lock & Approve' button endpoint — advances batch from payroll_review
    to approved. Equivalent to /advance but semantically tied to the Review CTA.
    Admin bypass: admin role may call this from any step — the advance runs with force=True."""
    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    caller = getattr(request.state, "user", None)
    is_admin_caller = caller and caller.get("role") == "admin"

    if batch.status != "payroll_review" and not is_admin_caller:
        return JSONResponse(
            {"ok": False, "error": f"Batch is in '{batch.status}', not 'payroll_review'. Cannot approve."},
            status_code=400,
        )

    success, new_status, blockers = advance_batch(db, batch, triggered_by="user", force=is_admin_caller)
    if not success:
        return JSONResponse({"ok": False, "status": batch.status, "blockers": blockers}, status_code=400)

    return JSONResponse({"ok": True, "status": new_status, "blockers": blockers})


# ── Go back one stage ─────────────────────────────────────────────────────────

@router.post("/{batch_id}/go-back")
def workflow_go_back(batch_id: int, db: Session = Depends(get_db)):
    """Move the batch back one stage in the workflow pipeline."""
    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    current = batch.status
    try:
        idx = STAGE_ORDER.index(current)
    except ValueError:
        return JSONResponse({"ok": False, "error": f"Unknown stage: {current}"}, status_code=400)

    if idx == 0:
        return JSONResponse({"ok": False, "error": "Already at the first stage"}, status_code=400)

    prev_stage = STAGE_ORDER[idx - 1]
    old_status = current

    # Cleanup depending on what we're undoing
    if current in ("approved", "export_ready"):
        # Undo payroll run — clear driver balances
        db.query(DriverBalance).filter(
            DriverBalance.payroll_batch_id == batch_id
        ).delete()
        batch.finalized_at = None

    elif current == "stubs_sending":
        # Going back to export_ready — clear email send logs so they can resend fresh
        db.query(EmailSendLog).filter(
            EmailSendLog.payroll_batch_id == batch_id
        ).delete()

    batch.status = prev_stage

    db.add(BatchWorkflowLog(
        payroll_batch_id=batch_id,
        from_status=old_status,
        to_status=prev_stage,
        triggered_by="user",
        notes=f"Went back from {old_status}",
    ))
    db.commit()

    return JSONResponse({"ok": True, "status": prev_stage})


# ── Rates check ──────────────────────────────────────────────────────────────

@router.get("/{batch_id}/rates-check")
def workflow_rates_check(batch_id: int, db: Session = Depends(get_db)):
    """Get rides with z_rate=0, grouped by service name, with suggested rates."""
    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    # Get unpriced rides grouped by service_name
    unpriced = (
        db.query(
            Ride.service_name,
            func.count(Ride.ride_id).label("count"),
            func.sum(Ride.net_pay).label("total_net_pay"),
            func.array_agg(func.distinct(Person.full_name)).label("drivers"),
        )
        .join(Person, Person.person_id == Ride.person_id)
        .filter(Ride.payroll_batch_id == batch_id, Ride.z_rate == 0)
        .group_by(Ride.service_name)
        .order_by(func.count(Ride.ride_id).desc())
        .all()
    )

    groups = []
    for row in unpriced:
        service_name = row.service_name or "Unknown"

        # Look for existing rate in z_rate_service
        existing_rate = (
            db.query(ZRateService)
            .filter(ZRateService.service_name == service_name)
            .first()
        )

        suggested_rate = float(existing_rate.default_rate) if existing_rate else None

        groups.append({
            "service_name": service_name,
            "count": int(row.count),
            "total_net_pay": round(float(row.total_net_pay or 0), 2),
            "drivers": list(row.drivers) if row.drivers else [],
            "suggested_rate": suggested_rate,
            "service_id": existing_rate.z_rate_service_id if existing_rate else None,
        })

    total_unpriced = sum(g["count"] for g in groups)

    return JSONResponse({
        "batch_id": batch_id,
        "total_unpriced": total_unpriced,
        "groups": groups,
    })


# ── Payroll preview with validation warnings ─────────────────────────────────

@router.get("/{batch_id}/payroll-preview")
def workflow_payroll_preview(batch_id: int, db: Session = Depends(get_db)):
    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    from sqlalchemy import text as _text
    override_rows = db.execute(
        _text("SELECT person_id FROM payroll_withheld_override WHERE batch_id = :b"),
        {"b": batch_id},
    ).fetchall()
    override_ids = {r[0] for r in override_rows}

    manual_withhold_rows = db.execute(
        _text("SELECT person_id, note FROM payroll_manual_withhold"),
    ).fetchall()
    manual_withhold_map = {r[0]: r[1] for r in manual_withhold_rows}

    data = _build_summary(db, batch_id=batch_id, auto_save=False, override_ids=override_ids or None, manual_withhold_ids=set(manual_withhold_map.keys()) or None)
    rows = data["rows"]
    totals = data["totals"]

    # Build validation warnings
    warnings = []

    # Drivers missing paycheck_code
    person_ids = [r["person_id"] for r in rows if not r["withheld"]]
    if person_ids:
        missing_pay_code = (
            db.query(Person)
            .filter(
                Person.person_id.in_(person_ids),
                (Person.paycheck_code.is_(None)) | (Person.paycheck_code == ""),
            )
            .all()
        )
        if missing_pay_code:
            names = [p.full_name for p in missing_pay_code[:5]]
            warnings.append({
                "severity": "warning",
                "title": f"{len(missing_pay_code)} drivers missing Paychex code",
                "description": f"Won't be included in Paychex CSV: {', '.join(names)}" + ("..." if len(missing_pay_code) > 5 else ""),
                "type": "missing_pay_code",
                "count": len(missing_pay_code),
                "affected": [
                    {"person_id": p.person_id, "name": p.full_name, "paycheck_code": p.paycheck_code or ""}
                    for p in missing_pay_code
                ],
            })

    # Drivers missing email
    all_person_ids = [r["person_id"] for r in rows]
    if all_person_ids:
        missing_email = (
            db.query(Person)
            .filter(
                Person.person_id.in_(all_person_ids),
                (Person.email.is_(None)) | (Person.email == ""),
            )
            .all()
        )
        if missing_email:
            names = [p.full_name for p in missing_email[:5]]
            warnings.append({
                "severity": "info",
                "title": f"{len(missing_email)} drivers missing email",
                "description": f"Won't receive paystub: {', '.join(names)}" + ("..." if len(missing_email) > 5 else ""),
                "type": "missing_email",
                "count": len(missing_email),
                "affected": [
                    {"person_id": p.person_id, "name": p.full_name, "email": p.email or ""}
                    for p in missing_email
                ],
            })

    # Negative margins (z_rate > net_pay) — possible data entry error
    for r in rows:
        if r["net_pay"] < 0:
            warnings.append({
                "severity": "error",
                "title": f"Negative net pay for {r['person']}",
                "description": f"Net pay: ${r['net_pay']:.2f}",
                "type": "negative_net",
            })

    # Rides with z_rate > net_pay per ride (aggregate check)
    negative_margin_ride_rows = (
        db.query(Ride.service_name, Ride.z_rate, Ride.net_pay, func.count(Ride.ride_id).label("cnt"))
        .filter(Ride.payroll_batch_id == batch_id, Ride.z_rate > Ride.net_pay, Ride.net_pay > 0)
        .group_by(Ride.service_name, Ride.z_rate, Ride.net_pay)
        .all()
    )
    negative_margin_rides = sum(int(r.cnt) for r in negative_margin_ride_rows)
    if negative_margin_rides > 0:
        neg_details = []
        for r in negative_margin_ride_rows:
            neg_details.append({
                "service_name": r.service_name or "Unknown",
                "z_rate": round(float(r.z_rate or 0), 2),
                "net_pay": round(float(r.net_pay or 0), 2),
                "count": int(r.cnt),
            })
        warnings.append({
            "severity": "warning",
            "title": f"{negative_margin_rides} rides with negative margin",
            "description": "Driver rate exceeds company rate — check rate assignments",
            "type": "negative_margin",
            "count": negative_margin_rides,
            "affected": neg_details,
        })

    # Late cancellation detection (EverDriven / source="maz")
    # Rides canceled within 2 hours get paid at half price.
    # Detect: net_pay is 40-55% of z_rate (both > 0).
    late_cancel_rides = (
        db.query(Ride, Person.full_name)
        .join(Person, Person.person_id == Ride.person_id)
        .filter(
            Ride.payroll_batch_id == batch_id,
            Ride.source == "maz",
            Ride.z_rate > 0,
            Ride.net_pay > 0,
            Ride.net_pay >= Ride.z_rate * 0.40,
            Ride.net_pay <= Ride.z_rate * 0.55,
        )
        .all()
    )
    if late_cancel_rides:
        lc_list = []
        for ride, driver_name in late_cancel_rides:
            z = float(ride.z_rate)
            n = float(ride.net_pay)
            lc_list.append({
                "driver": driver_name,
                "route": ride.service_name or "Unknown",
                "z_rate": round(z, 2),
                "net_pay": round(n, 2),
                "ratio": round(n / z, 2) if z else 0,
            })
        warnings.append({
            "severity": "warning",
            "title": "Late Cancellations Detected",
            "description": f"{len(lc_list)} EverDriven rides appear to be 2-hour cancellations (paid at ~50% of rate)",
            "type": "late_cancellation",
            "count": len(lc_list),
            "rides": lc_list,
        })

    # Net pay change detection — compares discrete per-route rates between this batch
    # and the most recent prior completed batch.  No averages — every rate is a set
    # number per route.  z_rate = driver pay, net_pay = partner pay per ride.
    # We use MAX() as the discrete value selector; all rides on a given route in one
    # batch carry the same stamped rate, so MAX == the actual rate.
    current_routes = (
        db.query(
            Ride.service_name,
            Ride.source,
            func.max(Ride.net_pay).label("partner_now"),
            func.max(Ride.z_rate).label("driver_now"),
        )
        .filter(Ride.payroll_batch_id == batch_id, Ride.net_pay > 0)
        .group_by(Ride.service_name, Ride.source)
        .all()
    )

    # Find the most recent prior completed batch (same source if possible)
    prior_batch = (
        db.query(PayrollBatch)
        .filter(
            PayrollBatch.payroll_batch_id < batch_id,
            PayrollBatch.status == "complete",
        )
        .order_by(PayrollBatch.payroll_batch_id.desc())
        .first()
    )

    net_pay_changes = []
    if prior_batch:
        prior_rates = (
            db.query(
                Ride.service_name,
                Ride.source,
                func.max(Ride.net_pay).label("partner_before"),
                func.max(Ride.z_rate).label("driver_before"),
            )
            .filter(
                Ride.payroll_batch_id == prior_batch.payroll_batch_id,
                Ride.net_pay > 0,
            )
            .group_by(Ride.service_name, Ride.source)
            .all()
        )
        prior_map = {(r.service_name, r.source): r for r in prior_rates}

        for route in current_routes:
            sn, src = route.service_name, route.source
            if not sn:
                continue
            partner_now = float(route.partner_now or 0)
            driver_now = float(route.driver_now or 0)
            if partner_now == 0:
                continue

            prior = prior_map.get((sn, src))
            if not prior:
                continue
            partner_before = float(prior.partner_before or 0)
            driver_before = float(prior.driver_before or 0)
            if partner_before == 0:
                continue

            partner_delta = round(partner_now - partner_before, 2)
            driver_delta = round(driver_now - driver_before, 2)
            margin_now = round(partner_now - driver_now, 2)
            margin_before = round(partner_before - driver_before, 2)
            margin_delta = round(margin_now - margin_before, 2)

            # Only surface routes where something actually changed
            if partner_delta == 0 and driver_delta == 0:
                continue

            net_pay_changes.append({
                "route": sn,
                "partner_before": partner_before,
                "partner_now": partner_now,
                "partner_delta": partner_delta,
                "driver_before": driver_before,
                "driver_now": driver_now,
                "driver_delta": driver_delta,
                "margin_delta": margin_delta,
            })

        # Sort: biggest margin gain first; negative margin delta at bottom
        net_pay_changes.sort(key=lambda x: x["margin_delta"], reverse=True)

    if net_pay_changes:
        warnings.append({
            "severity": "info",
            "title": "Net Pay Changes Detected",
            "description": f"{len(net_pay_changes)} routes changed vs prior batch — review margin impact below",
            "type": "net_pay_change",
            "count": len(net_pay_changes),
            "rides": net_pay_changes,
        })

    # Fetch emails for all drivers in this batch
    all_person_ids = [r["person_id"] for r in rows]
    email_map = {
        p.person_id: p.email or ""
        for p in db.query(Person).filter(Person.person_id.in_(all_person_ids)).all()
    } if all_person_ids else {}

    # Format rows for frontend
    drivers_out = []
    withheld_out = []
    for r in rows:
        entry = {
            "id": r["person_id"],
            "name": r["person"],
            "pay_code": r["code"],
            "email": email_map.get(r["person_id"], ""),
            "days": r["days"],
            "rides": r["rides"],
            "miles": round(float(r["miles"] or 0), 1),
            "net_pay": r["net_pay"],
            "partner_pays": r["partner_pays"],
            "driver_pay": r["driver_pay"],
            "deduction": r["deduction"],
            "carried_over": r["from_last_period"],
            "pay_this_period": r["pay_this_period"],
            "status": "withheld" if r["withheld"] else "paid",
            "withheld_amount": r["withheld_amount"],
            "force_pay_override": r["person_id"] in override_ids,
            "manual_withhold_note": manual_withhold_map.get(r["person_id"]),
            "missing_paycheck_code": r.get("missing_paycheck_code", False),
            "balance_source": r.get("balance_source"),
        }
        if r["withheld"]:
            withheld_out.append(entry)
        else:
            drivers_out.append(entry)

    total_withheld = sum(r["withheld_amount"] for r in rows if r["withheld"])

    return JSONResponse({
        "batch_id": batch_id,
        "company": _display_company(batch.company_name or ""),
        "period_start": batch.period_start.isoformat() if batch.period_start else None,
        "period_end": batch.period_end.isoformat() if batch.period_end else None,
        "drivers": drivers_out,
        "withheld": withheld_out,
        "totals": totals,
        "warnings": warnings,
        "stats": {
            "driver_count": len(rows),
            "total_pay": totals["pay_this_period"],
            "withheld_amount": round(total_withheld, 2),
            "withheld_count": len(withheld_out),
        },
    })


# ── Email template (batch-level) ─────────────────────────────────────────────

@router.get("/{batch_id}/email-template")
def workflow_get_email_template(batch_id: int, db: Session = Depends(get_db)):
    """Return the resolved email template for this batch (batch override → default)."""
    from backend.routes.email_templates import get_template
    tmpl = get_template(db, batch_id=batch_id)
    return JSONResponse(tmpl)


@router.post("/{batch_id}/email-template")
async def workflow_save_email_template(batch_id: int, request: Request, db: Session = Depends(get_db)):
    """Save a batch-level email template override (JSON body: {subject, body})."""
    from backend.db.models import EmailTemplate
    data = await request.json()
    subject = data.get("subject", "").strip()
    body = data.get("body", "").strip()
    tmpl = db.query(EmailTemplate).filter(
        EmailTemplate.scope == "batch",
        EmailTemplate.payroll_batch_id == batch_id,
    ).first()
    if tmpl:
        tmpl.subject = subject
        tmpl.body = body
    else:
        tmpl = EmailTemplate(scope="batch", payroll_batch_id=batch_id, subject=subject, body=body)
        db.add(tmpl)
    db.commit()
    return JSONResponse({"ok": True})


# ── Stubs status ─────────────────────────────────────────────────────────────

@router.get("/{batch_id}/stubs-status")
def workflow_stubs_status(batch_id: int, db: Session = Depends(get_db)):
    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    # All drivers in this batch — use a subquery to deduplicate by person_id
    # rather than .distinct() on the full Person object, which fails on PostgreSQL
    # json columns (cc_compliance, firstalt_compliance have no equality operator).
    person_ids_in_batch = (
        db.query(Ride.person_id)
        .filter(Ride.payroll_batch_id == batch_id)
        .distinct()
    )
    drivers = (
        db.query(Person)
        .filter(Person.person_id.in_(person_ids_in_batch))
        .all()
    )

    # Email send logs for this batch
    logs = (
        db.query(EmailSendLog)
        .filter(EmailSendLog.payroll_batch_id == batch_id)
        .all()
    )
    log_map = {log.person_id: log for log in logs}

    # Withheld drivers (carried-over balance > 0) — don't send them paystubs
    from backend.db.models import DriverBalance
    withheld_ids = {
        b.person_id
        for b in db.query(DriverBalance).filter(
            DriverBalance.payroll_batch_id == batch_id,
            DriverBalance.carried_over > 0,
        ).all()
    }

    results = []
    counts = {"sent": 0, "failed": 0, "no_email": 0, "withheld": 0, "pending": 0}

    for person in drivers:
        log = log_map.get(person.person_id)
        if person.person_id in withheld_ids:
            status = "withheld"
            counts["withheld"] += 1
        elif not person.email:
            status = "no_email"
            counts["no_email"] += 1
        elif log and log.status == "sent":
            status = "sent"
            counts["sent"] += 1
        elif log and log.status == "failed":
            status = "failed"
            counts["failed"] += 1
        else:
            status = "pending"
            counts["pending"] += 1

        results.append({
            "person_id": person.person_id,
            "name": person.full_name,
            "email": person.email,
            "status": status,
            "error": log.error_message if log and log.status == "failed" else None,
            "sent_at": log.sent_at.isoformat() if log and log.sent_at else None,
        })

    return JSONResponse({
        "batch_id": batch_id,
        "drivers": results,
        "counts": counts,
        "total": len(results),
    })


# ── Batch summary (JSON) ─────────────────────────────────────────────────────

@router.get("/{batch_id}/batch-summary")
def workflow_batch_summary(batch_id: int, db: Session = Depends(get_db)):
    """Return a structured JSON summary for the frontend summary page."""
    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    from backend.db.models import DriverBalance

    week_num = (
        db.query(func.count(PayrollBatch.payroll_batch_id))
        .filter(
            PayrollBatch.source == batch.source,
            PayrollBatch.period_start <= batch.period_start,
        )
        .scalar() or 1
    )
    week_label = f"Week {week_num}"

    # Per-driver ride stats
    driver_rows = (
        db.query(
            Ride.person_id,
            Person.full_name,
            Person.paycheck_code,
            func.count(Ride.ride_id).label("rides"),
            func.coalesce(func.sum(Ride.miles), 0).label("miles"),
            func.coalesce(func.sum(Ride.net_pay), 0).label("partner_paid"),
            func.coalesce(func.sum(Ride.z_rate), 0).label("driver_pay"),
            func.coalesce(func.sum(Ride.deduction), 0).label("deduction"),
        )
        .join(Person, Person.person_id == Ride.person_id)
        .filter(Ride.payroll_batch_id == batch_id)
        .group_by(Ride.person_id, Person.full_name, Person.paycheck_code)
        .order_by(Person.full_name)
        .all()
    )

    # Balance records for this batch (withheld / carried over)
    balance_records = (
        db.query(DriverBalance)
        .filter(DriverBalance.payroll_batch_id == batch_id)
        .all()
    )
    balance_map = {b.person_id: float(b.carried_over or 0) for b in balance_records}

    drivers_out = []
    totals = {
        "rides": 0,
        "miles": 0.0,
        "partner_paid": 0.0,
        "driver_cost": 0.0,
        "withheld": 0.0,
        "payout": 0.0,
        "margin": 0.0,
    }

    for row in driver_rows:
        carried_over = balance_map.get(row.person_id, 0.0)
        is_withheld = carried_over > 0
        partner_paid = round(float(row.partner_paid), 2)
        driver_pay = round(float(row.driver_pay), 2)
        miles = round(float(row.miles), 3)
        deduction = round(float(row.deduction), 2)
        paid_this_period = 0.0 if is_withheld else driver_pay

        drivers_out.append({
            "person_id": row.person_id,
            "name": row.full_name,
            "pay_code": row.paycheck_code or "",
            "rides": int(row.rides),
            "miles": miles,
            "partner_paid": partner_paid,
            "driver_pay": driver_pay,
            "deduction": deduction,
            "withheld_amount": carried_over if is_withheld else 0.0,
            "paid_this_period": paid_this_period,
            "is_withheld": is_withheld,
        })

        totals["rides"] += int(row.rides)
        totals["miles"] = round(totals["miles"] + miles, 3)
        totals["partner_paid"] = round(totals["partner_paid"] + partner_paid, 2)
        totals["driver_cost"] = round(totals["driver_cost"] + driver_pay, 2)
        totals["withheld"] = round(totals["withheld"] + (carried_over if is_withheld else 0.0), 2)
        totals["payout"] = round(totals["payout"] + paid_this_period, 2)
        totals["margin"] = round(totals["partner_paid"] - totals["driver_cost"], 2)

    return JSONResponse({
        "batch": {
            "id": batch_id,
            "company": _display_company(batch.company_name or ""),
            "week_label": week_label,
            "period_start": batch.period_start.isoformat() if batch.period_start else None,
            "period_end": batch.period_end.isoformat() if batch.period_end else None,
            "batch_ref": batch.batch_ref,
            "source": batch.source,
            "status": batch.status,
            "finalized_at": batch.finalized_at.isoformat() if batch.finalized_at else None,
        },
        "totals": totals,
        "drivers": drivers_out,
    })


# ── Excel column spec (mirrors mom's FA_Summary_paroll.xlsx exactly) ─────────
# Col: A=Driver Name, B=Pay Code, C=Rides, D=Miles, E=Partner Pays,
#      F=Driver Pay, G=Deduction, H=Withheld (Y/N), I=Carried Over, J=Paid This Period
_MOM_HEADERS = [
    "Driver Name", "Pay Code", "Rides", "Miles",
    "Partner Pays", "Driver Pay", "Deduction",
    "Withheld (Y/N)", "Carried Over", "Paid This Period",
]
# Column widths in characters — taken directly from mom's file
_MOM_COL_WIDTHS = [36.0, 12.0, 12.16, 13.16, 13.16, 10.16, 9.33, 8.66, 10.83, 13.33]
# Money column indices (1-based): E=5 F=6 G=7 I=9 J=10
_MONEY_COLS = {5, 6, 7, 9, 10}
_MONEY_FMT = '"$"#,##0.00'
# Fixed colors from mom's file — use full ARGB (8 chars, alpha=FF) so openpyxl
# writes the correct fgColor and the value round-trips as expected.
# Values verified by openpyxl introspection of Prod_SP_Acumen International_04032026 (2) (3).xlsx
_HEADER_FILL_HEX        = "FF0F1729"   # dark navy    — column header row (Driver Name / Pay Code / …)
_TOTALS_FILL_HEX        = "FF1F4E78"   # blue         — TOTALS row fill (white font)
_SECTION_FILL_HEX       = "FFA24B10"   # orange-brown — Paychex Flex + Paid on Week headers
_UNPAID_SECTION_FILL_HEX = "FF548235"  # green        — Unpaid on Week header
_PAID_DRIVER_FONT_HEX   = "FFA20000"   # dark red     — driver names in Paid on Week list (unused; kept for compat)
_UNPAID_DRIVER_FONT_HEX = "FF388600"   # dark green   — driver names in Unpaid on Week list (unused; kept for compat)
# Semantic driver-name font colors (applied in main table + both sub-sections)
_FONT_NO_CODE    = "FFC00000"  # red   — driver missing paycheck_code
_FONT_WITHHELD   = "FF548235"  # green — driver withheld this period

# SP PAY SUMMARY columns (matches Acumen xlsx tab exactly)
_SP_PAY_HEADERS = [
    "BATCH ID", "SP COMPANY", "DRIVER CODE", "SERVICE PERIOD",
    "DRIVER NAME", "SERVICE DAYS", "RUNS", "MILES",
    "SPIFF", "GROSS PAY", "DEDUCTION", "NET PAY",
]
# SP ITEMIZED REPORT columns
_SP_ITEMIZED_HEADERS = [
    "BATCH ID", "SP COMPANY", "DRIVER NAME", "DRIVE CODE",
    "DATE", "TRIP CODE", "TRIP NAME", "CANCELLATION REASON",
    "MILES", "SPIFF", "GROSS PAY", "DEDUCTION", "NET PAY",
]

# Release threshold: driver is withheld when combined < $100.
# WARNING: Juhar W9 was released at $76 — meaning at some point the threshold
# was applied differently or a manual override was used. If releasing a driver
# whose this-week earnings are < $100, a WARNING is logged for audit.
_RELEASE_THRESHOLD_DOLLARS: float = 100.0


def _period_label_mom(batch: PayrollBatch) -> str:
    """'MM/DD/YYYY - MM/DD/YYYY - Week N' — mom's row-1 format."""
    ps = getattr(batch, "week_start", None) or getattr(batch, "period_start", None)
    pe = getattr(batch, "week_end", None) or getattr(batch, "period_end", None)
    if ps and pe:
        week_num = _wl(ps, pe)  # returns "Week N"
        return f"{ps.strftime('%m/%d/%Y')} - {pe.strftime('%m/%d/%Y')} - {week_num}"
    if ps:
        return ps.strftime("%m/%d/%Y")
    return f"Batch {batch.payroll_batch_id}"


def _period_label_payroll_summary(batch: PayrollBatch) -> str:
    """'Period: Mon DD, YYYY – Mon DD, YYYY' — Payroll Summary row-2 format."""
    ps = getattr(batch, "week_start", None) or getattr(batch, "period_start", None)
    pe = getattr(batch, "week_end", None) or getattr(batch, "period_end", None)
    if ps and pe:
        return f"Period: {ps.strftime('%b %d, %Y')} – {pe.strftime('%b %d, %Y')}"
    if ps:
        return f"Period: {ps.strftime('%b %d, %Y')}"
    return ""


def _apply_header_style(ws, row_num: int, n_cols: int, fill_hex: str) -> None:
    """Bold white text on solid fill for a header row."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    font = Font(bold=True, color="FFFFFFFF")
    fill = PatternFill("solid", fgColor=fill_hex)
    center = Alignment(horizontal="center")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = center


def _build_sp_pay_summary_tab(ws, batch: PayrollBatch, ride_rows: list) -> None:
    """
    Populate the SP PAY SUMMARY tab.

    Reproduces the structure Brandon emails:
    Row 1: column headers
    Rows 2+: one row per driver with aggregated RUNS, MILES, GROSS PAY, etc.

    ride_rows: list of dicts with keys person, code, rides, miles,
               partner_pays (=gross_pay), deduction, net_pay.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import date as _date

    ws.title = "SP PAY SUMMARY"

    batch_ref = batch.batch_ref or ""
    # Render-time display override: FA/Acumen batches always show "Acumen International"
    # regardless of how company_name is stored in the DB row.
    _raw_company = (batch.company_name or "").lower()
    if _raw_company in ("firstalt", "acumen") or (batch.source or "").lower() == "acumen":
        company = "Acumen International"
    else:
        company = batch.company_name or ""
    ps = getattr(batch, "week_start", None) or getattr(batch, "period_start", None)
    pe = getattr(batch, "week_end", None) or getattr(batch, "period_end", None)
    svc_period = ""
    if ps and pe:
        svc_period = f"{ps.strftime('%m/%d/%Y')} - {pe.strftime('%m/%d/%Y')}"

    # Header row
    ws.append(_SP_PAY_HEADERS)
    _apply_header_style(ws, ws.max_row, len(_SP_PAY_HEADERS), _HEADER_FILL_HEX)

    # Derive service_days per driver from ride data (distinct dates).
    # We aggregate rides in the caller so we receive one row per driver.
    for r in ride_rows:
        gross = round(float(r.get("partner_pays") or 0), 2)
        ded = round(float(r.get("deduction") or 0), 2)
        net = round(gross - ded, 2)
        ws.append([
            batch_ref,
            company,
            "-",
            svc_period,
            r.get("person") or "",
            r.get("service_days") or 0,
            int(r.get("rides") or 0),
            round(float(r.get("miles") or 0), 1),
            0,          # SPIFF — always 0 in practice
            gross,
            ded,
            net,
        ])

    # Column widths (generous for readability)
    _SP_PAY_WIDTHS = [20, 24, 14, 28, 32, 14, 8, 10, 8, 12, 12, 12]
    for idx, w in enumerate(_SP_PAY_WIDTHS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w


def _build_sp_itemized_tab(ws, batch: PayrollBatch, trip_rows: list) -> None:
    """
    Populate the SP ITEMIZED REPORT tab.

    trip_rows: list of dicts with keys:
        person, code, date (date obj), trip_code, trip_name,
        cancellation_reason, miles, gross_pay, deduction, net_pay.
    Sorted by driver name then date already.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    ws.title = "SP ITEMIZED REPORT"

    batch_ref = batch.batch_ref or ""
    # Render-time display override: FA/Acumen batches always show "Acumen International"
    _raw_company = (batch.company_name or "").lower()
    if _raw_company in ("firstalt", "acumen") or (batch.source or "").lower() == "acumen":
        company = "Acumen International"
    else:
        company = batch.company_name or ""

    # Header row
    ws.append(_SP_ITEMIZED_HEADERS)
    _apply_header_style(ws, ws.max_row, len(_SP_ITEMIZED_HEADERS), _HEADER_FILL_HEX)

    for t in trip_rows:
        gross = round(float(t.get("gross_pay") or 0), 2)
        ded = round(float(t.get("deduction") or 0), 2)
        net = round(float(t.get("net_pay") or 0), 2)
        ws.append([
            batch_ref,
            company,
            t.get("person") or "",
            "-",
            t.get("date"),          # openpyxl writes datetime.date as Excel date
            t.get("trip_code") or "",
            t.get("trip_name") or "",
            t.get("cancellation_reason") or None,
            round(float(t.get("miles") or 0), 1),
            0,                      # SPIFF always 0
            gross,
            ded,
            net,
        ])
        # DATE column (col 5) — render as "Apr 13, 2026" to match Brandon's format
        ws.cell(row=ws.max_row, column=5).number_format = "mmm d, yyyy"

    # Column widths
    _SP_ITEM_WIDTHS = [20, 24, 32, 14, 14, 12, 44, 24, 10, 8, 12, 12, 12]
    for idx, w in enumerate(_SP_ITEM_WIDTHS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w


def _build_payroll_summary_tab(
    ws,
    batch: PayrollBatch,
    rows: list,
    totals: dict,
    llc_title: str,
    db=None,
) -> None:
    """
    Populate the Payroll Summary tab.

    Layout (matches W14 Acumen xlsx canonical format):
      R1:  "{LLC name} — Payroll Summary"
      R2:  "Period: Mon DD, YYYY – Mon DD, YYYY"
      R3:  blank
      R4:  column headers (blue fill)
      R5+: per-driver data
      +1:  TOTALS row (orange fill)
      +4 blank rows
      Paychex Flex Amount row — mom keys C; reconciliation note in G
      blank
      Paid on Week section
      per-driver paid lines (Name | Code | $amount)
      Total line
      blank
      Unpaid on Week section
      per-driver withheld lines
      blank row

    WARNING notes about release rule are written to a comment on the withheld
    driver's name cell when their this-week earnings were < $100 but the carry
    released (manual override path).
    """
    import logging
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    ws.title = "Payroll Summary"

    period_str = _period_label_payroll_summary(batch)
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL_HEX)
    totals_font = Font(bold=True, color="FFFFFFFF")
    totals_fill = PatternFill("solid", fgColor=_TOTALS_FILL_HEX)
    section_font = Font(bold=True, color="FFFFFFFF")
    section_fill = PatternFill("solid", fgColor=_SECTION_FILL_HEX)
    paid_driver_font   = Font(color=_PAID_DRIVER_FONT_HEX)
    unpaid_driver_font = Font(color=_UNPAID_DRIVER_FONT_HEX)
    center = Alignment(horizontal="center")

    # R1: title
    ws.append([llc_title])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.row_dimensions[ws.max_row].height = 22.0

    # R2: period
    ws.append([period_str])
    ws.row_dimensions[ws.max_row].height = 16.0

    # R3: blank
    ws.append([])

    # R4: column headers
    ws.append(_MOM_HEADERS)
    hdr_row = ws.max_row
    for col in range(1, len(_MOM_HEADERS) + 1):
        cell = ws.cell(row=hdr_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws.row_dimensions[hdr_row].height = 16.0

    # R5+: data rows
    first_data_row = hdr_row + 1
    for r in rows:
        carried = round(float(r.get("from_last_period") or 0), 2)
        driver_pay = round(float(r.get("driver_pay") or 0), 2)
        withheld = bool(r.get("withheld"))
        pay_this = round(float(r.get("pay_this_period") or 0), 2)

        # Release rule WARNING: if a driver with prior balance is being paid
        # but their THIS-WEEK earnings alone are below threshold, log it.
        if not withheld and carried > 0 and driver_pay < _RELEASE_THRESHOLD_DOLLARS:
            logging.warning(
                "[payroll-excel] Releasing held balance for %s: this-week=$%.2f "
                "(below $%.0f threshold). carried=$%.2f. Verify this is intentional.",
                r.get("person", "?"), driver_pay, _RELEASE_THRESHOLD_DOLLARS, carried,
            )

        ws.append([
            r.get("person") or "",
            r.get("code") or "",
            int(r.get("rides") or 0),
            round(float(r.get("miles") or 0), 3),
            round(float(r.get("partner_pays") or 0), 2),
            driver_pay,
            round(float(r.get("deduction") or 0), 2),
            "Yes" if withheld else "No",
            carried,
            pay_this,
        ])
        data_row = ws.max_row
        for col in (2, 3, 4):
            ws.cell(row=data_row, column=col).alignment = center
        for col in _MONEY_COLS:
            ws.cell(row=data_row, column=col).number_format = _MONEY_FMT
        # Semantic font on driver name (col 1)
        missing_code = bool(r.get("missing_paycheck_code"))
        if missing_code:
            ws.cell(row=data_row, column=1).font = Font(color=_FONT_NO_CODE)
        elif withheld:
            ws.cell(row=data_row, column=1).font = Font(color=_FONT_WITHHELD)

    last_data_row = ws.max_row

    # TOTALS row
    ws.append([
        "TOTALS", "",
        totals.get("rides", 0),
        round(float(totals.get("miles") or 0), 3),
        round(float(totals.get("partner_pays") or 0), 2),
        round(float(totals.get("driver_pay") or 0), 2),
        round(float(totals.get("deduction") or 0), 2),
        "",
        round(float(totals.get("carried_over") or 0), 2),
        round(float(totals.get("pay_this_period") or 0), 2),
    ])
    totals_row_num = ws.max_row
    ws.row_dimensions[totals_row_num].height = 16.0
    for col in range(1, len(_MOM_HEADERS) + 1):
        cell = ws.cell(row=totals_row_num, column=col)
        cell.font = totals_font
        cell.fill = totals_fill
    for col in _MONEY_COLS:
        ws.cell(row=totals_row_num, column=col).number_format = _MONEY_FMT

    # 4 blank rows
    for _ in range(4):
        ws.append([])

    # Paychex Flex Amount row
    # Mom keys the Paychex total in column C.
    # Column G shows difference: J(total paid) - C(paychex amount) — should be 0 when matched.
    paychex_row = ws.max_row + 1
    paid_total = round(float(totals.get("pay_this_period") or 0), 2)
    ws.append([
        "Paychex Flex Amount",
        None,
        None,           # ← mom keys in the amount here (col C)
        None,
        None,
        None,
        f"=J{totals_row_num}-C{paychex_row}",   # reconciliation check: should = 0
        None,
        None,
        paid_total,     # col J = Z-Pay total paid this period (read-only reference)
    ])
    ws.cell(row=paychex_row, column=1).font = section_font
    ws.cell(row=paychex_row, column=1).fill = section_fill
    ws.cell(row=paychex_row, column=10).number_format = _MONEY_FMT

    # blank
    ws.append([])

    # Paid on Week section
    # Only lists drivers whose previously-withheld balance is being released
    # this week (i.e. their person_id appears in payroll_withheld_override for
    # this batch_id).  One Excel row per driver_balance entry — a driver can
    # appear multiple times if they have multiple held-balance records (e.g.
    # Nuraynie's 4 rows across W9-W14).
    ws.append(["Paid on Week"])
    ws.cell(row=ws.max_row, column=1).font = section_font
    ws.cell(row=ws.max_row, column=1).fill = section_fill

    paid_on_week_entries = []  # list of (person_name, pay_code, amount)
    if db is not None:
        from sqlalchemy import text as _paidwk_text
        # Drivers whose withheld balance is being released this batch.
        override_rows = db.execute(
            _paidwk_text(
                "SELECT person_id FROM payroll_withheld_override WHERE batch_id = :b"
            ),
            {"b": batch.payroll_batch_id},
        ).fetchall()
        released_person_ids = [row[0] for row in override_rows]

        if released_person_ids:
            is_acumen_batch = (batch.source or "").lower() == "acumen"
            for pid in released_person_ids:
                person = db.query(Person).filter(Person.person_id == pid).first()
                if person is None:
                    continue
                person_name = person.full_name or ""
                pay_code = (
                    person.paycheck_code if is_acumen_batch else person.paycheck_code_maz
                ) or ""
                # All driver_balance rows for this person across ALL batches
                # represent held amounts being released now.
                balance_rows = (
                    db.query(DriverBalance)
                    .filter(DriverBalance.person_id == pid)
                    .order_by(DriverBalance.payroll_batch_id.asc())
                    .all()
                )
                for bal in balance_rows:
                    amount = round(float(bal.carried_over or 0), 2)
                    paid_on_week_entries.append((person_name, pay_code, amount))

    for person_name, pay_code, amount in paid_on_week_entries:
        ws.append([person_name, pay_code, amount])
        ws.cell(row=ws.max_row, column=3).number_format = _MONEY_FMT
        # Semantic font on driver name in Paid on Week sub-section
        if not pay_code:
            ws.cell(row=ws.max_row, column=1).font = Font(color=_FONT_NO_CODE)

    # Total row (always present; $0 when no releases this week)
    paid_sum = round(sum(amt for _, _, amt in paid_on_week_entries), 2)
    ws.append(["Total", None, paid_sum])
    ws.cell(row=ws.max_row, column=1).font = section_font
    ws.cell(row=ws.max_row, column=3).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=3).number_format = _MONEY_FMT

    # blank
    ws.append([])

    # Unpaid on Week section
    ws.append(["Unpaid on Week"])
    ws.cell(row=ws.max_row, column=1).font = section_font
    ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor=_UNPAID_SECTION_FILL_HEX)
    withheld_rows = [r for r in rows if r.get("withheld")]
    for r in withheld_rows:
        withheld_amt = round(float(r.get("withheld_amount") or 0), 2)
        ws.append([r.get("person") or "", r.get("code") or "", withheld_amt])
        ws.cell(row=ws.max_row, column=3).number_format = _MONEY_FMT
        # Semantic font on driver name in Unpaid sub-section
        _missing = bool(r.get("missing_paycheck_code"))
        if _missing:
            ws.cell(row=ws.max_row, column=1).font = Font(color=_FONT_NO_CODE)
        else:
            ws.cell(row=ws.max_row, column=1).font = Font(color=_FONT_WITHHELD)
    # Withheld total
    ws.append([])

    # Column widths — exact match to mom's file
    for col_idx, width in enumerate(_MOM_COL_WIDTHS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width


# ── Export Excel ─────────────────────────────────────────────────────────────

@router.get("/{batch_id}/export-excel")
def workflow_export_excel(batch_id: int, db: Session = Depends(get_db)):
    """
    Return a multi-tab .xlsx payroll download that matches the Brandon email format.

    Acumen (FA) batches → 3 tabs:
      1. SP PAY SUMMARY      — per-driver aggregated roll-up (regenerated from DB)
      2. SP ITEMIZED REPORT  — per-trip detail (regenerated from DB)
      3. Payroll Summary     — Z-Pay driver-pay summary with Paychex reconciliation row

    Maz (ED) batches → 1 tab:
      1. Payroll Summary     — same format, titled "Maz — Payroll Summary"

    The under-$100 withhold rule is handled by _build_summary (already correct in DB).
    Release rule: when a withheld driver's this-week earnings are < $100 but they are
    being paid (manual override), a WARNING is logged — see _build_payroll_summary_tab.

    TODO (confirm with Malik before building): add a second tab to Maz xlsx mirroring
    SP ITEMIZED REPORT structure generated from parsed ED ride data.
    """
    import io
    import openpyxl
    from fastapi.responses import StreamingResponse

    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    # _build_summary walks all prior batches so from_last_period (carried) is correct.
    # Pass override/manual-withhold IDs so the export exactly matches the approved
    # payroll state — otherwise a manual-withhold driver could appear as "paid" in
    # the export with carry-forward = 0 instead of their correct held amount.
    # Load force-pay and manual-withhold overrides so the export matches the
    # approved payroll state exactly.  Without these, a manually-withheld driver
    # could appear as "paid" with carry-forward = 0 instead of their accumulated
    # held amount.  Wrapped in try/except so SQLite test environments (which may
    # not have created these tables) don't crash.
    _override_ids: set[int] | None = None
    _manual_withhold_ids: set[int] | None = None
    try:
        from sqlalchemy import text as _sql_text
        _override_rows = db.execute(
            _sql_text("SELECT person_id FROM payroll_withheld_override WHERE batch_id = :b"),
            {"b": batch_id},
        ).fetchall()
        _override_ids = {r[0] for r in _override_rows} or None
        _manual_rows = db.execute(
            _sql_text("SELECT person_id FROM payroll_manual_withhold"),
        ).fetchall()
        _manual_withhold_ids = {r[0] for r in _manual_rows} or None
    except Exception:
        # Tables may not exist (e.g. after DB restore or in test environments).
        # Roll back the aborted transaction so subsequent queries on this session
        # don't fail with "current transaction is aborted".
        db.rollback()

    data = _build_summary(
        db,
        batch_id=batch_id,
        override_ids=_override_ids,
        manual_withhold_ids=_manual_withhold_ids,
    )
    rows = data["rows"]
    totals = data["totals"]

    is_acumen = (batch.source or "").lower() == "acumen"

    if is_acumen:
        llc_title = "FirstAlt — Payroll Summary"
        # Query per-trip data for SP ITEMIZED REPORT tab
        # cancellation_reason is NOT a DB column — derive from z_rate_source='canceled_trip'
        trip_rows_raw = (
            db.query(
                Person.full_name.label("person"),
                Person.paycheck_code.label("code"),
                Ride.ride_start_ts,
                Ride.source_ref,
                Ride.service_name,
                Ride.z_rate_source,
                Ride.miles,
                Ride.gross_pay,
                Ride.deduction,
                Ride.net_pay,
            )
            .join(Person, Person.person_id == Ride.person_id)
            .filter(Ride.payroll_batch_id == batch_id)
            .order_by(Person.full_name.asc(), Ride.ride_start_ts.asc())
            .all()
        )

        from sqlalchemy import func as _func
        # Service days per driver (distinct calendar dates).
        # Use func.date() rather than cast(..., Date) — more reliable for
        # timestamptz columns in Postgres and avoids type-resolution issues
        # that caused the svc_days_map to silently return 0 for every driver.
        svc_days_raw = (
            db.query(
                Ride.person_id,
                _func.count(_func.distinct(_func.date(Ride.ride_start_ts))).label("svc_days"),
            )
            .filter(Ride.payroll_batch_id == batch_id)
            .group_by(Ride.person_id)
            .all()
        )
        svc_days_map = {r.person_id: int(r.svc_days) for r in svc_days_raw}

        # Enrich summary rows with service_days for SP PAY SUMMARY tab
        rows_with_svc = []
        pid_map = {r["person"]: r for r in rows}  # name → summary row
        for r in rows:
            rows_with_svc.append({**r, "service_days": 0})  # default; will patch below
        # Rebuild with person_id-keyed svc_days
        for r in rows:
            # rows from _build_summary don't have person_id re-mapped to svc_days_map
            # We need to look up by person_id which is in r["person_id"]
            r["service_days"] = svc_days_map.get(r.get("person_id"), 0)

        trip_rows = []
        for t in trip_rows_raw:
            trip_date = None
            if t.ride_start_ts:
                trip_date = t.ride_start_ts.date()
            trip_rows.append({
                "person": t.person,
                "code": t.code or "-",
                "date": trip_date,
                "trip_code": t.source_ref or "",
                "trip_name": t.service_name or "",
                # canceled_trip z_rate_source signals FA still invoiced but driver got $0
                "cancellation_reason": "Canceled" if t.z_rate_source == "canceled_trip" else None,
                "miles": float(t.miles or 0),
                "gross_pay": float(t.gross_pay or 0),
                "deduction": float(t.deduction or 0),
                "net_pay": float(t.net_pay or 0),
            })

        if batch.sp_file_bytes:
            # ── Passthrough path ────────────────────────────────────────────
            # Load the original FA xlsx Brandon emailed so Tabs 1 & 2 are
            # byte-faithful.  Drop any stale Z-Pay payroll tab (could be from
            # a previous export or legacy generation), then append the fresh
            # Payroll Summary tab built from live DB data.
            wb = openpyxl.load_workbook(io.BytesIO(batch.sp_file_bytes))
            for stale in ("Payroll", "Payroll  ", "Payroll Summary"):
                if stale in wb.sheetnames:
                    del wb[stale]
            ws_ps = wb.create_sheet("Payroll Summary")
            _build_payroll_summary_tab(ws_ps, batch, rows, totals, llc_title, db=db)
        else:
            # ── Legacy fallback ─────────────────────────────────────────────
            # No original file stored — regenerate all 3 tabs from DB data.
            # Keep this path until sp_file_bytes is backfilled for older batches.
            wb = openpyxl.Workbook()
            # Tab 1: SP PAY SUMMARY
            ws1 = wb.active
            _build_sp_pay_summary_tab(ws1, batch, rows)
            # Tab 2: SP ITEMIZED REPORT
            ws2 = wb.create_sheet("SP ITEMIZED REPORT")
            _build_sp_itemized_tab(ws2, batch, trip_rows)
            # Tab 3: Mom's exact payroll tab ("Payroll  " with two trailing spaces)
            ws3 = wb.create_sheet()  # title set inside _build_mom_payroll_tab
            _build_mom_payroll_tab(ws3, rows, totals)

    else:
        # Maz / EverDriven — single tab only
        llc_title = "Maz — Payroll Summary"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        _build_payroll_summary_tab(ws1, batch, rows, totals, llc_title)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Stamp export time so the workflow gate advances.
    # Reused for the Excel export — the workflow stage advances on this timestamp.
    from datetime import datetime, timezone as _tz
    if not batch.paychex_exported_at:
        batch.paychex_exported_at = datetime.now(_tz.utc)
        db.commit()

    filename = f"payroll_{_safe_slug(batch.company_name or 'batch')}_{_fmt_period(batch)}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Mom's exact Tab 3 format ─────────────────────────────────────────────────
# Reproduces the "Payroll  " (TWO trailing spaces) tab from the canonical
# W14_FA_Master_Payroll.xlsx reference.
#
# 6-column layout:
#   A (Person)           = driver full name
#   B (Code)             = paycheck_code (Paychex Worker ID)
#   C (Payroll)          = driver pay this period (z_rate sum)
#   D (Unpaid / Pending) = withheld/carried amount (blank if driver is paid)
#   E (To Paid )         = blank — mom fills from Paychex; SUM formula in totals
#   F (Total )           = =C-D+E per row

def _build_mom_payroll_tab(ws, rows: list, totals: dict) -> None:
    """
    Populate *ws* with mom's exact 6-column payroll tab.
    rows/totals come from _build_summary.

    Structure verified against ~/Desktop/W14_FA_Master_Payroll.xlsx:
    - Per-driver col F: =C-D+E formula
    - Total row: SUM formulas for cols C/D/E/F
    - Paychex row: =D-C in col E, =F_total-C in col F
    - Unpaid on Week: =SUM(E_paychex:E_prev) in col E
    - Code col is int (not string) when possible
    - Sheet title has TWO trailing spaces: "Payroll  "

    Uses ws._current_row to track row numbers (not ws.max_row) because
    openpyxl's max_row only counts content-bearing rows and will miscount
    after appending blank rows — causing wrong formula row references.
    """
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    ws.title = "Payroll  "  # TWO trailing spaces — matches mom's reference

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    def _append(row_vals):
        ws.append(row_vals)
        return ws._current_row  # returns the row index just written

    # R1: "Summary"
    r = _append(["Summary"])
    ws.cell(r, 1).font = bold

    # R2: column headers (note trailing spaces on "To Paid " and "Total ")
    r = _append(["Person", "Code", "Payroll", "Unpaid / Pending", "To Paid ", "Total "])
    for col in range(1, 7):
        ws.cell(r, col).font = bold
        ws.cell(r, col).alignment = center

    # R3+: per-driver rows (alphabetical order from _build_summary)
    first_data_row = ws._current_row + 1
    for dr in rows:
        driver_pay = round(float(dr.get("driver_pay") or 0), 2)
        withheld_amt = round(float(dr.get("withheld_amount") or 0), 2)
        is_withheld = bool(dr.get("withheld"))
        code = dr.get("code")
        # Cast code to int when possible to match mom's int format
        try:
            code_val = int(code) if code else None
        except (TypeError, ValueError):
            code_val = code or None
        data_row = ws._current_row + 1
        _append([
            dr.get("person") or "",
            code_val,
            driver_pay if driver_pay else None,
            withheld_amt if is_withheld else None,
            None,
            f"=C{data_row}-D{data_row}+E{data_row}",
        ])
    last_data_row = ws._current_row

    _append([])  # blank row

    # Totals row with SUM formulas
    total_row = _append([
        "Total",
        None,
        f"=SUM(C{first_data_row}:C{last_data_row})",
        f"=SUM(D{first_data_row}:D{last_data_row})",
        f"=SUM(E{first_data_row}:E{last_data_row})",
        f"=SUM(F{first_data_row}:F{last_data_row})",
    ])
    ws.cell(total_row, 1).font = bold

    _append([])  # blank
    _append([])  # blank

    # Paychex Flex Amount row — mom enters Paychex total into col C
    paychex_row = ws._current_row + 1
    _append([
        "Paychex Flex Amound ",   # matches mom's exact (typo) spelling
        None,
        None,
        None,
        f"=D{paychex_row}-C{paychex_row}",
        f"=F{total_row}-C{paychex_row}",
    ])
    ws.cell(paychex_row, 1).font = bold

    _append([])  # blank

    # "Unpaid on Week"
    unpaid_label_row = ws._current_row + 1
    _append([
        "Unpaid on Week",
        None,
        None,
        None,
        f"=SUM(E{paychex_row}:E{unpaid_label_row - 1})",
        None,
    ])
    ws.cell(unpaid_label_row, 1).font = bold

    # Withheld drivers sub-list (Person | Code | Payroll only — cols D/E/F stay None)
    for dr in rows:
        if not dr.get("withheld"):
            continue
        code = dr.get("code")
        try:
            code_val = int(code) if code else None
        except (TypeError, ValueError):
            code_val = code or None
        _append([
            dr.get("person") or "",
            code_val,
            round(float(dr.get("driver_pay") or 0), 2) or None,
        ])

    _append([])  # blank
    _append([])  # blank

    # "Paid on Weeks" label
    paid_row = _append(["Paid on Weeks"])
    ws.cell(paid_row, 1).font = bold

    # ~14 filler empty rows so visual layout matches mom's W14 reference
    for _ in range(14):
        _append([])

    # Final "Total" footer (text only — no values)
    final_total_row = _append(["Total", None, None, None, None, None])
    ws.cell(final_total_row, 1).font = bold

    # Column widths matching mom's file
    widths = [36.0, 10.0, 12.0, 16.0, 12.0, 12.0]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


# ── Backward-compat shim ─────────────────────────────────────────────────────
# The old single-sheet export helper is preserved as a thin wrapper around the
# new _build_payroll_summary_tab so existing tests and any callers that imported
# _build_mom_excel directly continue to work without modification.

def _build_mom_excel(wb, rows: list, totals: dict, period_label: str) -> None:
    """
    Backward-compat wrapper.  Writes the old single "Payroll_Summary" sheet
    (underscore — legacy tab name) to *wb*.  Use _build_payroll_summary_tab for
    new code that needs the full layout with Paychex row and Paid/Unpaid sections.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.active
    ws.title = "Payroll_Summary"

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL_HEX)
    totals_font = Font(bold=True, color="FFFFFFFF")
    totals_fill = PatternFill("solid", fgColor=_TOTALS_FILL_HEX)
    center = Alignment(horizontal="center")

    ws.append([period_label])
    ws.row_dimensions[1].height = 22.0

    ws.append(_MOM_HEADERS)
    hdr_row = ws.max_row
    for col_idx in range(1, len(_MOM_HEADERS) + 1):
        cell = ws.cell(row=hdr_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    first_data_row = hdr_row + 1
    for r in rows:
        carried = round(float(r.get("from_last_period") or 0), 2)
        ws.append([
            r.get("person") or "",
            r.get("code") or "",
            int(r.get("rides") or 0),
            round(float(r.get("miles") or 0), 3),
            round(float(r.get("partner_pays") or 0), 2),
            round(float(r.get("driver_pay") or 0), 2),
            round(float(r.get("deduction") or 0), 2),
            "Yes" if r.get("withheld") else "No",
            carried,
            round(float(r.get("pay_this_period") or 0), 2),
        ])
        data_row = ws.max_row
        for col_idx in (2, 3, 4):
            ws.cell(row=data_row, column=col_idx).alignment = center
        for col_idx in _MONEY_COLS:
            ws.cell(row=data_row, column=col_idx).number_format = _MONEY_FMT

    last_data_row = ws.max_row
    totals_row_values = [
        "TOTALS", "",
        totals.get("rides", 0),
        round(float(totals.get("miles") or 0), 3),
        round(float(totals.get("partner_pays") or 0), 2),
        round(float(totals.get("driver_pay") or 0), 2),
        round(float(totals.get("deduction") or 0), 2),
        "",
        round(float(totals.get("carried_over") or 0), 2),
        round(float(totals.get("pay_this_period") or 0), 2),
    ]
    ws.append(totals_row_values)
    totals_row_num = ws.max_row
    ws.row_dimensions[totals_row_num].height = 16.0
    for col_idx in range(1, len(_MOM_HEADERS) + 1):
        cell = ws.cell(row=totals_row_num, column=col_idx)
        cell.font = totals_font
        cell.fill = totals_fill
    for col_idx in _MONEY_COLS:
        ws.cell(row=totals_row_num, column=col_idx).number_format = _MONEY_FMT

    for col_idx, width in enumerate(_MOM_COL_WIDTHS, start=1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width



# ── Generate Paychex (explicit pre-email signal) ─────────────────────────────

@router.post("/{batch_id}/generate-paychex")
def workflow_generate_paychex(batch_id: int, db: Session = Depends(get_db)):
    """Generate and validate the Paychex Excel in-memory, then stamp paychex_exported_at.

    This endpoint is the explicit "Paychex confirmed" action that must succeed
    before paystub emails are sent.  The UI calls it at the start of the
    stubs_sending stage so the operator sees a clear green signal before any
    email fires.

    Order of operations:
      1. Maz/EverDriven batches skip generation (mom submits manually) → ok + skipped=True.
      2. If paychex_exported_at is already set → idempotent, return ok + existing timestamp.
      3. Build the Paychex Excel in-memory and validate it is non-empty.
         Any exception → 500, no stamp written.
      4. Stamp paychex_exported_at and commit.
      5. Return ok=True + generated_at timestamp.
    """
    import io as _io
    import logging as _logging
    from datetime import datetime, timezone as _tz

    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    source = (batch.source or "").lower()
    is_maz = source == "maz"

    # Maz batches — mom enters Paychex manually, nothing to generate.
    if is_maz:
        return JSONResponse({"ok": True, "skipped": True, "reason": "EverDriven — Paychex submitted manually"})

    # Already generated — idempotent.
    if batch.paychex_exported_at:
        return JSONResponse({
            "ok": True,
            "generated_at": batch.paychex_exported_at.isoformat(),
            "already_generated": True,
        })

    # Build in-memory to validate nothing will blow up when the real export runs.
    try:
        import openpyxl as _openpyxl
        from sqlalchemy import text as _sql_text
        _ovr_ids: set[int] | None = None
        _mw_ids: set[int] | None = None
        try:
            _ovr_rows = db.execute(
                _sql_text("SELECT person_id FROM payroll_withheld_override WHERE batch_id = :b"),
                {"b": batch_id},
            ).fetchall()
            _ovr_ids = {r[0] for r in _ovr_rows} or None
            _mw_rows = db.execute(
                _sql_text("SELECT person_id FROM payroll_manual_withhold"),
            ).fetchall()
            _mw_ids = {r[0] for r in _mw_rows} or None
        except Exception:
            pass  # override tables absent in test environments — graceful

        paychex_data = _build_summary(db, batch_id=batch_id, override_ids=_ovr_ids, manual_withhold_ids=_mw_ids)
        paid_count = sum(1 for r in paychex_data["rows"] if not r["withheld"] and r["pay_this_period"] > 0)

        if paid_count > 0:
            wb = _openpyxl.Workbook()
            ws_check = wb.active
            _build_payroll_summary_tab(
                ws_check, batch, paychex_data["rows"], paychex_data["totals"],
                "FirstAlt — Payroll Summary",
            )
            buf = _io.BytesIO()
            wb.save(buf)
            if buf.tell() == 0:
                return JSONResponse(
                    {"error": "Paychex Excel generation produced an empty file."},
                    status_code=500,
                )
        # paid_count == 0 → all withheld, valid — still stamp so the gate advances.
    except Exception as exc:
        _logging.getLogger("zpay.workflow").error(
            "generate-paychex pre-flight failed for batch %s: %s", batch_id, exc,
        )
        return JSONResponse(
            {"error": f"Paychex generation failed: {exc}"},
            status_code=500,
        )

    # Stamp and commit — only reached on success.
    now = datetime.now(_tz.utc)
    batch.paychex_exported_at = now
    db.commit()

    return JSONResponse({"ok": True, "generated_at": now.isoformat()})


# ── Export PDF ───────────────────────────────────────────────────────────────

@router.get("/{batch_id}/export-pdf")
def workflow_export_pdf(batch_id: int, db: Session = Depends(get_db)):
    """Return a landscape PDF payroll summary for the batch."""
    import io
    from fastapi.responses import StreamingResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    week_num = (
        db.query(func.count(PayrollBatch.payroll_batch_id))
        .filter(
            PayrollBatch.source == batch.source,
            PayrollBatch.period_start <= batch.period_start,
        )
        .scalar() or 1
    )
    week_label = f"Week {week_num}"
    company = _display_company(batch.company_name or "")

    # Company-aware PDF colors
    _co = (batch.company_name or "").lower()
    if "acumen" in _co or "first" in _co:
        pdf_header_color = "#4A1525"
        pdf_totals_color = "#9B2C3D"
    elif "maz" in _co or "ever" in _co:
        pdf_header_color = "#0F1D3A"
        pdf_totals_color = "#1E3A6E"
    else:
        pdf_header_color = "#2563EB"
        pdf_totals_color = "#DBEAFE"
    period_start = batch.period_start.isoformat() if batch.period_start else "—"
    period_end = batch.period_end.isoformat() if batch.period_end else "—"
    batch_ref = batch.batch_ref or f"#{batch_id}"

    # ── Per-driver summary via _build_summary ─────────────────────────────────
    # _build_summary walks all prior batches per person, so from_last_period is
    # correct even when the DriverBalance row lives on an older batch or has
    # been cleared (auto-save deletes it once the driver is paid this period).
    data = _build_summary(db, batch_id=batch_id)
    rows = data["rows"]
    totals = data["totals"]

    # ── Build PDF ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], alignment=TA_CENTER, fontSize=16)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], alignment=TA_CENTER, fontSize=10)
    footer_style = ParagraphStyle("footer", parent=styles["Normal"], alignment=TA_RIGHT, fontSize=8, textColor=colors.grey)

    story = []

    # Header block
    story.append(Paragraph(company, title_style))
    story.append(Paragraph(f"{week_label} &nbsp;&nbsp;|&nbsp;&nbsp; {period_start} – {period_end} &nbsp;&nbsp;|&nbsp;&nbsp; Ref: {batch_ref}", sub_style))
    story.append(Spacer(1, 0.5 * cm))

    # Table data — 7 columns matching the summary PDF column set (minus Pay Code,
    # Miles, Deduction which the workflow PDF has always omitted for compactness)
    col_headers = ["Driver Name", "Rides", "Partner Pays", "Driver Pay", "Carried Over", "Withheld", "Paid This Period"]
    table_data = [col_headers]

    for r in rows:
        carried = r["from_last_period"]
        partner_pays = r["partner_pays"]
        driver_pay = r["driver_pay"]
        withheld_amt = r["withheld_amount"]
        paid = r["pay_this_period"]

        table_data.append([
            r["person"],
            str(r["rides"]),
            f"${partner_pays:,.2f}",
            f"${driver_pay:,.2f}",
            f"${carried:,.2f}" if carried else "—",
            f"${withheld_amt:,.2f}" if r["withheld"] else "—",
            f"${paid:,.2f}",
        ])

    # Totals row
    table_data.append([
        "TOTALS",
        str(totals["rides"]),
        f"${totals['partner_pays']:,.2f}",
        f"${totals['driver_pay']:,.2f}",
        f"${totals['carried_over']:,.2f}",
        "—",
        f"${totals['pay_this_period']:,.2f}",
    ])

    page_w = page_size[0] - 3 * cm
    col_widths = [
        page_w * 0.26, page_w * 0.07, page_w * 0.14,
        page_w * 0.14, page_w * 0.13, page_w * 0.12, page_w * 0.14,
    ]

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(pdf_header_color)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor(pdf_totals_color)),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F1F5F9")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(tbl)

    story.append(Spacer(1, 0.5 * cm))
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    story.append(Paragraph(f"Generated: {generated_at}", footer_style))

    doc.build(story)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{_safe_slug(batch.company_name or "payroll")}_{_fmt_period(batch)}.pdf"',
        },
    )

# ── Preview stub (dry-run email preview) ─────────────────────────────────────

@router.get("/{batch_id}/preview-stub/{person_id}")
def workflow_preview_stub(batch_id: int, person_id: int, db: Session = Depends(get_db)):
    """Return a JSON preview of the email that would be sent, without sending it."""
    from backend.routes.email import _build_payweek
    from backend.routes.email_templates import get_template, render_template
    from backend.services.email_service import _body_to_html

    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    person = db.query(Person).filter(Person.person_id == person_id).first()
    if not person:
        return JSONResponse({"error": "Person not found"}, status_code=404)

    rides = (
        db.query(Ride)
        .filter(Ride.payroll_batch_id == batch_id, Ride.person_id == person_id)
        .order_by(Ride.ride_start_ts.asc())
        .all()
    )

    payweek = _build_payweek(batch)
    company = batch.company_name or ""
    total_pay = sum(float(r.z_rate or 0) for r in rides)

    tmpl = get_template(db, person_id=person_id, batch_id=batch_id)
    from backend.routes.email_templates import build_signature_html
    ctx = {
        "driver_name": person.full_name,
        "first_name": (person.full_name.split() or ["Driver"])[0],
        "week_start": batch.week_start.isoformat() if batch.week_start else payweek,
        "week_end": batch.week_end.isoformat() if batch.week_end else payweek,
        "total_pay": f"{total_pay:.2f}",
        "ride_count": str(len(rides)),
        "company_name": company,
        "signature_html": build_signature_html(company),
    }
    subject, body = render_template(tmpl, ctx)
    html_email = _body_to_html(body, company=company, subject=subject)

    return JSONResponse({
        "subject": subject,
        "body_html": html_email,
        "driver_name": person.full_name,
        "email": person.email or "",
    })


# ── SMTP connectivity test ────────────────────────────────────────────────────

@router.get("/smtp-test")
def smtp_test():
    from backend.services.email_service import _IPv4SMTP
    import smtplib, socket
    results = {}
    # Test IPv4-forced port 587
    try:
        s = _IPv4SMTP("smtp.gmail.com", 587, timeout=10)
        s.ehlo()
        s.starttls()
        s.ehlo()
        s.quit()
        results["587_ipv4"] = "OK"
    except Exception as e:
        results["587_ipv4"] = str(e)
    # Test default port 587
    try:
        s = smtplib.SMTP("smtp.gmail.com", 587, timeout=10)
        s.ehlo(); s.starttls(); s.ehlo(); s.quit()
        results["587_default"] = "OK"
    except Exception as e:
        results["587_default"] = str(e)
    try:
        ip = socket.gethostbyname("smtp.gmail.com")
        results["dns_ipv4"] = ip
    except Exception as e:
        results["dns_ipv4"] = str(e)
    return JSONResponse(results)


# ── Send stubs (bulk) ────────────────────────────────────────────────────────

@router.post("/{batch_id}/send-stubs")
def workflow_send_stubs(
    batch_id: int,
    request: Request,
    confirmed_recipient_count: int = Query(..., description="Must match actual driver count — belt+suspenders against accidental fire"),
    test_recipient_override: str | None = Query(None, description="Admin-only: redirect ALL stubs to this email instead of drivers"),
    db: Session = Depends(get_db),
):
    """Send paystubs to all unsent drivers in the batch.

    Belt-and-suspenders gates:
      - confirmed_recipient_count must match the actual pending driver count.
      - test_recipient_override is admin-only; backend validates the calling user's role.
      - Batch must be in export_ready or stubs_sending status.

    Order of operations (critical):
      1. Generate Paychex Excel in-memory and validate it has rows.
         If generation fails, abort — no emails go out.
      2. Stamp paychex_exported_at on the batch so the workflow gate advances.
      3. Send paystubs to drivers (or override recipient for test sends).

    For Maz/EverDriven batches, mom submits Paychex manually, so step 1 is
    skipped — proceed directly to email send.  When all paid drivers are
    withheld (zero-row Paychex file), skip generation but continue to emails.
    """
    import logging as _logging

    # ── Auth gate for test_recipient_override ────────────────────────────────
    if test_recipient_override is not None:
        user = getattr(request.state, "user", None)
        if not user or user.get("role") != "admin":
            return JSONResponse(
                {"error": "test_recipient_override requires admin role"},
                status_code=403,
            )

    # ── Status gate ──────────────────────────────────────────────────────────
    # Re-fetched below after batch lookup, but do a quick check here.
    import io as _io
    from backend.routes.email import _generate_pdf, _build_payweek
    from backend.services.email_service import send_paystub

    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    # ── Fix 2: Paychex Excel generation BEFORE any email send ────────────────
    source = (batch.source or "").lower()
    is_maz = source == "maz"

    if not is_maz and not batch.paychex_exported_at:
        # Build the Excel in-memory to validate it won't fail.
        # Any exception here aborts before a single email goes out.
        try:
            import openpyxl as _openpyxl
            from sqlalchemy import text as _sql_text2
            _ovr_ids: set[int] | None = None
            _mw_ids: set[int] | None = None
            try:
                _ovr_rows = db.execute(
                    _sql_text2("SELECT person_id FROM payroll_withheld_override WHERE batch_id = :b"),
                    {"b": batch_id},
                ).fetchall()
                _ovr_ids = {r[0] for r in _ovr_rows} or None
                _mw_rows = db.execute(
                    _sql_text2("SELECT person_id FROM payroll_manual_withhold"),
                ).fetchall()
                _mw_ids = {r[0] for r in _mw_rows} or None
            except Exception:
                pass  # graceful degradation — override tables absent in test environments
            _paychex_data = _build_summary(db, batch_id=batch_id, override_ids=_ovr_ids, manual_withhold_ids=_mw_ids)
            _paid_count = sum(1 for r in _paychex_data["rows"] if not r["withheld"] and r["pay_this_period"] > 0)
            if _paid_count > 0:
                # Validate workbook builds without error
                _wb = _openpyxl.Workbook()
                _ws_check = _wb.active
                _build_payroll_summary_tab(
                    _ws_check, batch, _paychex_data["rows"], _paychex_data["totals"],
                    "FirstAlt — Payroll Summary",
                )
                _buf = _io.BytesIO()
                _wb.save(_buf)
                if _buf.tell() == 0:
                    return JSONResponse(
                        {"error": "Paychex Excel generation produced an empty file — aborting before any emails were sent."},
                        status_code=500,
                    )
            # _paid_count == 0 means all withheld — valid case, skip generation but continue.
        except Exception as _exc:
            _logging.getLogger("zpay.workflow").error(
                "Paychex Excel pre-flight failed for batch %s — aborting stub send: %s",
                batch_id, _exc,
            )
            return JSONResponse(
                {"error": f"Paychex Excel generation failed — aborting before any emails were sent. Detail: {_exc}"},
                status_code=500,
            )

        # Stamp paychex_exported_at so the workflow gate doesn't re-block
        from datetime import datetime, timezone as _tz
        batch.paychex_exported_at = datetime.now(_tz.utc)
        db.commit()

    payweek = _build_payweek(batch)
    company = batch.company_name or ""

    # For test sends, include all drivers with email (ignore already_sent filter —
    # test sends never count as "already sent" for real send purposes).
    is_test_send = test_recipient_override is not None

    if is_test_send:
        # All drivers in batch with an email address — ignore sent status
        drivers = (
            db.query(Person)
            .join(Ride, Ride.person_id == Person.person_id)
            .filter(
                Ride.payroll_batch_id == batch_id,
                Person.email.isnot(None),
                Person.email != "",
            )
            .distinct()
            .all()
        )
    else:
        # Normal path: only drivers who haven't been sent yet
        already_sent = (
            db.query(EmailSendLog.person_id)
            .filter(
                EmailSendLog.payroll_batch_id == batch_id,
                EmailSendLog.status == "sent",
                EmailSendLog.is_test == False,  # noqa: E712
            )
            .subquery()
        )
        drivers = (
            db.query(Person)
            .join(Ride, Ride.person_id == Person.person_id)
            .filter(
                Ride.payroll_batch_id == batch_id,
                Person.email.isnot(None),
                Person.email != "",
                ~Person.person_id.in_(db.query(already_sent.c.person_id)),
            )
            .distinct()
            .all()
        )

    # ── confirmed_recipient_count gate ───────────────────────────────────────
    actual_count = len(drivers)
    if confirmed_recipient_count != actual_count:
        return JSONResponse(
            {
                "ok": False,
                "error": (
                    f"Driver count mismatch: you confirmed {confirmed_recipient_count} "
                    f"but the batch has {actual_count} eligible drivers. "
                    "Refresh and try again."
                ),
            },
            status_code=400,
        )

    sent = 0
    failed = 0
    for person in drivers:
        rides = (
            db.query(Ride)
            .filter(Ride.payroll_batch_id == batch_id, Ride.person_id == person.person_id)
            .order_by(Ride.ride_start_ts.asc())
            .all()
        )

        pdf_path = _generate_pdf(person, rides, company, payweek)
        total_pay = sum(float(r.z_rate or 0) for r in rides)

        # Use override recipient for test sends
        effective_email = test_recipient_override if is_test_send else person.email

        try:
            send_paystub(
                to_email=effective_email,
                driver_name=person.full_name,
                company=company,
                payweek=payweek,
                pdf_path=pdf_path,
                person_id=person.person_id,
                payroll_batch_id=batch_id,
                week_start=batch.week_start.isoformat() if batch.week_start else "",
                week_end=batch.week_end.isoformat() if batch.week_end else "",
                total_pay=f"{total_pay:.2f}",
                ride_count=len(rides),
                db=db,
            )
            if not is_test_send:
                # Clear any old failed logs and record success (real send only)
                db.query(EmailSendLog).filter(
                    EmailSendLog.payroll_batch_id == batch_id,
                    EmailSendLog.person_id == person.person_id,
                    EmailSendLog.status == "failed",
                ).delete()
            db.add(EmailSendLog(
                payroll_batch_id=batch_id,
                person_id=person.person_id,
                status="sent",
                is_test=is_test_send,
            ))
            db.commit()
            sent += 1
        except Exception as exc:
            import traceback
            _logging.getLogger("zpay.workflow").error(
                "Failed to send stub to %s <%s>: %s\n%s",
                person.full_name, effective_email, exc, traceback.format_exc(),
            )
            db.add(EmailSendLog(
                payroll_batch_id=batch_id,
                person_id=person.person_id,
                status="failed",
                error_message=str(exc)[:200],
                is_test=is_test_send,
            ))
            db.commit()
            failed += 1

    return JSONResponse({
        "ok": True,
        "sent": sent,
        "failed": failed,
        "total_drivers": actual_count,
        "is_test": is_test_send,
    })


# ── Resend all stubs (admin retry — real send to all drivers) ────────────────

@router.post("/{batch_id}/resend-stubs")
def workflow_resend_stubs(
    batch_id: int,
    confirmed_recipient_count: int = Query(..., description="Must match actual driver count"),
    db: Session = Depends(get_db),
    _admin=Depends(require_role("admin")),
):
    """Admin-only: resend real paystubs to ALL drivers in the batch, including
    those already marked as sent. Use case: original send failed for some drivers.
    Creates fresh EmailSendLog entries (is_test=False).
    """
    from backend.routes.email import _generate_pdf, _build_payweek
    from backend.services.email_service import send_paystub
    import logging as _log
    import traceback as _tb

    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    # Admin bypass: admin role may resend from any batch status (e.g. export_ready)
    if batch.status not in ("stubs_sending", "complete", "export_ready", "approved"):
        return JSONResponse(
            {"ok": False, "error": f"Batch status '{batch.status}' does not allow resend."},
            status_code=400,
        )

    payweek = _build_payweek(batch)
    company = batch.company_name or ""

    drivers = (
        db.query(Person)
        .join(Ride, Ride.person_id == Person.person_id)
        .filter(
            Ride.payroll_batch_id == batch_id,
            Person.email.isnot(None),
            Person.email != "",
        )
        .distinct()
        .all()
    )

    actual_count = len(drivers)
    if confirmed_recipient_count != actual_count:
        return JSONResponse(
            {
                "ok": False,
                "error": (
                    f"Driver count mismatch: you confirmed {confirmed_recipient_count} "
                    f"but found {actual_count} drivers. Refresh and retry."
                ),
            },
            status_code=400,
        )

    sent = 0
    failed = 0
    for person in drivers:
        rides = (
            db.query(Ride)
            .filter(Ride.payroll_batch_id == batch_id, Ride.person_id == person.person_id)
            .order_by(Ride.ride_start_ts.asc())
            .all()
        )
        pdf_path = _generate_pdf(person, rides, company, payweek)
        total_pay = sum(float(r.z_rate or 0) for r in rides)
        try:
            send_paystub(
                to_email=person.email,
                driver_name=person.full_name,
                company=company,
                payweek=payweek,
                pdf_path=pdf_path,
                person_id=person.person_id,
                payroll_batch_id=batch_id,
                week_start=batch.week_start.isoformat() if batch.week_start else "",
                week_end=batch.week_end.isoformat() if batch.week_end else "",
                total_pay=f"{total_pay:.2f}",
                ride_count=len(rides),
                db=db,
            )
            db.add(EmailSendLog(
                payroll_batch_id=batch_id,
                person_id=person.person_id,
                status="sent",
                is_test=False,
            ))
            db.commit()
            sent += 1
        except Exception as exc:
            _log.getLogger("zpay.workflow").error(
                "resend-stubs failed for person %s: %s\n%s", person.person_id, exc, _tb.format_exc()
            )
            db.add(EmailSendLog(
                payroll_batch_id=batch_id,
                person_id=person.person_id,
                status="failed",
                error_message=str(exc)[:200],
                is_test=False,
            ))
            db.commit()
            failed += 1

    return JSONResponse({"ok": True, "sent": sent, "failed": failed, "total_drivers": actual_count})


# ── Send single stub (for progress-bar flow) ────────────────────────────────

@router.post("/{batch_id}/send-stub/{person_id}")
def workflow_send_single_stub(batch_id: int, person_id: int, db: Session = Depends(get_db)):
    """Send a paystub to one driver. Used by the frontend progress-bar loop."""
    from backend.routes.email import _generate_pdf, _build_payweek
    from backend.services.email_service import send_paystub

    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    person = db.query(Person).filter(Person.person_id == person_id).first()
    if not person:
        return JSONResponse({"ok": False, "status": "not_found", "error": "Driver not found"}, status_code=404)
    if not person.email:
        return JSONResponse({"ok": True, "status": "no_email", "name": person.full_name})

    # Skip withheld drivers — they're not receiving pay this period
    from backend.db.models import DriverBalance
    balance = db.query(DriverBalance).filter(
        DriverBalance.person_id == person_id,
        DriverBalance.payroll_batch_id == batch_id,
    ).first()
    if balance and balance.carried_over and float(balance.carried_over) > 0:
        return JSONResponse({"ok": True, "status": "no_email", "name": person.full_name})

    # Skip if already sent
    already = db.query(EmailSendLog).filter(
        EmailSendLog.payroll_batch_id == batch_id,
        EmailSendLog.person_id == person_id,
        EmailSendLog.status == "sent",
    ).first()
    if already:
        return JSONResponse({"ok": True, "status": "already_sent", "name": person.full_name})

    payweek = _build_payweek(batch)
    rides = (
        db.query(Ride)
        .filter(Ride.payroll_batch_id == batch_id, Ride.person_id == person_id)
        .order_by(Ride.ride_start_ts.asc())
        .all()
    )

    pdf_path = _generate_pdf(person, rides, batch.company_name or "", payweek)
    total_pay = sum(float(r.z_rate or 0) for r in rides)

    try:
        send_paystub(
            to_email=person.email,
            driver_name=person.full_name,
            company=batch.company_name or "",
            payweek=payweek,
            pdf_path=pdf_path,
            person_id=person_id,
            payroll_batch_id=batch_id,
            week_start=batch.week_start.isoformat() if batch.week_start else "",
            week_end=batch.week_end.isoformat() if batch.week_end else "",
            total_pay=f"{total_pay:.2f}",
            ride_count=len(rides),
            db=db,
        )
        # Clear old failed logs, record success
        db.query(EmailSendLog).filter(
            EmailSendLog.payroll_batch_id == batch_id,
            EmailSendLog.person_id == person_id,
            EmailSendLog.status == "failed",
        ).delete()
        db.add(EmailSendLog(
            payroll_batch_id=batch_id,
            person_id=person_id,
            status="sent",
        ))
        db.commit()
        return JSONResponse({"ok": True, "status": "sent", "name": person.full_name})
    except Exception as exc:
        import logging, traceback
        logging.getLogger("zpay.workflow").error(
            "send-stub failed for person %s: %s\n%s", person_id, exc, traceback.format_exc()
        )
        db.add(EmailSendLog(
            payroll_batch_id=batch_id,
            person_id=person_id,
            status="failed",
            error_message=str(exc)[:200],
        ))
        db.commit()
        return JSONResponse({"ok": False, "status": "failed", "name": person.full_name, "error": str(exc)[:200]})


# ── Retry single stub ────────────────────────────────────────────────────────

@router.post("/{batch_id}/retry-stub/{person_id}")
def workflow_retry_stub(batch_id: int, person_id: int, db: Session = Depends(get_db)):
    from backend.routes.email import _generate_pdf, _build_payweek
    from backend.services.email_service import send_paystub

    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    person = db.query(Person).filter(Person.person_id == person_id).first()
    if not person or not person.email:
        return JSONResponse({"error": "Driver not found or no email"}, status_code=404)

    # Skip withheld drivers — they're not receiving pay this period
    from backend.db.models import DriverBalance
    balance = db.query(DriverBalance).filter(
        DriverBalance.person_id == person_id,
        DriverBalance.payroll_batch_id == batch_id,
    ).first()
    if balance and balance.carried_over and float(balance.carried_over) > 0:
        return JSONResponse({"ok": True, "status": "no_email", "name": person.full_name})

    # Delete old failed log
    db.query(EmailSendLog).filter(
        EmailSendLog.payroll_batch_id == batch_id,
        EmailSendLog.person_id == person_id,
        EmailSendLog.status == "failed",
    ).delete()
    db.commit()

    payweek = _build_payweek(batch)
    rides = (
        db.query(Ride)
        .filter(Ride.payroll_batch_id == batch_id, Ride.person_id == person_id)
        .order_by(Ride.ride_start_ts.asc())
        .all()
    )

    pdf_path = _generate_pdf(person, rides, batch.company_name or "", payweek)
    total_pay = sum(float(r.z_rate or 0) for r in rides)

    try:
        send_paystub(
            to_email=person.email,
            driver_name=person.full_name,
            company=batch.company_name or "",
            payweek=payweek,
            pdf_path=pdf_path,
            person_id=person_id,
            payroll_batch_id=batch_id,
            week_start=batch.week_start.isoformat() if batch.week_start else "",
            week_end=batch.week_end.isoformat() if batch.week_end else "",
            total_pay=f"{total_pay:.2f}",
            ride_count=len(rides),
            db=db,
        )
        return JSONResponse({"ok": True, "status": "sent"})
    except Exception as exc:
        import logging, traceback
        logging.getLogger("zpay.workflow").error(
            "retry-stub failed for person %s: %s\n%s", person_id, exc, traceback.format_exc()
        )
        db.add(EmailSendLog(
            payroll_batch_id=batch_id,
            person_id=person_id,
            status="failed",
            error_message=str(exc)[:200],
        ))
        db.commit()
        return JSONResponse({"ok": False, "error": str(exc)[:200]}, status_code=500)


# ── Direct withheld toggle (for summary page corrections) ────────────────────

@router.post("/{batch_id}/set-withheld/{person_id}")
async def workflow_set_withheld(batch_id: int, person_id: int, request: Request, db: Session = Depends(get_db)):
    """
    Directly flip a driver's withheld status for this batch by editing
    their DriverBalance record. Used from the summary page to correct
    mistakes without re-running the full workflow.

    body: { "withheld": true | false }
    """
    from sqlalchemy import text
    body = await request.json()
    withheld = body.get("withheld", True)

    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)

    # Calculate combined pay for this driver (this week + any carry-in from last batch)
    data = _build_summary(db, batch_id=batch_id, auto_save=False)
    driver_row = next((r for r in data["rows"] if r["person_id"] == person_id), None)
    if not driver_row:
        return JSONResponse({"error": "Driver not in this batch"}, status_code=404)

    combined = round(driver_row["driver_pay"] + driver_row["from_last_period"], 2)

    # Update (or create) DriverBalance record
    existing = db.query(DriverBalance).filter(
        DriverBalance.payroll_batch_id == batch_id,
        DriverBalance.person_id == person_id,
    ).first()

    if withheld:
        # Move driver to withheld: store full combined amount as carry-over
        db.execute(
            text("DELETE FROM payroll_withheld_override WHERE batch_id = :b AND person_id = :p"),
            {"b": batch_id, "p": person_id},
        )
        if existing:
            existing.carried_over = combined
        else:
            db.add(DriverBalance(person_id=person_id, payroll_batch_id=batch_id, carried_over=combined))
    else:
        # Move driver to paid: zero out the balance, mark as force-paid
        db.execute(
            text("INSERT INTO payroll_withheld_override (batch_id, person_id) VALUES (:b, :p) ON CONFLICT DO NOTHING"),
            {"b": batch_id, "p": person_id},
        )
        if existing:
            existing.carried_over = 0
        else:
            db.add(DriverBalance(person_id=person_id, payroll_batch_id=batch_id, carried_over=0))

    db.commit()
    return JSONResponse({"ok": True, "withheld": withheld, "combined": combined})


# ── Withheld override endpoints ─────────────────────────────────────────────

@router.post("/{batch_id}/override-withheld/{person_id}")
def workflow_override_withheld(batch_id: int, person_id: int, db: Session = Depends(get_db)):
    """Force-pay a withheld driver for this batch regardless of the $100 threshold."""
    from sqlalchemy import text
    db.execute(
        text("INSERT INTO payroll_withheld_override (batch_id, person_id) VALUES (:b, :p) ON CONFLICT DO NOTHING"),
        {"b": batch_id, "p": person_id},
    )
    db.commit()
    return JSONResponse({"ok": True})


@router.delete("/{batch_id}/override-withheld/{person_id}")
def workflow_remove_withheld_override(batch_id: int, person_id: int, db: Session = Depends(get_db)):
    """Remove a force-pay override — driver goes back to normal withheld logic."""
    from sqlalchemy import text
    db.execute(
        text("DELETE FROM payroll_withheld_override WHERE batch_id = :b AND person_id = :p"),
        {"b": batch_id, "p": person_id},
    )
    db.commit()
    return JSONResponse({"ok": True})


@router.post("/{batch_id}/manual-withhold/{person_id}")
async def workflow_set_manual_withhold(batch_id: int, person_id: int, request: Request, db: Session = Depends(get_db)):
    """Manually withhold a driver's pay regardless of amount, with a note."""
    from sqlalchemy import text
    body = await request.json()
    note = body.get("note", "").strip()
    db.execute(
        text("INSERT INTO payroll_manual_withhold (person_id, note) VALUES (:p, :n) ON CONFLICT (person_id) DO UPDATE SET note = EXCLUDED.note, created_at = now()"),
        {"p": person_id, "n": note},
    )
    db.commit()
    return JSONResponse({"ok": True})

@router.delete("/{batch_id}/manual-withhold/{person_id}")
def workflow_clear_manual_withhold(batch_id: int, person_id: int, db: Session = Depends(get_db)):
    """Release a manual withhold — driver gets paid normally next batch."""
    from sqlalchemy import text
    db.execute(
        text("DELETE FROM payroll_manual_withhold WHERE person_id = :p"),
        {"p": person_id},
    )
    db.commit()
    return JSONResponse({"ok": True})


# ── Inline edit endpoints ───────────────────────────────────────────────────

@router.patch("/{batch_id}/update-person/{person_id}")
async def workflow_update_person(batch_id: int, person_id: int, request: Request, db: Session = Depends(get_db)):
    """Update a person's paycheck_code or email inline from the review step."""
    body = await request.json()
    person = db.query(Person).filter(Person.person_id == person_id).first()
    if not person:
        return JSONResponse({"error": "Person not found"}, status_code=404)

    if "paycheck_code" in body:
        person.paycheck_code = body["paycheck_code"].strip() or None
    if "email" in body:
        person.email = body["email"].strip() or None

    db.commit()
    return JSONResponse({
        "ok": True,
        "person_id": person.person_id,
        "paycheck_code": person.paycheck_code or "",
        "email": person.email or "",
    })


@router.patch("/{batch_id}/update-ride-rate")
async def workflow_update_ride_rate(batch_id: int, request: Request, db: Session = Depends(get_db)):
    """Update z_rate for rides in a batch matching a service_name.

    ``mode`` controls scope:
      - "default" (default): update ZRateService.default_rate (permanent) and
        every matching ride in this batch.
      - "late_cancellation": save as ZRateService.late_cancellation_rate so
        future late-cancel rides (net_pay 40–55% of default_rate) on this
        route get this rate automatically. Only late-cancel rides in the
        current batch are re-rated; regular rides keep their original rate.
      - "batch_only": update only matching rides in this batch; leave the
        permanent default_rate and late_cancellation_rate untouched.

    Legacy ``batch_only: true`` is still accepted and mapped to mode=batch_only.
    """
    from decimal import Decimal

    body = await request.json()
    service_name = body.get("service_name", "").strip()
    z_rate = body.get("z_rate")
    mode = (body.get("mode") or "").strip().lower()
    if not mode:
        mode = "batch_only" if bool(body.get("batch_only", False)) else "default"

    if mode not in ("default", "late_cancellation", "batch_only"):
        return JSONResponse({"error": f"invalid mode: {mode}"}, status_code=400)

    if not service_name or z_rate is None:
        return JSONResponse({"error": "service_name and z_rate required"}, status_code=400)

    rate_val = float(z_rate)
    if rate_val < 0:
        return JSONResponse({"error": "z_rate cannot be negative"}, status_code=400)

    svc = db.query(ZRateService).filter(ZRateService.service_name == service_name).first()

    if mode == "late_cancellation":
        # Persist as the per-service late-cancellation rate.
        if svc:
            svc.late_cancellation_rate = Decimal(str(rate_val))

        # Re-rate only late-cancellation rides in THIS batch
        # (net_pay is 40–55% of the existing z_rate).
        updated = (
            db.query(Ride)
            .filter(
                Ride.payroll_batch_id == batch_id,
                Ride.service_name == service_name,
                Ride.z_rate > 0,
                Ride.net_pay > 0,
                Ride.net_pay >= Ride.z_rate * 0.40,
                Ride.net_pay <= Ride.z_rate * 0.55,
            )
            .update({"z_rate": rate_val}, synchronize_session=False)
        )
    else:
        # default: update rides in this batch AND every other non-finalized batch
        # for the same service so the permanent rate truly persists across batches.
        # batch_only: only this batch, svc row untouched.
        if mode == "default":
            # Update ALL open/unfinalized batches (do not touch status=complete/approved)
            from backend.db.models import PayrollBatch
            open_batch_ids = [
                row.payroll_batch_id
                for row in db.query(PayrollBatch.payroll_batch_id).filter(
                    PayrollBatch.status.notin_(["complete", "approved"])
                ).all()
            ]
            updated = (
                db.query(Ride)
                .filter(
                    Ride.payroll_batch_id.in_(open_batch_ids),
                    Ride.service_name == service_name,
                )
                .update({"z_rate": rate_val, "z_rate_source": "permanent_override"}, synchronize_session=False)
            )
            # Also update the current batch explicitly in case it was already finalized
            # or not in open_batch_ids (edge case)
            db.query(Ride).filter(
                Ride.payroll_batch_id == batch_id,
                Ride.service_name == service_name,
            ).update({"z_rate": rate_val, "z_rate_source": "permanent_override"}, synchronize_session=False)
            if svc:
                svc.default_rate = Decimal(str(rate_val))
        else:
            # batch_only: only this batch
            updated = (
                db.query(Ride)
                .filter(Ride.payroll_batch_id == batch_id, Ride.service_name == service_name)
                .update({"z_rate": rate_val}, synchronize_session=False)
            )

    db.commit()
    return JSONResponse({"ok": True, "rides_updated": updated, "mode": mode})


# ── Manual adjustment route picker ──────────────────────────────────────────

# TODO: add auth dependency once auth system is wired (see batches.py for pattern)
@router.get("/{batch_id}/routes")
def workflow_batch_routes(batch_id: int, db: Session = Depends(get_db)):
    """Return available routes for the manual-adjustment route picker.

    Filters ``ZRateService`` rows to the batch's ``source`` and ``company_name``,
    then LEFT JOINs to the most-recent non-manual ride per route to surface
    ``last_miles`` and ``last_ride_date``.

    Response shape (list, ordered by service_name):
        [
          {
            "z_rate_service_id": int,
            "service_name": str,
            "default_rate": float,
            "last_miles": float,
            "last_ride_date": "YYYY-MM-DD" | null
          },
          ...
        ]
    """
    import logging
    from decimal import Decimal
    from sqlalchemy import text as _text

    logger = logging.getLogger(__name__)

    batch = db.query(PayrollBatch).filter(
        PayrollBatch.payroll_batch_id == batch_id
    ).first()
    if not batch:
        return JSONResponse({"error": f"batch_id {batch_id} not found"}, status_code=404)

    try:
        rows = db.execute(
            _text("""
                SELECT
                    svc.z_rate_service_id,
                    svc.service_name,
                    svc.default_rate,
                    recent.miles       AS last_miles,
                    recent.last_date   AS last_ride_date
                FROM z_rate_service svc
                LEFT JOIN LATERAL (
                    SELECT
                        r.miles,
                        DATE(r.ride_start_ts) AS last_date
                    FROM ride r
                    WHERE r.service_name = svc.service_name
                      AND r.source != 'manual'
                    ORDER BY r.ride_start_ts DESC
                    LIMIT 1
                ) recent ON TRUE
                WHERE svc.active = TRUE
                  AND (
                        (svc.source IS NULL OR svc.source = '')
                        OR lower(svc.source) = lower(:batch_source)
                  )
                  AND (
                        (svc.company_name IS NULL OR svc.company_name = '')
                        OR lower(svc.company_name) = lower(:batch_company)
                  )
                ORDER BY svc.service_name ASC
            """),
            {
                "batch_source": batch.source or "",
                "batch_company": batch.company_name or "",
            },
        ).fetchall()

        result = [
            {
                "z_rate_service_id": r.z_rate_service_id,
                "service_name": r.service_name,
                "default_rate": float(r.default_rate) if r.default_rate is not None else 0.0,
                "last_miles": float(r.last_miles) if r.last_miles is not None else 0.0,
                "last_ride_date": str(r.last_ride_date) if r.last_ride_date else None,
            }
            for r in rows
        ]
        return JSONResponse(result)

    except Exception as exc:
        logger.exception("workflow_batch_routes failed for batch_id=%s", batch_id)
        return JSONResponse({"error": str(exc)}, status_code=500)

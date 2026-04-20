import csv
import io
import re
from io import BytesIO
from pathlib import Path
from datetime import date, datetime

from fastapi import APIRouter, Depends, Form, Request, Query
from fastapi.responses import StreamingResponse, RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import func, cast, Date
from sqlalchemy.orm import Session
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.db import get_db
from backend.db.models import Person, Ride, PayrollBatch, DriverBalance

router = APIRouter(prefix="/summary", tags=["summary"])

_templates = None
def templates():
    global _templates
    if _templates is None:
        templates_dir = Path(__file__).resolve().parents[1] / "templates"
        _templates = Jinja2Templates(directory=str(templates_dir))
    return _templates


def _batch_period_label(batch: PayrollBatch, acumen_rank: int | None = None) -> str:
    """Return 'Week N · M/D – M/D' for a batch."""
    start = batch.period_start.strftime("%-m/%-d") if batch.period_start else "?"
    end = batch.period_end.strftime("%-m/%-d") if batch.period_end else "?"
    if batch.source == "maz" and batch.batch_ref:
        m = re.search(r'W(\d+)$', batch.batch_ref or '')
        week_num = int(m.group(1)) if m else None
    else:
        week_num = acumen_rank
    if week_num:
        return f"Week {week_num} · {start} – {end}"
    return f"{start} – {end}"


def _build_week_rank_map(batches: list) -> dict:
    """Map payroll_batch_id → week number for acumen batches (ranked by week_start)."""
    acumen = sorted(
        [b for b in batches if b.source == "acumen" and b.week_start],
        key=lambda b: b.week_start,
    )
    return {b.payroll_batch_id: rank for rank, b in enumerate(acumen, start=1)}


COLUMNS = [
    "Driver Name", "Pay Code", "Rides", "Miles",
    "Partner Pays", "Driver Pay", "Deduction",
    "Withheld (Y/N)", "Carried Over", "Paid This Period",
]


def _get_companies(db: Session) -> list[str]:
    rows = (
        db.query(PayrollBatch.company_name)
        .distinct()
        .order_by(PayrollBatch.company_name.asc())
        .all()
    )
    return [r[0] for r in rows]


PAY_THRESHOLD = 100.0


def _resolve_latest_open_batch(
    db: Session,
    company: str | None = None,
    source: str | None = None,
) -> int | None:
    """
    Return the payroll_batch_id of the most recent open (not finalized) batch.
    A batch is considered open when finalized_at IS NULL.
    Filters by company or source when provided.
    Returns None if no open batch exists.
    """
    q = (
        db.query(PayrollBatch)
        .filter(PayrollBatch.finalized_at.is_(None))
    )
    if source:
        q = q.filter(PayrollBatch.source == source)
    elif company:
        q = q.filter(PayrollBatch.company_name == company)
    batch = q.order_by(PayrollBatch.period_end.desc(), PayrollBatch.created_at.desc()).first()
    return batch.payroll_batch_id if batch else None


def _build_summary(
    db: Session,
    company: str | None = None,
    source: str | None = None,
    batch_id: int | None = None,
    start: date | None = None,
    end: date | None = None,
    auto_save: bool = False,
    override_ids: set[int] | None = None,
    manual_withhold_ids: set[int] | None = None,
) -> dict:
    """
    Returns rows + totals for the summary page.

    When batch_id is provided:
    - Looks up the previous batch for the same company
    - Reads driver_balance for that previous batch as "from_last_period"
    - combined = net_pay + from_last_period
    - If combined < $100 → withheld; auto-saves combined to driver_balance for this batch
    - If combined >= $100 → driver gets paid; clears any driver_balance record for this batch

    When batch_id is NOT provided and no date range is given:
    - Resolves to the latest open (finalized_at IS NULL) batch for the given company/source.
    - If no open batch exists, returns empty rows with a "no_active_batch" signal.
    """
    # Resolve to latest open batch when no explicit batch_id or date range is supplied
    if not batch_id and not start and not end:
        batch_id = _resolve_latest_open_batch(db, company=company, source=source)
        if batch_id is None:
            return {"rows": [], "totals": {
                "rides": 0, "miles": 0.0, "partner_pays": 0.0,
                "driver_pay": 0.0, "deduction": 0.0, "carried_over": 0.0,
                "days": 0, "net_pay": 0.0, "pay_this_period": 0.0,
            }, "no_active_batch": True}

    ride_date = func.coalesce(
        cast(Ride.ride_start_ts, Date),
    )

    q = (
        db.query(
            Person.person_id.label("person_id"),
            Person.full_name.label("person"),
            Person.paycheck_code.label("code"),
            func.min(ride_date).label("first_date"),
            func.max(ride_date).label("last_date"),
            func.count(Ride.ride_id).label("rides"),
            func.coalesce(func.sum(Ride.miles), 0).label("miles"),
            func.count(func.distinct(ride_date)).label("days"),
            func.coalesce(func.sum(Ride.net_pay), 0).label("gross_earned"),
            func.coalesce(func.sum(Ride.deduction), 0).label("total_deduction"),
            func.coalesce(func.sum(Ride.z_rate), 0).label("z_rate_total"),
        )
        .join(Ride, Ride.person_id == Person.person_id)
        .join(PayrollBatch, PayrollBatch.payroll_batch_id == Ride.payroll_batch_id)
    )

    # batch_id is the primary scope — always prefer it over a broad company scan
    if batch_id:
        q = q.filter(PayrollBatch.payroll_batch_id == batch_id)
    elif source:
        q = q.filter(PayrollBatch.source == source)
    elif company:
        q = q.filter(PayrollBatch.company_name == company)
    if start:
        q = q.filter(ride_date >= start)
    if end:
        q = q.filter(ride_date <= end)

    q = q.group_by(Person.person_id, Person.full_name, Person.external_id)
    q = q.order_by(Person.full_name.asc())

    rows_raw = q.all()

    # ── Carried-over amounts from previous batch ───────────────────────────────
    carried_map: dict[int, float] = {}
    if batch_id:
        current_batch = db.query(PayrollBatch).filter(
            PayrollBatch.payroll_batch_id == batch_id
        ).first()
        if current_batch:
            prev_batch = (
                db.query(PayrollBatch)
                .filter(
                    PayrollBatch.company_name == current_batch.company_name,
                    PayrollBatch.period_start < current_batch.period_start,
                )
                .order_by(PayrollBatch.period_start.desc())
                .first()
            )
            if prev_batch:
                prev_balances = (
                    db.query(DriverBalance)
                    .filter(DriverBalance.payroll_batch_id == prev_batch.payroll_batch_id)
                    .all()
                )
                carried_map = {b.person_id: round(float(b.carried_over or 0), 2) for b in prev_balances}

    def fmt(d):
        return d.strftime("%-m/%-d/%Y") if d else ""

    rows = []
    total_rides = 0
    total_miles = 0.0
    total_partner = 0.0
    total_driver = 0.0
    total_deduction = 0.0
    total_carried = 0.0
    total_pay = 0.0

    for r in rows_raw:
        rides = int(r.rides or 0)
        miles = round(float(r.miles or 0), 3)
        # partner_pays = what the partner (FA/ED) pays
        partner_pays = round(float(r.gross_earned or 0), 2)
        deduction = round(float(r.total_deduction or 0), 2)
        # driver_pay = what Maz pays the driver (sum of z_rate on rides)
        driver_pay = round(float(r.z_rate_total or 0), 2)
        days = int(r.days or 0)
        active = f"{fmt(r.first_date)} – {fmt(r.last_date)}" if r.first_date else ""
        from_last = carried_map.get(r.person_id, 0.0)
        combined = round(driver_pay + from_last, 2)
        withheld = combined < PAY_THRESHOLD
        if override_ids and r.person_id in override_ids:
            withheld = False  # force pay regardless of threshold
        if manual_withhold_ids and r.person_id in manual_withhold_ids:
            withheld = True  # manual override always withholds
        pay_this_period = 0.0 if withheld else combined

        rows.append({
            "person_id": r.person_id,
            "person": r.person or "",
            "code": r.code or "",
            "rides": rides,
            "miles": miles,
            "partner_pays": partner_pays,
            "driver_pay": driver_pay,
            "deduction": deduction,
            "active_between": active,
            "days": days,
            "net_pay": partner_pays,  # kept for JSON API compat
            "from_last_period": from_last,
            "pay_this_period": pay_this_period,
            "withheld": withheld,
            "withheld_amount": combined if withheld else 0.0,
        })
        total_rides += rides
        total_miles += miles
        total_partner += partner_pays
        total_driver += driver_pay
        total_deduction += deduction
        total_carried += from_last if withheld else 0.0
        total_pay += pay_this_period

    totals = {
        "rides": total_rides,
        "miles": round(total_miles, 3),
        "partner_pays": round(total_partner, 2),
        "driver_pay": round(total_driver, 2),
        "deduction": round(total_deduction, 2),
        "carried_over": round(total_carried, 2),
        "days": int(sum(r["days"] for r in rows)),
        "net_pay": round(total_partner, 2),
        "pay_this_period": round(total_pay, 2),
    }

    # ── Auto-save withheld balances for the current batch ─────────────────────
    if batch_id and auto_save:
        for row in rows:
            existing = (
                db.query(DriverBalance)
                .filter(
                    DriverBalance.person_id == row["person_id"],
                    DriverBalance.payroll_batch_id == batch_id,
                )
                .first()
            )
            if row["withheld"]:
                # Store the combined withheld amount so the NEXT batch can carry it forward
                if existing:
                    existing.carried_over = row["withheld_amount"]
                else:
                    db.add(DriverBalance(
                        person_id=row["person_id"],
                        payroll_batch_id=batch_id,
                        carried_over=row["withheld_amount"],
                    ))
            else:
                # Driver is being paid this period — clear any stale balance record
                if existing:
                    db.delete(existing)
        db.commit()

    return {"rows": rows, "totals": totals}


# ── Summary page ──────────────────────────────────────────────────────────────

@router.get("/", name="summary_page")
def summary_page(
    request: Request,
    company: str | None = Query(None),
    batch_id: int | None = Query(None),
    start: date | None = Query(None),
    end: date | None = Query(None),
    db: Session = Depends(get_db),
):
    _wants_json = (
        "application/json" in request.headers.get("content-type", "")
        or "application/json" in request.headers.get("accept", "")
    )

    companies = _get_companies(db)

    # Default to first company if none selected
    selected_company = company or (companies[0] if companies else None)

    # Batches for the selected company
    batches = []
    if selected_company:
        batches = (
            db.query(PayrollBatch)
            .filter(PayrollBatch.company_name == selected_company)
            .order_by(PayrollBatch.period_start.desc())
            .all()
        )

    # GET is always read-only — never auto-save on page load
    data = _build_summary(db, company=selected_company, batch_id=batch_id, start=start, end=end, auto_save=False)

    if _wants_json:
        try:
            rows = data["rows"]
            totals = data["totals"]
            rank_map = _build_week_rank_map(batches)
            periods = [
                {"label": _batch_period_label(b, rank_map.get(b.payroll_batch_id)), "batch_id": b.payroll_batch_id}
                for b in batches
            ]
            drivers_out = []
            withheld_out = []
            for r in rows:
                entry = {
                    "id": r["person_id"],
                    "name": r["person"],
                    "pay_code": r["code"],
                    "days": r["days"],
                    "net_pay": r["net_pay"],
                    "carried_over": r["from_last_period"],
                    "pay_this_period": r["pay_this_period"],
                    "status": "withheld" if r["withheld"] else "paid",
                    "override": False,
                    "withheld": r["withheld_amount"],
                }
                if r["withheld"]:
                    withheld_out.append(entry)
                else:
                    drivers_out.append(entry)
            total_withheld = sum(r["withheld_amount"] for r in rows if r["withheld"])
            current_batch_id = batches[0].payroll_batch_id if batches else None
            selected_bid = batch_id or current_batch_id
            # Find label for selected batch to populate Period stat card
            selected_batch_obj = next((b for b in batches if b.payroll_batch_id == selected_bid), None)
            selected_week_label = _batch_period_label(selected_batch_obj, rank_map.get(selected_bid)) if selected_batch_obj else None
            return JSONResponse({
                "company": selected_company,
                "batch_id": selected_bid,
                "period": selected_week_label,
                "periods": periods,
                "drivers": drivers_out,
                "withheld": withheld_out,
                "stats": {
                    "driver_count": len(rows),
                    "total_pay": totals["pay_this_period"],
                    "withheld_amount": round(total_withheld, 2),
                },
            })
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=500)

    return templates().TemplateResponse(
        request,
        "summary.html",
        {
            "rows": data["rows"],
            "totals": data["totals"],
            "companies": companies,
            "selected_company": selected_company,
            "batches": batches,
            "selected_batch_id": batch_id,
            "start": start,
            "end": end,
            "payroll_run": False,
        },
    )


# ── Run Payroll (POST — commits withheld balances) ────────────────────────────

@router.post("/run", name="summary_run")
async def summary_run(
    request: Request,
    overrides: str = Form(""),
    company: str | None = Query(None),
    batch_id: int | None = Query(None),
    db: Session = Depends(get_db),
):
    """
    Explicitly run payroll for a batch. This is the only place that writes
    DriverBalance records — the GET route is strictly read-only.
    Must have a batch_id to commit withheld balances.

    overrides: comma-separated person_id values for drivers who should be
    force-paid this period regardless of the $100 threshold.
    """
    # Support JSON body from Next.js frontend
    accept = request.headers.get("accept", "")
    content_type = request.headers.get("content-type", "")
    is_json = "json" in accept or "json" in content_type

    if is_json:
        try:
            body = await request.json()
            batch_id = batch_id or body.get("batch_id")
            company = company or body.get("company")
            overrides = body.get("overrides", "")
        except Exception:
            pass

    override_ids = set(int(x) for x in overrides.split(",") if x.strip().isdigit()) if isinstance(overrides, str) else set()

    if not batch_id:
        if is_json:
            return JSONResponse({"error": "no_batch", "message": "Select a batch to run payroll"}, status_code=400)
        redirect = f"/summary/?company={company}" if company else "/summary/"
        return RedirectResponse(url=redirect + "&error=no_batch", status_code=303)

    companies = _get_companies(db)
    selected_company = company or (companies[0] if companies else None)

    # Run with auto_save=True to commit balances
    _build_summary(db, company=selected_company, batch_id=batch_id, auto_save=True, override_ids=override_ids)

    if is_json:
        return JSONResponse({"ok": True, "batch_id": batch_id, "company": selected_company})

    redirect = f"/summary/?company={company}&batch_id={batch_id}&ran=1" if company else f"/summary/?batch_id={batch_id}&ran=1"
    return RedirectResponse(url=redirect, status_code=303)


# ── Excel export ──────────────────────────────────────────────────────────────

@router.get("/export/excel")
def summary_excel(
    company: str | None = Query(None),
    batch_id: int | None = Query(None),
    start: date | None = Query(None),
    end: date | None = Query(None),
    db: Session = Depends(get_db),
):
    # Always resolve to a single batch — never dump all-time data.
    # If batch_id wasn't provided, find the most recent batch for the company.
    if not batch_id and not start and not end:
        q = db.query(PayrollBatch).order_by(PayrollBatch.period_start.desc())
        if company:
            q = q.filter(PayrollBatch.company_name == company)
        latest = q.first()
        if latest:
            batch_id = latest.payroll_batch_id

    data = _build_summary(db, company=company, batch_id=batch_id, start=start, end=end)
    rows = data["rows"]
    totals = data["totals"]

    # Look up batch for period info
    batch = None
    if batch_id:
        batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == batch_id).first()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll Summary"

    # Company-aware colors
    co = (company or "").lower()
    if "acumen" in co or "first" in co:
        header_fill = PatternFill("solid", fgColor="4A1525")
        totals_fill = PatternFill("solid", fgColor="9B2C3D")
    elif "maz" in co or "ever" in co:
        header_fill = PatternFill("solid", fgColor="0F1D3A")
        totals_fill = PatternFill("solid", fgColor="1E3A6E")
    else:
        header_fill = PatternFill("solid", fgColor="0F1729")
        totals_fill = PatternFill("solid", fgColor="1E3A5F")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    money_fmt = '"$"#,##0.00'
    col_count = len(COLUMNS)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws["A1"]
    title_cell.value = f"{company or 'All Companies'} — Payroll Summary"
    title_color = "4A1525" if ("acumen" in co or "first" in co) else "0F1D3A" if ("maz" in co or "ever" in co) else "0F1729"
    title_cell.font = Font(bold=True, size=14, color=title_color)
    title_cell.alignment = left
    ws.row_dimensions[1].height = 24

    # Period row
    period_str = ""
    if batch and batch.week_start and batch.week_end:
        period_str = f"Period: {batch.week_start.strftime('%b %d, %Y')} – {batch.week_end.strftime('%b %d, %Y')}"
    elif batch and batch.period_start and batch.period_end:
        period_str = f"Period: {batch.period_start.strftime('%b %d, %Y')} – {batch.period_end.strftime('%b %d, %Y')}"
    elif start or end:
        period_str = f"Period: {start or '—'} to {end or '—'}"

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws["A2"].value = period_str
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws.row_dimensions[2].height = 16
    ws.append([])  # blank row 3
    header_row = 4

    # Header
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.row_dimensions[header_row].height = 22

    # Data rows
    row_fill_even = PatternFill("solid", fgColor="F8FAFC")
    row_fill_odd  = PatternFill("solid", fgColor="FFFFFF")
    # Money columns: Partner Pays(5), Driver Pay(6), Deduction(7), Carried Over(9), Paid This Period(10)
    money_cols = {5, 6, 7, 9, 10}

    for i, r in enumerate(rows):
        row_idx = header_row + 1 + i
        fill = row_fill_even if i % 2 == 0 else row_fill_odd
        carried = r["from_last_period"] if r["withheld"] else 0.0
        paid = r["pay_this_period"]
        vals = [
            r["person"],                            # Driver Name
            r["code"],                              # Pay Code
            r["rides"],                             # Rides
            r["miles"],                             # Miles
            r["partner_pays"],                      # Partner Pays
            r["driver_pay"],                        # Driver Pay
            r["deduction"],                         # Deduction
            "Yes" if r["withheld"] else "No",       # Withheld (Y/N)
            round(carried, 2) if r["withheld"] else 0.0,  # Carried Over
            paid,                                   # Paid This Period
        ]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = right if col_idx >= 3 else left
            if col_idx in money_cols and isinstance(val, (int, float)):
                cell.number_format = money_fmt

    # Totals
    totals_row_num = header_row + 1 + len(rows)
    totals_font = Font(bold=True, color="FFFFFF", size=11)
    totals_vals = [
        "TOTALS", "",
        totals["rides"], totals["miles"],
        totals["partner_pays"], totals["driver_pay"], totals["deduction"],
        "", totals["carried_over"], totals["pay_this_period"],
    ]
    for col_idx, val in enumerate(totals_vals, start=1):
        cell = ws.cell(row=totals_row_num, column=col_idx, value=val)
        cell.fill = totals_fill
        cell.font = totals_font
        cell.alignment = right if col_idx >= 3 else left
        if col_idx in money_cols and isinstance(val, (int, float)):
            cell.number_format = money_fmt

    # Column widths
    for i, w in enumerate([30, 12, 8, 10, 14, 14, 12, 14, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    co_slug = (company or "all").lower().replace(" ", "_")
    filename = f"zpay_{co_slug}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── PDF export ────────────────────────────────────────────────────────────────

@router.get("/export/pdf")
def summary_pdf(
    company: str | None = Query(None),
    batch_id: int | None = Query(None),
    start: date | None = Query(None),
    end: date | None = Query(None),
    db: Session = Depends(get_db),
):
    data = _build_summary(db, company=company, batch_id=batch_id, start=start, end=end)
    rows = data["rows"]
    totals = data["totals"]

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=(8.5 * inch, 11 * inch),
        leftMargin=0.5*inch, rightMargin=0.5*inch,
        topMargin=0.5*inch, bottomMargin=0.5*inch,
    )
    styles = getSampleStyleSheet()
    elements = []

    title_style = styles["Title"]
    title_style.textColor = colors.HexColor("#0f1729")
    elements.append(Paragraph(f"{company or 'All Companies'} — Payroll Summary", title_style))

    sub = f"Generated {datetime.now().strftime('%m/%d/%Y')}"
    if start or end:
        sub += f"  ·  {start or '—'} to {end or '—'}"
    elements.append(Paragraph(sub, styles["Normal"]))
    elements.append(Spacer(1, 14))

    table_data = [COLUMNS]
    for r in rows:
        carried = r["from_last_period"] if r["withheld"] else 0.0
        table_data.append([
            r["person"],
            r["code"] or "—",
            str(r["rides"]),
            f"{r['miles']:.1f}",
            f"${r['partner_pays']:,.2f}",
            f"${r['driver_pay']:,.2f}",
            f"${r['deduction']:,.2f}",
            "Yes" if r["withheld"] else "No",
            f"${carried:,.2f}" if r["withheld"] else "—",
            f"${r['pay_this_period']:,.2f}",
        ])

    table_data.append([
        "TOTALS", "",
        str(totals["rides"]),
        f"{totals['miles']:.1f}",
        f"${totals['partner_pays']:,.2f}",
        f"${totals['driver_pay']:,.2f}",
        f"${totals['deduction']:,.2f}",
        "",
        f"${totals['carried_over']:,.2f}",
        f"${totals['pay_this_period']:,.2f}",
    ])

    col_widths = [1.4*inch, 0.6*inch, 0.4*inch, 0.5*inch, 0.75*inch, 0.75*inch, 0.65*inch, 0.6*inch, 0.7*inch, 0.8*inch]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        # Header
        ("BACKGROUND",    (0,0), (-1,0), colors.HexColor("#0f1729")),
        ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0), 8),
        ("ALIGN",         (0,0), (-1,0), "CENTER"),
        ("BOTTOMPADDING", (0,0), (-1,0), 8),
        ("TOPPADDING",    (0,0), (-1,0), 8),
        # Body
        ("FONTNAME",      (0,1), (-1,-2), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-2), 8),
        ("ROWBACKGROUNDS",(0,1), (-1,-2), [colors.white, colors.HexColor("#f1f5f9")]),
        ("ALIGN",         (2,1), (-1,-2), "RIGHT"),
        ("ALIGN",         (0,1), (1,-2), "LEFT"),
        ("BOTTOMPADDING", (0,1), (-1,-2), 5),
        ("TOPPADDING",    (0,1), (-1,-2), 5),
        # Paid This Period column bold
        ("FONTNAME",      (-1,1), (-1,-2), "Helvetica-Bold"),
        # Totals row
        ("BACKGROUND",    (0,-1), (-1,-1), colors.HexColor("#1e3a5f")),
        ("TEXTCOLOR",     (0,-1), (-1,-1), colors.white),
        ("FONTNAME",      (0,-1), (-1,-1), "Helvetica-Bold"),
        ("ALIGN",         (3,-1), (-1,-1), "RIGHT"),
        ("TOPPADDING",    (0,-1), (-1,-1), 8),
        ("BOTTOMPADDING", (0,-1), (-1,-1), 8),
        # Grid
        ("LINEBELOW",     (0,0), (-1,0), 0.5, colors.HexColor("#1e3a5f")),
        ("LINEABOVE",     (0,-1), (-1,-1), 0.5, colors.HexColor("#3b82f6")),
        ("GRID",          (0,0), (-1,-1), 0.25, colors.HexColor("#e2e8f0")),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)

    co_slug = (company or "all").lower().replace(" ", "_")
    filename = f"zpay_{co_slug}_{date.today().isoformat()}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Paychex CSV export ────────────────────────────────────────────────────────
# Standard Payroll Import (SPI) format — 16 columns, no header row.
# Cols: Client ID, Worker ID, Org, Job#, Pay Component, Rate, Rate#,
#       Hours, Units, Line Date, Amount, Check Seq#, Override State,
#       Override Local, Override Local Juris, Labor Assignment

PAYCHEX_CLIENT_IDS = {
    "acumen": "70189220",   # Acumen International / FirstAlt
    "maz": "17182126",      # Maz Services / EverDriven
}

@router.get("/export/paycheck-csv")
def export_paycheck_csv(
    payroll_batch_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """
    Export drivers paid this period in Paychex SPI format.
    Only includes drivers who are NOT withheld (pay_this_period > 0).
    """
    # Track that Paychex CSV was exported (for workflow gate)
    from datetime import timezone as _tz
    batch = db.query(PayrollBatch).filter(PayrollBatch.payroll_batch_id == payroll_batch_id).first()
    if batch and not batch.paychex_exported_at:
        batch.paychex_exported_at = datetime.now(_tz.utc)
        db.commit()

    # Pick the right client ID based on batch company
    co = (batch.company_name or "").lower() if batch else ""
    if "maz" in co or "ever" in co:
        paychex_client_id = PAYCHEX_CLIENT_IDS["maz"]
    else:
        paychex_client_id = PAYCHEX_CLIENT_IDS["acumen"]

    data = _build_summary(db, batch_id=payroll_batch_id)
    rows = data["rows"]

    # Pull Person records to get paycheck_code (= Paychex Worker ID)
    person_ids = [r["person_id"] for r in rows if not r["withheld"] and r["pay_this_period"] > 0]
    persons = {
        p.person_id: p
        for p in db.query(Person).filter(Person.person_id.in_(person_ids)).all()
    } if person_ids else {}

    output = io.StringIO()
    writer = csv.writer(output)
    # NO header row — SPI format uses fixed column positions

    for r in rows:
        if r["withheld"] or r["pay_this_period"] <= 0:
            continue

        person = persons.get(r["person_id"])
        # Worker ID: Maz batches use paycheck_code_maz (Maz Paychex ID);
        # all others use paycheck_code. Fall back to paycheck_code if maz field unset.
        worker_id = ""
        if person:
            if "maz" in co or "ever" in co:
                worker_id = getattr(person, "paycheck_code_maz", None) or getattr(person, "paycheck_code", None) or ""
            else:
                worker_id = getattr(person, "paycheck_code", None) or ""

        if not worker_id:
            continue  # Skip drivers without a Paychex ID — can't import them

        writer.writerow([
            paychex_client_id,              # Col A: Client ID
            worker_id,                       # Col B: Worker ID
            "",                              # Col C: Org (blank)
            "",                              # Col D: Job Number (blank)
            "1099-NEC",                      # Col E: Pay Component (exact name in Paychex)
            "",                              # Col F: Rate (blank)
            "",                              # Col G: Rate Number (blank)
            "",                              # Col H: Hours (blank — flat amount, not hourly)
            "",                              # Col I: Units (blank)
            "",                              # Col J: Line Date (blank)
            f"{r['pay_this_period']:.2f}",  # Col K: Amount
            "",                              # Col L: Check Seq Number (blank)
            "",                  # Col M: Override State (blank)
            "",                  # Col N: Override Local (blank)
            "",                  # Col O: Override Local Jurisdiction (blank)
            "",                  # Col P: Labor Assignment (blank)
        ])

    output.seek(0)
    filename = f"paychex_spi_{date.today().isoformat()}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

"""
audit_xlsx_integrity.py — Read-only xlsx integrity audit for a payroll batch.

Usage
-----
    python -m backend.scripts.audit_xlsx_integrity --batch 95
    python -m backend.scripts.audit_xlsx_integrity --batch 95 --against /path/to/mom.xlsx

What it does
------------
1. Queries the DB for the batch and regenerates the xlsx bytes using the same
   builder the download route uses (_build_maz_xlsx_bytes for Maz batches,
   or the inline Acumen builder for FA batches).
2. Runs _build_summary independently to get the canonical per-driver numbers.
3. Loads the generated xlsx via openpyxl and reads the Payroll Summary tab.
4. Diffs each driver row: xlsx value vs _build_summary value.
5. Optionally diffs against mom's xlsx file when --against is supplied.

Exit codes
----------
  0  everything matches
  1  drift detected between xlsx and DB, or --against diff found

This script is READ-ONLY — it never mutates the DB or any file on disk.
The generated xlsx is written to a NamedTemporaryFile and deleted after diff.
"""

from __future__ import annotations

import argparse
import io
import os
import sys
import tempfile
from pathlib import Path
from typing import Any

# ── Project root on sys.path ─────────────────────────────────────────────────
_PROJECT_ROOT = str(Path(__file__).resolve().parents[2])
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

# Required env vars before importing anything that touches the DB
os.environ.setdefault("ZPAY_SECRET_KEY", "audit-script-placeholder-not-used-for-auth")

import openpyxl  # noqa: E402
from sqlalchemy import create_engine  # noqa: E402
from sqlalchemy.orm import sessionmaker  # noqa: E402


# ── Column index constants (1-based, matching _MOM_HEADERS order) ────────────
_COL_DRIVER_NAME   = 1   # "Driver Name"
_COL_PAY_CODE      = 2   # "Pay Code"
_COL_RIDES         = 3   # "Rides"
_COL_MILES         = 4   # "Miles"
_COL_PARTNER_PAYS  = 5   # "Partner Pays"
_COL_DRIVER_PAY    = 6   # "Driver Pay"
_COL_DEDUCTION     = 7   # "Deduction"
_COL_WITHHELD      = 8   # "Withheld (Y/N)"
_COL_CARRIED_OVER  = 9   # "Carried Over"
_COL_PAID          = 10  # "Paid This Period"


def _fmt(v: Any) -> str:
    """Format a cell value for display."""
    if v is None:
        return "None"
    if isinstance(v, float):
        return f"{v:.2f}"
    return str(v)


def _cell_float(ws, row: int, col: int) -> float:
    """Read a numeric cell, returning 0.0 on None/non-numeric."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return 0.0
    try:
        return round(float(val), 2)
    except (TypeError, ValueError):
        return 0.0


def _cell_str(ws, row: int, col: int) -> str:
    val = ws.cell(row=row, column=col).value
    return "" if val is None else str(val).strip()


def _find_header_row(ws) -> int:
    """
    Find the row number of the Payroll Summary column-header row.

    The header row contains "Driver Name" in column A.
    Layout is always: R1=title, R2=period, R3=blank, R4=headers.
    We scan the first 10 rows defensively.
    """
    for row_num in range(1, 11):
        val = ws.cell(row=row_num, column=_COL_DRIVER_NAME).value
        if val and str(val).strip() == "Driver Name":
            return row_num
    raise ValueError("Could not locate 'Driver Name' header row in Payroll Summary tab")


def _read_payroll_summary_tab(ws) -> dict[str, dict]:
    """
    Parse the Payroll Summary tab into a per-driver dict.

    Returns:
        {driver_name: {partner_pays, driver_pay, deduction, carried_over, paid_this_period, withheld}}

    Stops reading when it hits the TOTALS row or a blank driver-name cell
    (which signals the start of the Paychex/Paid/Unpaid sections).
    """
    header_row = _find_header_row(ws)
    data_start = header_row + 1

    drivers: dict[str, dict] = {}
    for row_num in range(data_start, ws.max_row + 1):
        name = _cell_str(ws, row_num, _COL_DRIVER_NAME)
        if not name or name.upper() == "TOTALS":
            break
        drivers[name] = {
            "partner_pays":      _cell_float(ws, row_num, _COL_PARTNER_PAYS),
            "driver_pay":        _cell_float(ws, row_num, _COL_DRIVER_PAY),
            "deduction":         _cell_float(ws, row_num, _COL_DEDUCTION),
            "carried_over":      _cell_float(ws, row_num, _COL_CARRIED_OVER),
            "paid_this_period":  _cell_float(ws, row_num, _COL_PAID),
            "withheld":          _cell_str(ws, row_num, _COL_WITHHELD),
            "rides":             ws.cell(row=row_num, column=_COL_RIDES).value,
        }
    return drivers


def _summary_rows_to_dict(rows: list[dict]) -> dict[str, dict]:
    """
    Convert _build_summary rows list to the same shape as _read_payroll_summary_tab.
    """
    result: dict[str, dict] = {}
    for r in rows:
        name = (r.get("person") or "").strip()
        if not name:
            continue
        result[name] = {
            "partner_pays":     round(float(r.get("partner_pays") or 0), 2),
            "driver_pay":       round(float(r.get("driver_pay") or 0), 2),
            "deduction":        round(float(r.get("deduction") or 0), 2),
            "carried_over":     round(float(r.get("from_last_period") or 0), 2),
            "paid_this_period": round(float(r.get("pay_this_period") or 0), 2),
            "withheld":         "Yes" if r.get("withheld") else "No",
            "rides":            int(r.get("rides") or 0),
        }
    return result


def _diff_dicts(
    label_a: str,
    label_b: str,
    a: dict[str, dict],
    b: dict[str, dict],
    fields: list[str],
) -> list[str]:
    """
    Compare two per-driver dicts field by field.

    Returns a list of human-readable diff lines.  Empty list = no drift.
    """
    all_names = sorted(set(a) | set(b))
    diffs: list[str] = []

    for name in all_names:
        if name not in a:
            diffs.append(f"  [ONLY IN {label_b}] {name}")
            continue
        if name not in b:
            diffs.append(f"  [ONLY IN {label_a}] {name}")
            continue
        row_a = a[name]
        row_b = b[name]
        field_diffs = []
        for field in fields:
            va = row_a.get(field)
            vb = row_b.get(field)
            # Numeric comparison with a small tolerance for float representation
            if isinstance(va, float) and isinstance(vb, float):
                if abs(va - vb) > 0.005:
                    field_diffs.append(
                        f"    {field}: {label_a}={_fmt(va)}  {label_b}={_fmt(vb)}"
                    )
            elif str(va) != str(vb):
                field_diffs.append(
                    f"    {field}: {label_a}={_fmt(va)}  {label_b}={_fmt(vb)}"
                )
        if field_diffs:
            diffs.append(f"  ⚠  {name}")
            diffs.extend(field_diffs)
        else:
            diffs.append(f"  ✓  {name}")

    return diffs


def _generate_xlsx_bytes_for_batch(db, batch) -> bytes:
    """
    Regenerate xlsx bytes for the given batch using the production builder.

    Maz → _build_maz_xlsx_bytes (delegates to _build_summary internally).
    Acumen → inline builder that mirrors workflow_export_excel's Acumen path.
    """
    import io as _io
    import openpyxl as _openpyxl

    from backend.routes.workflow import (
        _build_maz_xlsx_bytes,
        _build_payroll_summary_tab,
        _build_sp_pay_summary_tab,
        _build_sp_itemized_tab,
    )
    from backend.routes.summary import _build_summary
    from backend.db.models import Ride, Person

    batch_id = batch.payroll_batch_id
    source = (batch.source or "").lower()

    if source == "maz":
        return _build_maz_xlsx_bytes(db, batch)

    # Acumen path — mirror workflow_export_excel logic
    _override_ids = None
    _manual_withhold_ids = None
    try:
        from sqlalchemy import text as _sql_text
        _override_rows = db.execute(
            _sql_text("SELECT person_id FROM payroll_withheld_override WHERE batch_id = :b"),
            {"b": batch_id},
        ).fetchall()
        _override_ids = {r[0] for r in _override_rows} or None
        _manual_rows = db.execute(
            _sql_text("SELECT person_id FROM payroll_manual_withhold"),
        ).fetchall()
        _manual_withhold_ids = {r[0] for r in _manual_rows} or None
    except Exception:
        db.rollback()

    data = _build_summary(
        db,
        batch_id=batch_id,
        override_ids=_override_ids,
        manual_withhold_ids=_manual_withhold_ids,
    )
    rows = data["rows"]
    totals = data["totals"]

    trip_rows_raw = (
        db.query(
            Person.full_name.label("person"),
            Person.paycheck_code.label("code"),
            Ride.ride_start_ts,
            Ride.source_ref,
            Ride.service_name,
            Ride.z_rate_source,
            Ride.miles,
            Ride.gross_pay,
            Ride.deduction,
            Ride.net_pay,
        )
        .join(Person, Person.person_id == Ride.person_id)
        .filter(Ride.payroll_batch_id == batch_id, Ride.removed_at.is_(None))
        .order_by(Person.full_name.asc(), Ride.ride_start_ts.asc())
        .all()
    )

    trip_rows = []
    for t in trip_rows_raw:
        trip_date = t.ride_start_ts.date() if t.ride_start_ts else None
        trip_rows.append({
            "person": t.person,
            "code": t.code or "-",
            "date": trip_date,
            "trip_code": t.source_ref or "",
            "trip_name": t.service_name or "",
            "cancellation_reason": "Canceled" if t.z_rate_source == "canceled_trip" else None,
            "miles": float(t.miles or 0),
            "gross_pay": float(t.gross_pay or 0),
            "deduction": float(t.deduction or 0),
            "net_pay": float(t.net_pay or 0),
        })

    llc_title = "FirstAlt — Payroll Summary"
    wb = _openpyxl.Workbook()
    ws1 = wb.active
    _build_sp_pay_summary_tab(ws1, batch, rows)
    ws2 = wb.create_sheet("SP ITEMIZED REPORT")
    _build_sp_itemized_tab(ws2, batch, trip_rows)
    ws_ps = wb.create_sheet("Payroll Summary")
    _build_payroll_summary_tab(ws_ps, batch, rows, totals, llc_title, db=db)

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def run_audit(batch_id: int, against_path: str | None = None) -> int:
    """
    Main audit logic.

    Returns 0 on clean, 1 on drift.
    """
    database_url = os.environ.get("DATABASE_URL") or os.environ.get("DATABASE_PUBLIC_URL")
    if not database_url:
        print("ERROR: DATABASE_URL or DATABASE_PUBLIC_URL env var required.", file=sys.stderr)
        return 1

    engine = create_engine(database_url)
    Session = sessionmaker(bind=engine)
    db = Session()

    try:
        from backend.db.models import PayrollBatch
        from backend.routes.summary import _build_summary

        batch = db.query(PayrollBatch).filter(
            PayrollBatch.payroll_batch_id == batch_id
        ).first()
        if batch is None:
            print(f"ERROR: Batch {batch_id} not found in DB.", file=sys.stderr)
            return 1

        source = (batch.source or "").lower()
        print(f"\nAudit — Batch {batch_id}  source={source}  status={batch.status}")
        print("=" * 64)

        # Step 1: regenerate xlsx bytes
        print("\n[1] Regenerating xlsx via production builder …")
        xlsx_bytes = _generate_xlsx_bytes_for_batch(db, batch)
        print(f"    Generated {len(xlsx_bytes):,} bytes")

        # Step 2: run _build_summary independently
        print("\n[2] Running _build_summary (canonical DB numbers) …")
        summary_data = _build_summary(db, batch_id=batch_id)
        summary_by_driver = _summary_rows_to_dict(summary_data["rows"])
        print(f"    {len(summary_by_driver)} drivers in DB summary")

        # Step 3: read xlsx Payroll Summary tab
        print("\n[3] Parsing xlsx Payroll Summary tab …")
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        if "Payroll Summary" not in wb.sheetnames:
            print("ERROR: 'Payroll Summary' tab not found in generated xlsx.", file=sys.stderr)
            return 1
        ws = wb["Payroll Summary"]
        xlsx_by_driver = _read_payroll_summary_tab(ws)
        print(f"    {len(xlsx_by_driver)} drivers in xlsx")

        # Step 4: diff xlsx vs DB
        print("\n[4] Diffing xlsx vs DB _build_summary …\n")
        _COMPARE_FIELDS = [
            "partner_pays", "driver_pay", "deduction",
            "carried_over", "paid_this_period", "withheld",
        ]
        diffs_db = _diff_dicts("xlsx", "DB", xlsx_by_driver, summary_by_driver, _COMPARE_FIELDS)
        for line in diffs_db:
            print(line)

        has_drift = any("⚠" in line or "ONLY IN" in line for line in diffs_db)
        db_drift_count = sum(1 for line in diffs_db if "⚠" in line or "ONLY IN" in line)
        ok_count = sum(1 for line in diffs_db if "✓" in line)

        print(f"\n    Result: {ok_count} drivers match, {db_drift_count} drivers differ")

        # Step 5: optional --against diff
        against_drift = False
        if against_path:
            print(f"\n[5] Diffing against mom's xlsx: {against_path} …\n")
            try:
                mom_wb = openpyxl.load_workbook(against_path, data_only=True)
                # Try common tab name variants
                mom_ws = None
                for candidate in ("Payroll Summary", "Payroll  ", "Payroll", "Sheet1"):
                    if candidate in mom_wb.sheetnames:
                        mom_ws = mom_wb[candidate]
                        print(f"    Using tab: '{candidate}'")
                        break
                if mom_ws is None:
                    print(
                        f"    WARNING: could not find Payroll Summary tab in {against_path}.\n"
                        f"    Available sheets: {mom_wb.sheetnames}"
                    )
                else:
                    mom_by_driver = _read_payroll_summary_tab(mom_ws)
                    print(f"    {len(mom_by_driver)} drivers in mom's xlsx\n")
                    diffs_mom = _diff_dicts(
                        "DB", "mom's xlsx",
                        summary_by_driver, mom_by_driver,
                        _COMPARE_FIELDS,
                    )
                    for line in diffs_mom:
                        print(line)
                    against_drift_count = sum(
                        1 for line in diffs_mom if "⚠" in line or "ONLY IN" in line
                    )
                    against_ok = sum(1 for line in diffs_mom if "✓" in line)
                    print(
                        f"\n    Result: {against_ok} drivers match, "
                        f"{against_drift_count} drivers differ vs mom's xlsx"
                    )
                    against_drift = against_drift_count > 0
            except FileNotFoundError:
                print(f"    ERROR: file not found: {against_path}", file=sys.stderr)
                return 1
            except Exception as exc:
                print(f"    ERROR reading mom's xlsx: {exc}", file=sys.stderr)
                return 1

        print()
        if has_drift or against_drift:
            print("RESULT: DRIFT DETECTED — exit 1")
            return 1
        else:
            print("RESULT: ALL MATCH — exit 0")
            return 0

    finally:
        db.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Read-only xlsx integrity audit for a Z-Pay payroll batch.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples
--------
  python -m backend.scripts.audit_xlsx_integrity --batch 95

  python -m backend.scripts.audit_xlsx_integrity --batch 95 \\
    --against ~/Downloads/Prod_SP_Acumen_International_05152026.xlsx

The script never writes to the DB or any file on disk.
""",
    )
    parser.add_argument(
        "--batch",
        type=int,
        required=True,
        metavar="BATCH_ID",
        help="Payroll batch ID to audit (e.g. 95)",
    )
    parser.add_argument(
        "--against",
        type=str,
        default=None,
        metavar="PATH",
        help="Optional: path to mom's xlsx file for column-by-column comparison",
    )
    args = parser.parse_args()
    sys.exit(run_audit(args.batch, args.against))


if __name__ == "__main__":
    main()

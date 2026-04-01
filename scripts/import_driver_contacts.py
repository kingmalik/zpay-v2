"""
Import driver contact info from Excel files into the Person table.

Source files (copied into container at /tmp/):
  - /tmp/acumen_drivers.xlsx  — Driver's TIN sheet: Driver Name, TIN, Email, Phone
  - /tmp/maz_drivers.xlsx     — Sheet1: Driver Name, Email, Status

Logic:
  - Fuzzy-match Driver Name -> Person.full_name (threshold 80%)
  - Update email, phone, paycheck_code (TIN) only if currently NULL in DB
  - Never overwrite existing data
"""

import os
import sys
import re
import unicodedata
from difflib import SequenceMatcher

import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session

# ── DB connection ─────────────────────────────────────────────────────────────
DATABASE_URL = os.environ.get("DATABASE_URL", "postgresql://zpay:zpay@db:5432/zpay")
engine = create_engine(DATABASE_URL)


# ── Name normalisation ────────────────────────────────────────────────────────
def _norm(name: str) -> str:
    if not name:
        return ""
    name = str(name).strip()
    # Remove sub-driver suffixes like "- Sub#2", "- Sub#3 Morning only"
    name = re.sub(r"\s*-\s*Sub#?\d+.*$", "", name, flags=re.IGNORECASE)
    # Strip unicode accents → ASCII equivalents
    name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    # Collapse whitespace, lowercase
    return " ".join(name.lower().split())


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, _norm(a), _norm(b)).ratio()


# ── Load all persons from DB ──────────────────────────────────────────────────
def load_persons(session: Session):
    rows = session.execute(
        text("SELECT person_id, full_name, email, phone, paycheck_code FROM person ORDER BY person_id")
    ).fetchall()
    return [dict(r._mapping) for r in rows]


def find_best_match(name: str, persons: list[dict], threshold=0.80):
    best = None
    best_score = 0.0
    for p in persons:
        score = _similarity(name, p["full_name"])
        if score > best_score:
            best_score = score
            best = p
    if best_score >= threshold:
        return best, best_score
    return None, best_score


# ── Read Excel rows ───────────────────────────────────────────────────────────
def read_acumen(path: str):
    """
    Reads the TIN sheet from the Acumen file.
    Columns: Driver Name | TIN | Email | LIC# | Phone
    TIN is stored as paycheck_code.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["TIN"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name:
            continue
        tin   = str(row[1]).strip() if row[1] else None
        email = str(row[2]).strip().lstrip("\t") if row[2] else None
        phone = str(row[4]).strip() if row[4] else None
        rows.append({"name": str(name).strip(), "email": email, "phone": phone, "paycheck_code": tin})
    wb.close()
    return rows


def read_maz(path: str):
    """
    Reads Sheet1 from the Maz file.
    Columns: Driver Name | Email | Status
    No phone or TIN.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name:
            continue
        email = str(row[1]).strip().lstrip("\t") if row[1] else None
        rows.append({"name": str(name).strip(), "email": email, "phone": None, "paycheck_code": None})
    wb.close()
    return rows


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    acumen_path = "/tmp/acumen_drivers.xlsx"
    maz_path    = "/tmp/maz_drivers.xlsx"

    for p in [acumen_path, maz_path]:
        if not os.path.exists(p):
            print(f"ERROR: File not found: {p}")
            print("Copy files into container first:")
            print(f'  docker cp "Downloads/Driver\'s TIN (2).xlsx" z-pay-app-1:/tmp/acumen_drivers.xlsx')
            print(f'  docker cp "Downloads/MAZ drivers (3).xlsx"   z-pay-app-1:/tmp/maz_drivers.xlsx')
            sys.exit(1)

    acumen_rows = read_acumen(acumen_path)
    maz_rows    = read_maz(maz_path)
    print(f"Loaded {len(acumen_rows)} Acumen rows, {len(maz_rows)} Maz rows")

    # Deduplicate by name (Acumen wins — it has TIN + phone)
    seen_names: dict[str, dict] = {}
    for row in acumen_rows:
        key = _norm(row["name"])
        seen_names[key] = row

    for row in maz_rows:
        key = _norm(row["name"])
        if key not in seen_names:
            seen_names[key] = row
        else:
            # Maz may have email where Acumen didn't
            existing = seen_names[key]
            if not existing.get("email") and row.get("email"):
                existing["email"] = row["email"]

    all_rows = list(seen_names.values())
    print(f"Unique driver names to process: {len(all_rows)}")

    updated = 0
    skipped = 0
    no_match = []

    with Session(engine) as session:
        persons = load_persons(session)

        for row in all_rows:
            person, score = find_best_match(row["name"], persons)

            if person is None:
                print(f"  No match ({score:.0%}): {row['name']}")
                no_match.append(row["name"])
                continue

            pid = person["person_id"]
            updates = {}

            if row.get("email") and not person["email"]:
                updates["email"] = row["email"]
            if row.get("phone") and not person["phone"]:
                updates["phone"] = row["phone"]
            if row.get("paycheck_code") and not person["paycheck_code"]:
                updates["paycheck_code"] = row["paycheck_code"]

            if not updates:
                skipped += 1
                continue

            set_clause = ", ".join(f"{k} = :{k}" for k in updates)
            updates["pid"] = pid
            session.execute(
                text(f"UPDATE person SET {set_clause} WHERE person_id = :pid"),
                updates,
            )

            fields = ", ".join(updates.keys() - {"pid"})
            print(f"  Updated ({score:.0%}): {person['full_name']} — {fields}")
            updated += 1

        session.commit()

    print()
    print(f"Summary: {updated} updated, {skipped} skipped (already had data), {len(no_match)} unmatched")
    if no_match:
        print("\nUnmatched drivers:")
        for name in no_match:
            print(f"  - {name}")


if __name__ == "__main__":
    main()

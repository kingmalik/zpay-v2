#!/usr/bin/env python3
"""
Z-Pay Truth Ledger — W1 through W14
Reads mom's weekly Excel files from Wheels of Unity Drive.
Produces:
  ~/Library/Application Support/zpay-backups/audit/truth_ledger_W1_W14.csv
  ~/Library/Application Support/zpay-backups/audit/truth_summary.md

Read-only on all source files. Zero writes elsewhere.
"""

import os
import csv
import re
import subprocess
import traceback
from pathlib import Path
from typing import Optional

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
WOU = Path(
    "/Users/malikmilion/Library/CloudStorage/"
    "GoogleDrive-milionmalik@gmail.com/"
    ".shortcut-targets-by-id/"
    "1rjKNX3VW7IXqaWGnwK6jAqhqwB0Vnp2C/Wheels of Unity"
)
FA_BASE = WOU / "Payroll/Acumen/2026"
MAZ_BASE = WOU / "Payroll/Maz/Reports 2026"

OUT_DIR = Path.home() / "Library/Application Support/zpay-backups/audit"
OUT_CSV = OUT_DIR / "truth_ledger_W1_W14.csv"
OUT_MD = OUT_DIR / "truth_summary.md"

WEEKS = list(range(1, 15))

# ---------------------------------------------------------------------------
# DB lookup — paycheck_code → person info (read-only, railway run)
# ---------------------------------------------------------------------------

def fetch_person_map():
    """Pull paycheck_code and paycheck_code_maz from prod via railway run.
    Returns two dicts:
      fa_map:  paycheck_code (int) -> {name, paycheck_code_maz}
      maz_map: paycheck_code_maz (int) -> {name, paycheck_code}
    """
    fa_map = {}
    maz_map = {}
    try:
        result = subprocess.run(
            [
                "railway", "run",
                "--service", "zpay-backend",
                "python3", "-c",
                (
                    "import os, psycopg2, json; "
                    "conn=psycopg2.connect(os.environ['DATABASE_URL']); "
                    "cur=conn.cursor(); "
                    "cur.execute('SELECT name, paycheck_code, paycheck_code_maz FROM person WHERE paycheck_code IS NOT NULL OR paycheck_code_maz IS NOT NULL'); "
                    "rows=cur.fetchall(); "
                    "print(json.dumps([{'name':r[0],'pc':r[1],'pc_maz':r[2]} for r in rows])); "
                    "conn.close()"
                ),
            ],
            capture_output=True,
            text=True,
            timeout=60,
            cwd="/Users/malikmilion/Desktop/zpay-v2-fresh",
        )
        if result.returncode == 0:
            import json
            # Find the JSON array in stdout (railway may emit other lines)
            for line in result.stdout.strip().splitlines():
                line = line.strip()
                if line.startswith("["):
                    rows = json.loads(line)
                    for r in rows:
                        name = r.get("name", "")
                        pc = r.get("pc")
                        pc_maz = r.get("pc_maz")
                        if pc:
                            fa_map[int(pc)] = {"name": name, "paycheck_code_maz": pc_maz}
                        if pc_maz:
                            maz_map[int(pc_maz)] = {"name": name, "paycheck_code": pc}
                    break
    except Exception as e:
        print(f"[WARN] DB lookup failed: {e} — proceeding without DB name enrichment")
    return fa_map, maz_map


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def find_xlsx(week_dir: Path) -> Optional[Path]:
    """Return first .xlsx in directory, ignoring temp files."""
    for p in sorted(week_dir.glob("*.xlsx")):
        if not p.name.startswith("~$"):
            return p
    return None


def to_float(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "")
    if s in ("", "-", "None", "nan", "N/A"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def is_withheld(v) -> bool:
    if v is None:
        return False
    return str(v).strip().lower() in ("withheld", "yes", "y", "1")


def normalize_code(v) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(str(v).strip().replace(",", "").split(".")[0])
    except (ValueError, TypeError):
        return None


# These signal the absolute end of driver data — accounting section begins
HARD_STOP_KEYWORDS = {
    "paychex flex amound", "paychex flex amount",
    "unpaid on week", "unpaid on week ",
    "paid on week", "paid on week ", "paid on weeks",
}

# These are subtotal/total rows to skip (don't stop, just ignore)
SKIP_KEYWORDS = {
    "total", "total ", "totals", "grand total", "subtotal",
    "total ", "total",
}

HEADER_KEYWORDS = {
    "summary", "person", "driver", "driver name",
    "acumen international — payroll summary",
    "firstalt — payroll summary",
    "all companies — payroll summary",
    "details",
}


def is_hard_stop_row(row, name_col=0) -> bool:
    """True if this row signals the absolute end of the driver block (accounting section)."""
    name = row[name_col] if len(row) > name_col else None
    if name is None:
        return False
    name_s = str(name).strip().lower()
    return name_s in HARD_STOP_KEYWORDS


def is_stop_row(row, name_col=0) -> bool:
    """True if this row signals end of the driver data block (TOTALS in new-format files)."""
    name = row[name_col] if len(row) > name_col else None
    if name is None:
        return False
    name_s = str(name).strip().lower()
    # Only hard-stop for the accounting keywords in old-format
    return name_s in HARD_STOP_KEYWORDS


def is_data_row(row, name_col=0, code_col=1) -> bool:
    """True if row looks like a driver record (has a name that isn't a header/total)."""
    name = row[name_col] if len(row) > name_col else None
    if name is None:
        return False
    name_s = str(name).strip()
    if not name_s:
        return False
    name_lower = name_s.lower()
    if name_lower in HARD_STOP_KEYWORDS or name_lower in SKIP_KEYWORDS or name_lower in HEADER_KEYWORDS:
        return False
    if name_lower.startswith("period:"):
        return False
    return True


# ---------------------------------------------------------------------------
# FA parsers
# ---------------------------------------------------------------------------

def _fa_find_payroll_sheet(wb):
    """Return (sheet_name, family) where family is 'old', 'w11', or 'new'."""
    names = wb.sheetnames
    normalized = {s.strip().lower().replace("_", " "): s for s in names}

    # Priority: most specific first
    # W12-W14 style: 'Payroll_Summary' or 'Payroll Summary' (with rides/miles cols)
    # W11 style: 'Payroll Summary' (with Active Between / Days / Net Pay cols)
    # W1-W10 style: 'Payroll ' or 'Payroll  ' (with Person/Code/Payroll cols)

    for raw, original in normalized.items():
        if raw == "payroll summary" or raw == "payroll_summary":
            # Distinguish W11 vs W12+ by header content
            ws = wb[original]
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                if row[0] in ("Driver", "Person", "Driver Name"):
                    row_list = [str(c).strip().lower() if c else "" for c in row]
                    if "rides" in row_list or "miles" in row_list:
                        return original, "new"
                    else:
                        return original, "w11"
            # Default: if no driver header found yet, assume new
            return original, "new"

    # Old-style Payroll tab (with trailing spaces)
    for raw, original in normalized.items():
        if raw.startswith("payroll"):
            return original, "old"

    return None, None


def parse_fa_old(ws, week, filepath) -> list[dict]:
    """FA W1-W10: Person | Code | Payroll | ... | Unpaid/Pending | To Paid | Total"""
    rows_out = []
    header_found = False
    col_person = 0
    col_code = 1
    col_payroll = 2
    col_unpaid = None
    col_to_paid = None
    col_total = None

    all_rows = list(ws.iter_rows(values_only=True))

    # Find header row
    for i, row in enumerate(all_rows):
        row_s = [str(c).strip().lower() if c else "" for c in row]
        if "person" in row_s or "driver" in row_s:
            header_found = True
            # Map columns
            for j, h in enumerate(row_s):
                if h == "person" or h == "driver":
                    col_person = j
                elif h == "code":
                    col_code = j
                elif h in ("payroll",):
                    col_payroll = j
                elif "unpaid" in h or "pending" in h:
                    col_unpaid = j
                elif "to paid" in h or "to be paid" in h:
                    col_to_paid = j
                elif h in ("total", "total "):
                    col_total = j
            # Parse data rows from here; stop at first Total/Paychex row
            for dr in all_rows[i + 1:]:
                if is_stop_row(dr, col_person):
                    break  # end of driver block
                if not is_data_row(dr, col_person, col_code):
                    continue
                name = str(dr[col_person]).strip()
                code = normalize_code(dr[col_code] if len(dr) > col_code else None)
                payroll = to_float(dr[col_payroll] if len(dr) > col_payroll else None)
                unpaid = to_float(dr[col_unpaid] if col_unpaid and len(dr) > col_unpaid else None)
                to_paid = to_float(dr[col_to_paid] if col_to_paid and len(dr) > col_to_paid else None)
                total = to_float(dr[col_total] if col_total and len(dr) > col_total else None)

                # Skip rows with no numeric payroll amount
                if payroll is None and total is None:
                    continue

                rows_out.append({
                    "driver_name_in_file": name,
                    "paycheck_code": code,
                    "week": week,
                    "llc": "FA",
                    "payroll_amount": payroll,
                    "unpaid_pending": unpaid,
                    "to_paid": to_paid,
                    "total": total,
                    "source_file": str(filepath),
                })
            break

    return rows_out


def parse_fa_w11(ws, week, filepath) -> list[dict]:
    """FA W11: Driver | Code | Active Between | Days | Net Pay | From Last Period | Pay This Period"""
    rows_out = []
    all_rows = list(ws.iter_rows(values_only=True))

    for i, row in enumerate(all_rows):
        row_s = [str(c).strip().lower() if c else "" for c in row]
        if "driver" in row_s or "person" in row_s:
            col_name = next((j for j, h in enumerate(row_s) if h in ("driver", "person")), 0)
            col_code = next((j for j, h in enumerate(row_s) if h == "code"), 1)
            col_net = next((j for j, h in enumerate(row_s) if "net pay" in h), 4)
            col_from = next((j for j, h in enumerate(row_s) if "from last" in h), 5)
            col_pay_this = next((j for j, h in enumerate(row_s) if "pay this period" in h), 6)

            for dr in all_rows[i + 1:]:
                # W11 format: stop at TOTALS row or accounting section
                name0 = str(dr[col_name]).strip().lower() if len(dr) > col_name and dr[col_name] else ""
                if name0 in SKIP_KEYWORDS or name0 in HARD_STOP_KEYWORDS:
                    break
                if not is_data_row(dr, col_name, col_code):
                    continue
                name = str(dr[col_name]).strip()
                code = normalize_code(dr[col_code] if len(dr) > col_code else None)
                net_pay = to_float(dr[col_net] if len(dr) > col_net else None)
                pay_this = dr[col_pay_this] if len(dr) > col_pay_this else None

                # "Withheld" means $0 paid this period
                withheld = is_withheld(pay_this)
                paid_amount = 0.0 if withheld else to_float(pay_this)

                if net_pay is None and paid_amount is None:
                    continue

                rows_out.append({
                    "driver_name_in_file": name,
                    "paycheck_code": code,
                    "week": week,
                    "llc": "FA",
                    "payroll_amount": net_pay,
                    "unpaid_pending": None,
                    "to_paid": paid_amount,
                    "total": net_pay,
                    "source_file": str(filepath),
                })
            break

    return rows_out


def parse_fa_new(ws, week, filepath) -> list[dict]:
    """FA W12-W14: DriverName | PayCode | Rides | Miles | PartnerPays | DriverPay | Deduction | Withheld | CarriedOver | PaidThisPeriod"""
    rows_out = []
    all_rows = list(ws.iter_rows(values_only=True))

    for i, row in enumerate(all_rows):
        row_s = [str(c).strip().lower() if c else "" for c in row]
        if "driver name" in row_s or "driver" in row_s or "person" in row_s:
            col_name = next((j for j, h in enumerate(row_s) if h in ("driver name", "driver", "person")), 0)
            col_code = next((j for j, h in enumerate(row_s) if "code" in h), 1)
            col_driver_pay = next((j for j, h in enumerate(row_s) if h == "driver pay"), 5)
            col_withheld = next((j for j, h in enumerate(row_s) if "withheld" in h), 7)
            col_carried = next((j for j, h in enumerate(row_s) if "carried" in h), 8)
            col_paid_this = next((j for j, h in enumerate(row_s) if "paid this" in h), 9)

            for dr in all_rows[i + 1:]:
                # FA new format: stop at any total/accounting row
                name0 = str(dr[col_name]).strip().lower() if len(dr) > col_name and dr[col_name] else ""
                if name0 in SKIP_KEYWORDS or name0 in HARD_STOP_KEYWORDS:
                    break
                if not is_data_row(dr, col_name, col_code):
                    continue
                name = str(dr[col_name]).strip()
                code = normalize_code(dr[col_code] if len(dr) > col_code else None)
                driver_pay = to_float(dr[col_driver_pay] if len(dr) > col_driver_pay else None)
                withheld_flag = dr[col_withheld] if len(dr) > col_withheld else None
                carried = to_float(dr[col_carried] if len(dr) > col_carried else None)
                paid_this = to_float(dr[col_paid_this] if len(dr) > col_paid_this else None)

                if driver_pay is None:
                    continue

                rows_out.append({
                    "driver_name_in_file": name,
                    "paycheck_code": code,
                    "week": week,
                    "llc": "FA",
                    "payroll_amount": driver_pay,
                    "unpaid_pending": carried if is_withheld(withheld_flag) else None,
                    "to_paid": paid_this,
                    "total": driver_pay,
                    "source_file": str(filepath),
                })
            break

    return rows_out


def parse_fa_week(week: int, filepath: Path) -> tuple[list[dict], Optional[str]]:
    """Parse one FA week file. Returns (rows, error_msg)."""
    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
        sheet_name, family = _fa_find_payroll_sheet(wb)
        if not sheet_name:
            wb.close()
            return [], f"No recognizable payroll tab in {filepath.name}"
        ws = wb[sheet_name]
        if family == "old":
            rows = parse_fa_old(ws, week, filepath)
        elif family == "w11":
            rows = parse_fa_w11(ws, week, filepath)
        else:  # new
            rows = parse_fa_new(ws, week, filepath)
        wb.close()
        return rows, None
    except Exception as e:
        return [], f"{type(e).__name__}: {e}\n{traceback.format_exc()}"


# ---------------------------------------------------------------------------
# Maz parsers
# ---------------------------------------------------------------------------

def _maz_find_payroll_sheet(wb):
    """Return (sheet_name, family).
    Families: 'old' (W1-W8 Payroll), 'master' (W9-W11 Master_Payroll),
              'new' (W12/W14 Payroll Summary), 'w13' (two Table 1 tabs).
    """
    names = wb.sheetnames
    normalized = {s.strip().lower().replace("_", " "): s for s in names}

    if "payroll summary" in normalized:
        return normalized["payroll summary"], "new"
    if "master payroll" in normalized or "master_payroll" in normalized:
        key = "master payroll" if "master payroll" in normalized else "master_payroll"
        return normalized.get(key, normalized.get("master_payroll")), "master"
    if "payroll" in normalized:
        return normalized["payroll"], "old"
    # W13: two Table 1 tabs — detect by absence of summary tab + presence of "table 1 (2)"
    table_tabs = [s for s in names if s.lower().startswith("table 1")]
    if len(table_tabs) >= 2:
        return None, "w13"

    return None, None


def parse_maz_old(ws, week, filepath) -> list[dict]:
    """Maz W1-W8: Person(0) | ?(1) | Code(2) | NetPay(3) | Payroll(4) | ... | Unpaid(8) | ... | Total(10)"""
    rows_out = []
    all_rows = list(ws.iter_rows(values_only=True))

    col_person = 0
    col_code = 2
    col_payroll = 4
    col_unpaid = 8
    col_total = 10

    # Find header to confirm column positions
    for i, row in enumerate(all_rows):
        row_s = [str(c).strip().lower() if c else "" for c in row]
        if "person" in row_s:
            col_person = row_s.index("person")
            if "code" in row_s:
                col_code = row_s.index("code")
            if "payroll" in row_s:
                col_payroll = row_s.index("payroll")
            for j, h in enumerate(row_s):
                if "unpaid" in h or "pending" in h:
                    col_unpaid = j
                    break
            for j, h in enumerate(row_s):
                if h in ("total", "total "):
                    col_total = j
                    break
            # Parse from next row
            for dr in all_rows[i + 1:]:
                if is_stop_row(dr, col_person):
                    break
                if not is_data_row(dr, col_person, col_code):
                    continue
                name = str(dr[col_person]).strip()
                code = normalize_code(dr[col_code] if len(dr) > col_code else None)
                payroll = to_float(dr[col_payroll] if len(dr) > col_payroll else None)
                unpaid = to_float(dr[col_unpaid] if len(dr) > col_unpaid else None)
                total = to_float(dr[col_total] if len(dr) > col_total else None)

                if payroll is None and total is None:
                    continue

                rows_out.append({
                    "driver_name_in_file": name,
                    "paycheck_code_maz": code,
                    "week": week,
                    "llc": "Maz",
                    "payroll_amount": payroll,
                    "unpaid_pending": unpaid,
                    "to_paid": None,
                    "total": total,
                    "source_file": str(filepath),
                })
            break

    return rows_out


def parse_maz_master(ws, week, filepath) -> list[dict]:
    """Maz W9-W11: Person | Code | ActiveBetween | Days | Runs | Miles | Gross | UnPaid | ToBePaid | NetPay"""
    rows_out = []
    all_rows = list(ws.iter_rows(values_only=True))

    for i, row in enumerate(all_rows):
        row_s = [str(c).strip().lower() if c else "" for c in row]
        if "person" in row_s:
            col_name = row_s.index("person")
            col_code = row_s.index("code") if "code" in row_s else 1
            col_gross = next((j for j, h in enumerate(row_s) if h == "gross"), 6)
            col_unpaid = next((j for j, h in enumerate(row_s) if "unpaid" in h or "un paid" in h), 7)
            col_to_be = next((j for j, h in enumerate(row_s) if "to be paid" in h), 8)
            col_net = next((j for j, h in enumerate(row_s) if "net pay" in h), 9)

            for dr in all_rows[i + 1:]:
                name0 = str(dr[col_name]).strip().lower() if len(dr) > col_name and dr[col_name] else ""
                if name0 in SKIP_KEYWORDS or name0 in HARD_STOP_KEYWORDS:
                    break
                if not is_data_row(dr, col_name, col_code):
                    continue
                name = str(dr[col_name]).strip()
                code = normalize_code(dr[col_code] if len(dr) > col_code else None)
                gross = to_float(dr[col_gross] if len(dr) > col_gross else None)
                unpaid = to_float(dr[col_unpaid] if len(dr) > col_unpaid else None)
                to_be = to_float(dr[col_to_be] if len(dr) > col_to_be else None)
                net = to_float(dr[col_net] if len(dr) > col_net else None)

                if gross is None and net is None:
                    continue

                rows_out.append({
                    "driver_name_in_file": name,
                    "paycheck_code_maz": code,
                    "week": week,
                    "llc": "Maz",
                    "payroll_amount": gross,
                    "unpaid_pending": unpaid,
                    "to_paid": to_be,
                    "total": net,
                    "source_file": str(filepath),
                })
            break

    return rows_out


def parse_maz_new(ws, week, filepath) -> list[dict]:
    """Maz W12/W14 Payroll Summary — same structure as FA new."""
    rows_out = []
    all_rows = list(ws.iter_rows(values_only=True))

    for i, row in enumerate(all_rows):
        row_s = [str(c).strip().lower() if c else "" for c in row]
        if "driver name" in row_s or "driver" in row_s or "person" in row_s:
            col_name = next((j for j, h in enumerate(row_s) if h in ("driver name", "driver", "person")), 0)
            col_code = next((j for j, h in enumerate(row_s) if "code" in h), 1)
            col_driver_pay = next((j for j, h in enumerate(row_s) if h == "driver pay"), 5)
            col_withheld = next((j for j, h in enumerate(row_s) if "withheld" in h), 7)
            col_carried = next((j for j, h in enumerate(row_s) if "carried" in h), 8)
            col_paid_this = next((j for j, h in enumerate(row_s) if "paid this" in h), 9)

            for dr in all_rows[i + 1:]:
                name0 = str(dr[col_name]).strip().lower() if len(dr) > col_name and dr[col_name] else ""
                if name0 in SKIP_KEYWORDS or name0 in HARD_STOP_KEYWORDS:
                    break
                if not is_data_row(dr, col_name, col_code):
                    continue
                name = str(dr[col_name]).strip()
                code = normalize_code(dr[col_code] if len(dr) > col_code else None)
                driver_pay = to_float(dr[col_driver_pay] if len(dr) > col_driver_pay else None)
                withheld_flag = dr[col_withheld] if len(dr) > col_withheld else None
                carried = to_float(dr[col_carried] if len(dr) > col_carried else None)
                paid_this = to_float(dr[col_paid_this] if len(dr) > col_paid_this else None)

                if driver_pay is None:
                    continue

                rows_out.append({
                    "driver_name_in_file": name,
                    "paycheck_code_maz": code,
                    "week": week,
                    "llc": "Maz",
                    "payroll_amount": driver_pay,
                    "unpaid_pending": carried if is_withheld(withheld_flag) else None,
                    "to_paid": paid_this,
                    "total": driver_pay,
                    "source_file": str(filepath),
                })
            break

    return rows_out


def parse_maz_w13(wb, week, filepath) -> list[dict]:
    """Maz W13: two Table 1 tabs. Each has Person | Code | ... | Net Pay (col 13).
    Combine both sub-tables per driver (sum Net Pay)."""
    rows_out = []
    table_tabs = sorted([s for s in wb.sheetnames if s.lower().startswith("table 1")])

    # Accumulate per driver
    driver_sums: dict[str, dict] = {}

    for tab in table_tabs:
        ws = wb[tab]
        all_rows = list(ws.iter_rows(values_only=True))
        col_name = 0
        col_code = 1
        col_net = 13

        for i, row in enumerate(all_rows):
            row_s = [str(c).strip().lower() if c else "" for c in row]
            if "person" in row_s:
                col_name = row_s.index("person")
                if "code" in row_s:
                    col_code = row_s.index("code")
                # Net Pay is col 13 (0-indexed) in observed data
                for j, h in enumerate(row_s):
                    if "net pay" in h:
                        col_net = j
                        break
                for dr in all_rows[i + 1:]:
                    if not is_data_row(dr, col_name, col_code):
                        continue
                    name = str(dr[col_name]).strip()
                    # Skip subtotal / details
                    if name.lower() in ("subtotal", "details", "total"):
                        continue
                    code = normalize_code(dr[col_code] if len(dr) > col_code else None)
                    net = to_float(dr[col_net] if len(dr) > col_net else None)
                    if net is None:
                        continue
                    key = name.lower()
                    if key not in driver_sums:
                        driver_sums[key] = {
                            "driver_name_in_file": name,
                            "paycheck_code_maz": code,
                            "net_sum": 0.0,
                        }
                    driver_sums[key]["net_sum"] += net
                    # Prefer non-None code
                    if code and not driver_sums[key]["paycheck_code_maz"]:
                        driver_sums[key]["paycheck_code_maz"] = code
                break

    for d in driver_sums.values():
        rows_out.append({
            "driver_name_in_file": d["driver_name_in_file"],
            "paycheck_code_maz": d["paycheck_code_maz"],
            "week": week,
            "llc": "Maz",
            "payroll_amount": round(d["net_sum"], 2),
            "unpaid_pending": None,
            "to_paid": None,
            "total": round(d["net_sum"], 2),
            "source_file": str(filepath),
        })

    return rows_out


def parse_maz_week(week: int, filepath: Path) -> tuple[list[dict], Optional[str]]:
    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
        sheet_name, family = _maz_find_payroll_sheet(wb)

        if family is None:
            wb.close()
            return [], f"No recognizable payroll tab in {filepath.name}"

        if family == "w13":
            rows = parse_maz_w13(wb, week, filepath)
            wb.close()
            return rows, None

        ws = wb[sheet_name]
        if family == "old":
            rows = parse_maz_old(ws, week, filepath)
        elif family == "master":
            rows = parse_maz_master(ws, week, filepath)
        else:
            rows = parse_maz_new(ws, week, filepath)

        wb.close()
        return rows, None
    except Exception as e:
        return [], f"{type(e).__name__}: {e}\n{traceback.format_exc()}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Fetching person map from DB...")
    fa_map, maz_map = fetch_person_map()
    print(f"  FA map: {len(fa_map)} records, Maz map: {len(maz_map)} records")

    all_rows = []
    skipped = []
    parsed_ok = 0

    fa_weeks_ok = []
    maz_weeks_ok = []

    # --- FA ---
    for week in WEEKS:
        week_dir = FA_BASE / f"Week {week}"
        filepath = find_xlsx(week_dir)
        if not filepath:
            skipped.append({"llc": "FA", "week": week, "path": str(week_dir), "reason": "No .xlsx file found"})
            continue
        rows, err = parse_fa_week(week, filepath)
        if err:
            skipped.append({"llc": "FA", "week": week, "path": str(filepath), "reason": err})
            continue
        if not rows:
            skipped.append({"llc": "FA", "week": week, "path": str(filepath), "reason": "Parser returned 0 rows"})
            continue
        # Enrich with paycheck_code_maz from DB
        for r in rows:
            r.setdefault("paycheck_code_maz", None)
            if r.get("paycheck_code") and r["paycheck_code"] in fa_map:
                db_rec = fa_map[r["paycheck_code"]]
                if not r["paycheck_code_maz"]:
                    r["paycheck_code_maz"] = db_rec.get("paycheck_code_maz")
        all_rows.extend(rows)
        parsed_ok += 1
        fa_weeks_ok.append(week)
        print(f"  FA W{week}: {len(rows)} drivers parsed from {filepath.name}")

    # --- Maz ---
    for week in WEEKS:
        week_dir = MAZ_BASE / f"Week {week}"
        filepath = find_xlsx(week_dir)
        if not filepath:
            skipped.append({"llc": "Maz", "week": week, "path": str(week_dir), "reason": "No .xlsx file found"})
            continue
        rows, err = parse_maz_week(week, filepath)
        if err:
            skipped.append({"llc": "Maz", "week": week, "path": str(filepath), "reason": err})
            continue
        if not rows:
            skipped.append({"llc": "Maz", "week": week, "path": str(filepath), "reason": "Parser returned 0 rows"})
            continue
        # Enrich with paycheck_code from DB
        for r in rows:
            r.setdefault("paycheck_code", None)
            if r.get("paycheck_code_maz") and r["paycheck_code_maz"] in maz_map:
                db_rec = maz_map[r["paycheck_code_maz"]]
                if not r["paycheck_code"]:
                    r["paycheck_code"] = db_rec.get("paycheck_code")
        all_rows.extend(rows)
        parsed_ok += 1
        maz_weeks_ok.append(week)
        print(f"  Maz W{week}: {len(rows)} drivers parsed from {filepath.name}")

    print(f"\nTotal rows: {len(all_rows)}, files parsed: {parsed_ok}, skipped: {len(skipped)}")

    # --- Write CSV ---
    fieldnames = [
        "paycheck_code", "paycheck_code_maz", "driver_name_in_file",
        "week", "llc", "payroll_amount", "unpaid_pending", "to_paid",
        "total", "source_file",
    ]
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in all_rows:
            writer.writerow(row)
    print(f"CSV written: {OUT_CSV}")

    # --- Build summary ---
    build_summary(all_rows, skipped, fa_weeks_ok, maz_weeks_ok, parsed_ok, fa_map, maz_map)
    print(f"Summary written: {OUT_MD}")


# ---------------------------------------------------------------------------
# Summary builder
# ---------------------------------------------------------------------------

def build_summary(all_rows, skipped, fa_weeks_ok, maz_weeks_ok, parsed_ok, fa_map, maz_map):
    lines = []
    lines.append("# Z-Pay Truth Ledger — W1 through W14 Summary")
    lines.append("")

    # Coverage
    def week_range_str(weeks):
        if not weeks:
            return "none"
        weeks_s = sorted(weeks)
        return f"W{weeks_s[0]}–W{weeks_s[-1]} ({len(weeks_s)} weeks)"

    lines.append("## Coverage")
    lines.append(f"- FA: {week_range_str(fa_weeks_ok)}")
    lines.append(f"- Maz: {week_range_str(maz_weeks_ok)}")
    lines.append("")

    # Parse stats
    lines.append("## Parse Stats")
    lines.append(f"- Files parsed successfully: {parsed_ok}")
    lines.append(f"- Files skipped: {len(skipped)}")
    if skipped:
        lines.append("")
        lines.append("### Skipped Files")
        for s in skipped:
            lines.append(f"- {s['llc']} W{s['week']}: `{s['path']}` — {s['reason'][:120]}")
    lines.append("")

    # Totals
    fa_rows = [r for r in all_rows if r["llc"] == "FA"]
    maz_rows = [r for r in all_rows if r["llc"] == "Maz"]

    def safe_sum(rows, field):
        return sum(r[field] for r in rows if r.get(field) is not None)

    fa_total = safe_sum(fa_rows, "payroll_amount")
    maz_total = safe_sum(maz_rows, "payroll_amount")
    overall_total = fa_total + maz_total

    lines.append("## Dollar Totals (payroll_amount column)")
    lines.append(f"- FA total:      ${fa_total:,.2f}")
    lines.append(f"- Maz total:     ${maz_total:,.2f}")
    lines.append(f"- Overall total: ${overall_total:,.2f}")
    lines.append("")

    # Driver count grid
    lines.append("## Driver Count per (Week, LLC)")
    lines.append("")
    lines.append("| Week | FA drivers | Maz drivers |")
    lines.append("|------|-----------|------------|")
    for w in WEEKS:
        fa_cnt = len([r for r in fa_rows if r["week"] == w])
        maz_cnt = len([r for r in maz_rows if r["week"] == w])
        lines.append(f"| W{w:02d}  | {fa_cnt:9d} | {maz_cnt:11d} |")
    lines.append("")

    # Nuraynie deep-dive
    lines.append("## Nuraynie Mohammed — Per-Week Detail")
    lines.append("")
    lines.append("Nuraynie has no paycheck_code (never enrolled in Paychex). Code=None in all files.")
    lines.append("Reconnaissance anchor: $1,198 across W10–W14. DB pre-wipe anchor: $1,595 (inflated suspicion).")
    lines.append("")

    nura_rows = [
        r for r in fa_rows
        if "nuraynie" in r["driver_name_in_file"].lower()
    ]

    # Deduplicate: some weeks have the driver appear multiple times (summary + sub-table).
    # Keep only the main data row per week (first occurrence with payroll_amount > 0).
    nura_by_week: dict[int, dict] = {}
    for r in nura_rows:
        w = r["week"]
        amount = r.get("payroll_amount") or 0
        if w not in nura_by_week:
            nura_by_week[w] = r
        else:
            # Prefer the row with the higher amount (avoids 0/None sub-rows)
            existing = nura_by_week[w].get("payroll_amount") or 0
            if amount > existing:
                nura_by_week[w] = r

    lines.append("| Week | payroll_amount | to_paid | notes |")
    lines.append("|------|---------------|---------|-------|")

    nura_total = 0.0
    for w in WEEKS:
        if w in nura_by_week:
            r = nura_by_week[w]
            amt = r.get("payroll_amount")
            to_paid = r.get("to_paid")
            amt_s = f"${amt:,.2f}" if amt is not None else "—"
            to_paid_s = f"${to_paid:,.2f}" if to_paid is not None else "—"
            nura_total += amt or 0
            lines.append(f"| W{w:02d}  | {amt_s} | {to_paid_s} | |")
        else:
            lines.append(f"| W{w:02d}  | not present | — | |")

    lines.append(f"| **TOTAL** | **${nura_total:,.2f}** | | |")
    lines.append("")

    # Drift callout
    prior_anchor = 1198.0  # reconnaissance figure
    db_anchor = 1595.0     # DB figure (pre-wipe, suspected inflation)

    lines.append("### Drift Analysis")
    drift_vs_recon = nura_total - prior_anchor
    drift_vs_db = nura_total - db_anchor

    lines.append(f"- Ledger total: **${nura_total:,.2f}**")
    lines.append(f"- Reconnaissance anchor (W10–W14): **${prior_anchor:,.2f}** — drift: **${drift_vs_recon:+,.2f}**")
    lines.append(f"- DB pre-wipe anchor: **${db_anchor:,.2f}** — drift: **${drift_vs_db:+,.2f}**")
    lines.append("")

    if abs(drift_vs_recon) > 1.0:
        lines.append("> **NOTE**: Ledger total differs from reconnaissance anchor. See per-week rows above.")
    if drift_vs_db < -10.0:
        lines.append("> **PRE-WIPE DB MAY HAVE INFLATED**: DB anchor ($1,595) is higher than ledger total by ${:.2f}. The pre-wipe DB likely carried ghost balances or double-counted Nuraynie's held amounts. Ledger total from mom's files is the ground truth.".format(abs(drift_vs_db)))
    elif drift_vs_db > 10.0:
        lines.append("> **LEDGER HIGHER THAN DB ANCHOR**: Investigate — ledger shows ${:.2f} more than pre-wipe DB. May indicate weeks not yet processed in DB at wipe time.".format(drift_vs_db))
    lines.append("")

    # Final notes
    lines.append("## Notes")
    lines.append("- FA W1-W10: parsed from 'Payroll ' tab (trailing-space variant). Payroll_amount = gross Payroll column.")
    lines.append("- FA W11: 'Payroll Summary' tab with Net Pay / Pay This Period columns. payroll_amount = Net Pay.")
    lines.append("- FA W12-W14: richer format with Rides/Miles/DriverPay/PaidThisPeriod. payroll_amount = Driver Pay.")
    lines.append("- FA W14 has two 'Payroll Summary' tabs: used the first one (FirstAlt period Apr 04–10).")
    lines.append("- Maz W1-W8: 'Payroll' tab. payroll_amount = Payroll column.")
    lines.append("- Maz W9-W11: 'Master_Payroll' tab. payroll_amount = Gross column.")
    lines.append("- Maz W12/W14: 'Payroll Summary' tab. payroll_amount = Driver Pay.")
    lines.append("- Maz W13: no summary tab. Two 'Table 1' sub-tables (ED raw trip data). payroll_amount = sum of Net Pay across both sub-tables per driver.")
    lines.append("- Rows with no paycheck_code (None) appear for drivers not yet enrolled in Paychex at time of that week's file.")
    lines.append(f"- DB enrichment: FA map {len(fa_map)} records, Maz map {len(maz_map)} records.")

    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    main()
